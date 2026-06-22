from __future__ import annotations

import datetime
import hashlib
import json
import math
import os
import re
import secrets
import sqlite3
import threading
import uuid
from functools import wraps
from io import BytesIO
from typing import Any
from urllib import parse as urlparse
from urllib import request as urlrequest

from flask import (
    Flask,
    abort,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook, load_workbook
from itsdangerous import BadSignature, BadTimeSignature, SignatureExpired, URLSafeTimedSerializer
from werkzeug.security import check_password_hash, generate_password_hash

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")
DATA_DIR = os.path.join(BASE_DIR, "data")
DATABASE = os.path.join(DATA_DIR, "dramas.db")
REMOTE_UPLOAD_DIR = os.path.join(DATA_DIR, "remote_uploads")
ICON_DIR = os.path.join(STATIC_DIR, "icons")
FAVICON_PATH = os.path.join(ICON_DIR, "app-icon.ico")
APPLE_TOUCH_ICON_PATH = os.path.join(ICON_DIR, "app-icon.png")

ALLOWED_FLAGS = {"是", "否"}
SORTABLE_FIELDS = {
    "date",
    "original_name",
    "new_name",
    "episodes",
    "duration",
    "review_passed",
    "uploaded",
    "uploader",
    "company",
    "created_at",
}
DEFAULT_SORT_FIELD = "date"
DEFAULT_SORT_DIR = "desc"


def _int_env(name: str, default: int) -> int:
    try:
        return int(os.environ.get(name, "") or default)
    except (TypeError, ValueError):
        return default


LICENSE_STATUS_VALUES = {"active", "disabled", "expired"}
LICENSE_EDITION_VALUES = {"basic", "pro", "enterprise"}
LICENSE_LIST_SORTABLE_FIELDS = {
    "license_key": "license_key",
    "licensee": "COALESCE(licensee, '')",
    "edition": "edition",
    "max_activations": "max_activations",
    "active_activations": "(SELECT COUNT(*) FROM license_activations la WHERE la.license_id = licenses.id AND (la.revoked_at IS NULL OR la.revoked_at = ''))",
    "expires_at": "COALESCE(expires_at, '9999-12-31')",
    "last_verified_at": "(SELECT COALESCE(MAX(la.last_verified_at), '') FROM license_activations la WHERE la.license_id = licenses.id)",
    "status": "status",
    "created_at": "created_at",
    "updated_at": "updated_at",
}
LICENSE_LIST_DEFAULT_SORT_FIELD = "created_at"
LICENSE_LIST_DEFAULT_SORT_DIR = "desc"
LICENSE_TOKEN_SALT = "desktop-license"
LICENSE_TOKEN_MAX_AGE_SECONDS = 60 * 60 * 24 * 30
ACCOUNT_TOKEN_SALT = "desktop-account"
TT_ACCOUNT_TOKEN_SALT = "tiktok-account"
ACCOUNT_TOKEN_MAX_AGE_SECONDS = _int_env("ACCOUNT_TOKEN_MAX_AGE_SECONDS", 60 * 60 * 24)
ACCOUNT_OFFLINE_GRACE_HOURS = _int_env("ACCOUNT_OFFLINE_GRACE_HOURS", 72)
ACCOUNT_DEFAULT_MAX_DEVICES = 1
USERNAME_RE = re.compile(r"^(?:[A-Za-z0-9_]{2,30}|[^@\s]+@[^@\s]+\.[^@\s]+)$")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
REMOTE_MESSAGE_STATUS_VALUES = {"pending", "sent", "running", "success", "failed", "canceled", "stopped"}
REMOTE_MESSAGE_TYPE_VALUES = {"text", "command", "image", "status", "log"}
REMOTE_SENDER_TYPE_VALUES = {"user", "client", "system"}
REMOTE_COMMAND_IMPORT_DRAMA_TITLES = "import_drama_titles"
REMOTE_COMMAND_KUAISHOU_UPLOAD_SERIES = "ks_upload_series"
REMOTE_COMMAND_KUAISHOU_START_QUEUE = "ks_start_queue"
REMOTE_COMMAND_KUAISHOU_STOP_QUEUE = "ks_stop_queue"
REMOTE_COMMAND_KUAISHOU_QUERY_STATUS = "ks_query_status"
REMOTE_COMMAND_ALLOWED_COMMANDS = {
    REMOTE_COMMAND_IMPORT_DRAMA_TITLES,
    REMOTE_COMMAND_KUAISHOU_UPLOAD_SERIES,
    REMOTE_COMMAND_KUAISHOU_START_QUEUE,
    REMOTE_COMMAND_KUAISHOU_STOP_QUEUE,
    REMOTE_COMMAND_KUAISHOU_QUERY_STATUS,
}
REMOTE_IMPORT_DRAMA_ALLOWED_STEPS = {
    "download",
    "rewrite_info",
    "material_transcode",
    "material_auto_repair",
    "auto_fill_info",
    "generate_poster",
    "generate_materials",
    "material_validate",
    "upload_series",
    "publish_materials",
}
REMOTE_IMPORT_DRAMA_ALLOWED_ERROR_STRATEGIES = {"skip", "stop"}
WEIXIN_API_BASE = "https://api.weixin.qq.com"
MINIDRAMA_TOKEN_REFRESH_MARGIN_SECONDS = 300
MINIDRAMA_TOKEN_REFRESH_LOCK_SECONDS = 60
MINIDRAMA_APP_ID_SETTING_KEY = "minidrama_app_id"
MINIDRAMA_APP_SECRET_SETTING_KEY = "minidrama_app_secret"
KUAISHOU_API_BASE = "https://ad.e.kuaishou.com"
KUAISHOU_TOKEN_REFRESH_MARGIN_SECONDS = 2 * 60 * 60
KUAISHOU_ACCESS_TOKEN_RENEW_MARGIN_SECONDS = 6 * 60 * 60
KUAISHOU_REFRESH_TOKEN_RENEW_MARGIN_SECONDS = 7 * 24 * 60 * 60
KUAISHOU_TOKEN_REFRESH_LOCK_SECONDS = 60
KUAISHOU_TOKEN_REFRESH_SCHEDULER_INTERVAL_SECONDS = _int_env('KUAISHOU_TOKEN_REFRESH_SCHEDULER_INTERVAL_SECONDS', 60 * 60)
KUAISHOU_TOKEN_REFRESH_STARTUP_DELAY_SECONDS = _int_env('KUAISHOU_TOKEN_REFRESH_STARTUP_DELAY_SECONDS', 10)
UPLOAD_RECORD_PLATFORMS = {"video_channel", "miniprogram", "kuaishou"}
UPLOAD_RECORD_PLATFORM_LABELS = {
    "video_channel": "微信视频号",
    "miniprogram": "微信小程序",
    "kuaishou": "快手",
}
PLATFORM_DRAMA_SORTABLE_FIELDS = {
    "date": "COALESCE(ur.date, substr(ur.record_time, 1, 10), d.date, '')",
    "record_time": "COALESCE(ur.record_time, ur.created_at, '')",
    "original_name": "COALESCE(ur.original_name, d.original_name, '')",
    "new_name": "COALESCE(ur.new_name, d.new_name, '')",
    "episodes": "COALESCE(ur.episodes, d.episodes, 0)",
    "uploaded": "COALESCE(ur.upload_status, '')",
    "uploader": "COALESCE(ur.uploader_display, ur.owner_username, '')",
    "company": "COALESCE(d.company, '')",
    "owner_username": "COALESCE(ur.owner_username, '')",
}

HEADER_MAP = {
    "日期": "date",
    "原剧名": "original_name",
    "新剧名": "new_name",
    "集数": "episodes",
    "时间(分钟)": "duration",
    "时间（分钟）": "duration",
    "是否审核通过": "review_passed",
    "是否上传": "uploaded",
    "素材": "materials",
    "推广语": "promo_text",
    "简介": "description",
    "公司": "company",
    "备注一": "remark1",
    "备注二": "remark2",
    "备注三": "remark3",
}
EXPORT_HEADERS = [
    "日期",
    "原剧名",
    "新剧名",
    "集数",
    "时间(分钟)",
    "是否审核通过",
    "是否上传",
    "素材",
    "推广语",
    "简介",
    "公司",
    "上传者",
]
LICENSE_EXPORT_HEADERS = [
    "激活码",
    "掩码",
    "授权对象",
    "版本",
    "状态",
    "最大设备数",
    "当前绑定设备数",
    "累计绑定记录数",
    "到期时间",
    "最近校验",
    "备注",
    "创建时间",
    "更新时间",
    "删除时间",
]

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", uuid.uuid4().hex)
app.config["JSON_AS_ASCII"] = False
app.config["JSON_SORT_KEYS"] = False
app.config["LICENSE_SIGNING_KEY"] = os.environ.get(
    "LICENSE_SIGNING_KEY",
    app.config["SECRET_KEY"],
)


@app.route("/favicon.ico")
def favicon():
    return send_file(FAVICON_PATH, mimetype="image/vnd.microsoft.icon", max_age=86400)


@app.route("/apple-touch-icon.png")
def apple_touch_icon():
    return send_file(APPLE_TOUCH_ICON_PATH, mimetype="image/png", max_age=86400)

# 红果短剧链路(数据API + 密钥管理)已隐藏并暂停注册。
# 如需恢复红果短剧/红果密钥页面和相关 API，取消下面注释即可。
# try:
#     from hongguo_bp import hongguo_bp
#     app.register_blueprint(hongguo_bp)
# except Exception as _hg_exc:  # noqa: BLE001
#     print(f"[warn] 红果 Blueprint 未加载: {_hg_exc}")


def ensure_data_dir() -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(REMOTE_UPLOAD_DIR, exist_ok=True)


def get_db() -> sqlite3.Connection:
    if "db" not in g:
        ensure_data_dir()
        conn = sqlite3.connect(DATABASE, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON;")
        conn.execute("PRAGMA journal_mode = WAL;")
        g.db = conn
    return g.db


def close_db(_: Exception | None = None) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


app.teardown_appcontext(close_db)


def init_db() -> None:
    ensure_data_dir()
    with app.app_context():
        db = get_db()
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                email TEXT,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'user',
                status TEXT NOT NULL DEFAULT 'active',
                max_devices INTEGER NOT NULL DEFAULT 1,
                edition TEXT NOT NULL DEFAULT 'pro',
                expires_at TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS tt_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                email TEXT,
                password_hash TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'active',
                max_devices INTEGER NOT NULL DEFAULT 1,
                edition TEXT NOT NULL DEFAULT 'pro',
                expires_at TEXT,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS dramas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT,
                original_name TEXT NOT NULL,
                new_name TEXT NOT NULL,
                episodes INTEGER,
                duration INTEGER,
                review_passed TEXT NOT NULL DEFAULT '否',
                uploaded TEXT NOT NULL DEFAULT '否',
                materials TEXT,
                promo_text TEXT,
                description TEXT,
                company TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(original_name, new_name)
            );

            CREATE TABLE IF NOT EXISTS licenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                license_key TEXT UNIQUE NOT NULL,
                license_key_masked TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'active',
                edition TEXT NOT NULL DEFAULT 'pro',
                licensee TEXT,
                max_activations INTEGER NOT NULL DEFAULT 1,
                expires_at TEXT,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS license_activations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                license_id INTEGER NOT NULL,
                machine_id TEXT NOT NULL,
                app_name TEXT,
                app_version TEXT,
                token_hash TEXT NOT NULL,
                activated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                last_verified_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                revoked_at TEXT,
                FOREIGN KEY (license_id) REFERENCES licenses(id) ON DELETE CASCADE,
                UNIQUE(license_id, machine_id)
            );

            CREATE TABLE IF NOT EXISTS user_devices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                machine_id TEXT NOT NULL,
                device_name TEXT,
                app_name TEXT,
                app_version TEXT,
                token_hash TEXT NOT NULL,
                logged_in_at TEXT,
                last_verified_at TEXT,
                revoked_at TEXT,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                UNIQUE(user_id, machine_id)
            );

            CREATE TABLE IF NOT EXISTS tt_user_devices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tt_user_id INTEGER NOT NULL,
                machine_id TEXT NOT NULL,
                device_name TEXT,
                app_name TEXT,
                app_version TEXT,
                token_hash TEXT NOT NULL,
                logged_in_at TEXT,
                last_verified_at TEXT,
                revoked_at TEXT,
                FOREIGN KEY (tt_user_id) REFERENCES tt_users(id) ON DELETE CASCADE,
                UNIQUE(tt_user_id, machine_id)
            );

            CREATE TABLE IF NOT EXISTS remote_clients (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                client_id TEXT UNIQUE NOT NULL,
                client_name TEXT NOT NULL,
                client_token_hash TEXT NOT NULL,
                owner_user_id INTEGER NOT NULL,
                machine_id TEXT,
                device_name TEXT,
                app_version TEXT,
                workspace_path TEXT,
                status TEXT NOT NULL DEFAULT 'offline',
                last_seen_at TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (owner_user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS remote_conversations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                remote_client_id INTEGER NOT NULL,
                owner_user_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'active',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (remote_client_id) REFERENCES remote_clients(id) ON DELETE CASCADE,
                FOREIGN KEY (owner_user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS remote_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                conversation_id INTEGER NOT NULL,
                sender_type TEXT NOT NULL,
                sender_user_id INTEGER,
                remote_client_id INTEGER,
                message_type TEXT NOT NULL,
                content_text TEXT,
                payload_json TEXT,
                status TEXT NOT NULL DEFAULT 'pending',
                result_json TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (conversation_id) REFERENCES remote_conversations(id) ON DELETE CASCADE,
                FOREIGN KEY (sender_user_id) REFERENCES users(id) ON DELETE SET NULL,
                FOREIGN KEY (remote_client_id) REFERENCES remote_clients(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS remote_attachments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                message_id INTEGER NOT NULL,
                file_type TEXT NOT NULL,
                original_name TEXT,
                stored_path TEXT NOT NULL,
                content_type TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (message_id) REFERENCES remote_messages(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS minidrama_token_cache (
                app_id TEXT PRIMARY KEY,
                access_token TEXT,
                expires_at INTEGER NOT NULL DEFAULT 0,
                refreshing_by TEXT,
                refreshing_until INTEGER NOT NULL DEFAULT 0,
                last_error TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS minidrama_apps (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                app_id TEXT UNIQUE NOT NULL,
                app_secret TEXT NOT NULL,
                name TEXT,
                is_default INTEGER NOT NULL DEFAULT 0,
                enabled INTEGER NOT NULL DEFAULT 1,
                created_at TEXT,
                updated_at TEXT,
                updated_by INTEGER,
                FOREIGN KEY (updated_by) REFERENCES users(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS kuaishou_token_cache (
                app_id TEXT PRIMARY KEY,
                access_token TEXT,
                expires_at INTEGER NOT NULL DEFAULT 0,
                refreshing_by TEXT,
                refreshing_until INTEGER NOT NULL DEFAULT 0,
                last_error TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS kuaishou_apps (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                app_id TEXT UNIQUE NOT NULL,
                app_secret TEXT NOT NULL,
                advertiser_id TEXT,
                name TEXT,
                access_token TEXT,
                refresh_token TEXT,
                access_token_expires_at INTEGER NOT NULL DEFAULT 0,
                refresh_token_expires_at INTEGER NOT NULL DEFAULT 0,
                is_default INTEGER NOT NULL DEFAULT 0,
                enabled INTEGER NOT NULL DEFAULT 1,
                created_at TEXT,
                updated_at TEXT,
                updated_by INTEGER,
                FOREIGN KEY (updated_by) REFERENCES users(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS app_settings (
                setting_key TEXT PRIMARY KEY,
                setting_value TEXT,
                is_secret INTEGER NOT NULL DEFAULT 0,
                updated_at TEXT,
                updated_by INTEGER
            );

            CREATE TABLE IF NOT EXISTS upload_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                owner_user_id INTEGER NOT NULL,
                owner_username TEXT,
                remote_client_id INTEGER,
                drama_id INTEGER,
                platform TEXT NOT NULL,
                platform_label TEXT,
                sync_key TEXT NOT NULL,
                record_time TEXT,
                date TEXT,
                upload_status TEXT,
                execution_mode TEXT,
                step_label TEXT,
                project_name TEXT,
                project_path TEXT,
                original_name TEXT,
                new_name TEXT,
                episodes INTEGER,
                video_file_count INTEGER,
                uploaded_video_count INTEGER,
                uploader_display TEXT,
                account_profile_id TEXT,
                account_profile_name TEXT,
                device_name TEXT,
                failure_reason TEXT,
                extra_info TEXT,
                series_id TEXT,
                mini_series_id TEXT,
                audit_status TEXT,
                selling_status TEXT,
                audit_reject_reason TEXT,
                audit_reject_detail TEXT,
                online_status TEXT,
                online_at TEXT,
                distribution_status TEXT,
                distribution_at TEXT,
                distribution_detail TEXT,
                submitted_at TEXT,
                raw_json TEXT,
                created_at TEXT,
                updated_at TEXT,
                FOREIGN KEY (owner_user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY (remote_client_id) REFERENCES remote_clients(id) ON DELETE SET NULL,
                FOREIGN KEY (drama_id) REFERENCES dramas(id) ON DELETE SET NULL,
                UNIQUE(owner_user_id, platform, sync_key)
            );

            CREATE TABLE IF NOT EXISTS drama_platform_status (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                drama_id INTEGER NOT NULL,
                platform TEXT NOT NULL,
                platform_label TEXT,
                latest_status TEXT,
                uploaded_video_count INTEGER,
                video_file_count INTEGER,
                last_record_id INTEGER,
                external_series_id TEXT,
                audit_status TEXT,
                selling_status TEXT,
                audit_reject_reason TEXT,
                online_status TEXT,
                distribution_status TEXT,
                updated_at TEXT,
                FOREIGN KEY (drama_id) REFERENCES dramas(id) ON DELETE CASCADE,
                FOREIGN KEY (last_record_id) REFERENCES upload_records(id) ON DELETE SET NULL,
                UNIQUE(drama_id, platform)
            );
            """
        )
        # Migrate: add source column if missing
        try:
            db.execute("ALTER TABLE dramas ADD COLUMN source TEXT DEFAULT NULL")
            db.commit()
        except Exception:
            pass  # column already exists
        for col_def in [
            "ALTER TABLE dramas ADD COLUMN uploader TEXT DEFAULT NULL",
            "ALTER TABLE dramas ADD COLUMN remark1 TEXT DEFAULT NULL",
            "ALTER TABLE dramas ADD COLUMN remark2 TEXT DEFAULT NULL",
            "ALTER TABLE dramas ADD COLUMN remark3 TEXT DEFAULT NULL",
            "ALTER TABLE licenses ADD COLUMN deleted_at TEXT DEFAULT NULL",
            "ALTER TABLE licenses ADD COLUMN deleted_by INTEGER DEFAULT NULL",
            "ALTER TABLE users ADD COLUMN email TEXT DEFAULT NULL",
            "ALTER TABLE users ADD COLUMN status TEXT NOT NULL DEFAULT 'active'",
            "ALTER TABLE users ADD COLUMN max_devices INTEGER NOT NULL DEFAULT 1",
            "ALTER TABLE users ADD COLUMN edition TEXT NOT NULL DEFAULT 'pro'",
            "ALTER TABLE users ADD COLUMN expires_at TEXT DEFAULT NULL",
            "ALTER TABLE users ADD COLUMN updated_at TEXT DEFAULT NULL",
            "ALTER TABLE upload_records ADD COLUMN audit_reject_reason TEXT DEFAULT NULL",
            "ALTER TABLE upload_records ADD COLUMN audit_reject_detail TEXT DEFAULT NULL",
            "ALTER TABLE upload_records ADD COLUMN online_status TEXT DEFAULT NULL",
            "ALTER TABLE upload_records ADD COLUMN online_at TEXT DEFAULT NULL",
            "ALTER TABLE upload_records ADD COLUMN distribution_status TEXT DEFAULT NULL",
            "ALTER TABLE upload_records ADD COLUMN distribution_at TEXT DEFAULT NULL",
            "ALTER TABLE upload_records ADD COLUMN distribution_detail TEXT DEFAULT NULL",
            "ALTER TABLE drama_platform_status ADD COLUMN audit_reject_reason TEXT DEFAULT NULL",
            "ALTER TABLE drama_platform_status ADD COLUMN online_status TEXT DEFAULT NULL",
            "ALTER TABLE drama_platform_status ADD COLUMN distribution_status TEXT DEFAULT NULL",
        ]:
            try:
                db.execute(col_def)
                db.commit()
            except Exception:
                pass
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_license_activations_license_id ON license_activations(license_id)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_license_activations_machine_id ON license_activations(machine_id)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_licenses_deleted_at ON licenses(deleted_at)"
        )
        db.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_users_email_unique ON users(lower(email)) WHERE email IS NOT NULL AND email != ''"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_user_devices_user_id ON user_devices(user_id)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_user_devices_machine_id ON user_devices(machine_id)"
        )
        db.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_tt_users_email_unique ON tt_users(lower(email)) WHERE email IS NOT NULL AND email != ''"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_tt_user_devices_user_id ON tt_user_devices(tt_user_id)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_tt_user_devices_machine_id ON tt_user_devices(machine_id)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_remote_clients_owner_user_id ON remote_clients(owner_user_id)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_remote_conversations_client_id ON remote_conversations(remote_client_id)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_remote_messages_conversation_id ON remote_messages(conversation_id)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_remote_messages_status ON remote_messages(status)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_minidrama_token_expires_at ON minidrama_token_cache(expires_at)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_minidrama_apps_default ON minidrama_apps(is_default, enabled)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_kuaishou_token_expires_at ON kuaishou_token_cache(expires_at)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_kuaishou_apps_default ON kuaishou_apps(is_default, enabled)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_dramas_created_at ON dramas(created_at)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_dramas_date ON dramas(date)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_dramas_company ON dramas(company)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_upload_records_owner_user_id ON upload_records(owner_user_id)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_upload_records_platform ON upload_records(platform)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_upload_records_record_time ON upload_records(record_time)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_upload_records_drama_id ON upload_records(drama_id)"
        )
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_drama_platform_status_drama_id ON drama_platform_status(drama_id)"
        )
        db.execute(
            """
            UPDATE users
            SET max_devices = ?
            WHERE max_devices IS NULL OR max_devices > ?
            """,
            (ACCOUNT_DEFAULT_MAX_DEVICES, ACCOUNT_DEFAULT_MAX_DEVICES),
        )
        migrate_legacy_minidrama_settings(db)
        seed_default_users(db)
        db.commit()


def seed_default_users(db: sqlite3.Connection) -> None:
    count = db.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if count:
        return
    users = [
        ("admin", generate_password_hash("admin123"), "admin"),
        ("user1", generate_password_hash("user123"), "user"),
    ]
    db.executemany(
        "INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)", users
    )

def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if "user_id" not in session:
            if request.path.startswith("/api/"):
                return jsonify({"error": "需要登录"}), 401
            return redirect(url_for("login", next=request.full_path))
        return view(*args, **kwargs)

    return wrapped


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if session.get("role") != "admin":
            if request.path.startswith("/api/"):
                return jsonify({"error": "权限不足"}), 403
            abort(403)
        return view(*args, **kwargs)

    return wrapped


def get_license_serializer() -> URLSafeTimedSerializer:
    return URLSafeTimedSerializer(app.config["LICENSE_SIGNING_KEY"])


def mask_license_key(license_key: str) -> str:
    value = str(license_key or "").strip()
    if len(value) <= 8:
        return value
    return f"{value[:4]}****{value[-4:]}"


def generate_license_key() -> str:
    parts = [
        "WXA",
        str(datetime.date.today().year),
        secrets.token_hex(2).upper(),
        secrets.token_hex(2).upper(),
        secrets.token_hex(2).upper(),
    ]
    return "-".join(parts)


def hash_token(token: str) -> str:
    return hashlib.sha256(str(token or "").encode("utf-8")).hexdigest()


def issue_license_token(*, license_row: sqlite3.Row, machine_id: str) -> str:
    serializer = get_license_serializer()
    payload = {
        "license_id": license_row["id"],
        "license_key": license_row["license_key"],
        "machine_id": machine_id,
        "edition": license_row["edition"],
        "expires_at": license_row["expires_at"],
        "issued_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
    }
    return serializer.dumps(payload, salt=LICENSE_TOKEN_SALT)


def verify_license_token(token: str) -> dict:
    serializer = get_license_serializer()
    try:
        return serializer.loads(
            token,
            salt=LICENSE_TOKEN_SALT,
            max_age=LICENSE_TOKEN_MAX_AGE_SECONDS,
        )
    except (BadSignature, BadTimeSignature, SignatureExpired):
        raise ValueError("授权 token 无效或已过期")


def normalize_email(value: str) -> str:
    return str(value or "").strip().lower()


def issue_account_token(*, user_row: sqlite3.Row, machine_id: str) -> str:
    serializer = get_license_serializer()
    payload = {
        "user_id": user_row["id"],
        "username": user_row["username"],
        "email": user_row["email"] or "",
        "machine_id": machine_id,
        "edition": user_row["edition"],
        "expires_at": user_row["expires_at"] or "",
        "issued_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
    }
    return serializer.dumps(payload, salt=ACCOUNT_TOKEN_SALT)


def verify_account_token(token: str) -> dict:
    serializer = get_license_serializer()
    try:
        return serializer.loads(
            token,
            salt=ACCOUNT_TOKEN_SALT,
            max_age=ACCOUNT_TOKEN_MAX_AGE_SECONDS,
        )
    except (BadSignature, BadTimeSignature, SignatureExpired):
        raise ValueError("账号登录凭证无效或已过期")


def issue_tt_account_token(*, user_row: sqlite3.Row, machine_id: str) -> str:
    serializer = get_license_serializer()
    payload = {
        "tt_user_id": user_row["id"],
        "username": user_row["username"],
        "email": user_row["email"] or "",
        "machine_id": machine_id,
        "edition": user_row["edition"],
        "expires_at": user_row["expires_at"] or "",
        "issued_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
    }
    return serializer.dumps(payload, salt=TT_ACCOUNT_TOKEN_SALT)


def verify_tt_account_token(token: str) -> dict:
    serializer = get_license_serializer()
    try:
        return serializer.loads(
            token,
            salt=TT_ACCOUNT_TOKEN_SALT,
            max_age=ACCOUNT_TOKEN_MAX_AGE_SECONDS,
        )
    except (BadSignature, BadTimeSignature, SignatureExpired):
        raise ValueError("TT账号登录凭证无效或已过期")


def parse_iso_datetime(value: str | None) -> datetime.datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.datetime.fromisoformat(text)
    except ValueError:
        return None


def is_license_expired(license_row: sqlite3.Row) -> bool:
    expires_at = parse_iso_datetime(license_row["expires_at"])
    if not expires_at:
        return False
    return expires_at <= datetime.datetime.now()


def is_user_account_expired(user_row: sqlite3.Row) -> bool:
    expires_at = parse_iso_datetime(user_row["expires_at"])
    if not expires_at:
        return False
    return expires_at <= datetime.datetime.now()


def get_user_by_account(db: sqlite3.Connection, account: str) -> sqlite3.Row | None:
    value = str(account or "").strip()
    if not value:
        return None
    return db.execute(
        """
        SELECT *
        FROM users
        WHERE username = ? OR lower(COALESCE(email, '')) = lower(?)
        """,
        (value, value),
    ).fetchone()


def get_tt_user_by_account(db: sqlite3.Connection, account: str) -> sqlite3.Row | None:
    value = str(account or "").strip()
    if not value:
        return None
    return db.execute(
        """
        SELECT *
        FROM tt_users
        WHERE username = ? OR lower(COALESCE(email, '')) = lower(?)
        """,
        (value, value),
    ).fetchone()


def _users_has_email_column(db: sqlite3.Connection) -> bool:
    try:
        rows = db.execute("PRAGMA table_info(users)").fetchall()
    except Exception:
        return False
    for row in rows:
        if isinstance(row, sqlite3.Row):
            name = str(row["name"] or "")
        else:
            name = str(row[1] if len(row) > 1 else "")
        if name == "email":
            return True
    return False


def _find_existing_registration_conflict(
    db: sqlite3.Connection,
    *,
    username: str,
    email: str,
) -> str | None:
    username = str(username or "").strip()
    email = normalize_email(str(email or ""))
    existing_username = db.execute(
        "SELECT id FROM users WHERE username = ?",
        (username,),
    ).fetchone()
    if existing_username:
        return "用户名已存在"
    if email and _users_has_email_column(db):
        existing_email = db.execute(
            "SELECT id FROM users WHERE lower(COALESCE(email, '')) = lower(?)",
            (email,),
        ).fetchone()
        if existing_email:
            return "邮箱已存在"
    return None


def current_active_user_device_count(db: sqlite3.Connection, user_id: int) -> int:
    row = db.execute(
        """
        SELECT COUNT(*) AS cnt
        FROM user_devices
        WHERE user_id = ? AND (revoked_at IS NULL OR revoked_at = '')
        """,
        (user_id,),
    ).fetchone()
    return int(row["cnt"] if row else 0)


def current_active_tt_user_device_count(db: sqlite3.Connection, tt_user_id: int) -> int:
    row = db.execute(
        """
        SELECT COUNT(*) AS cnt
        FROM tt_user_devices
        WHERE tt_user_id = ? AND (revoked_at IS NULL OR revoked_at = '')
        """,
        (tt_user_id,),
    ).fetchone()
    return int(row["cnt"] if row else 0)


def ensure_tt_account_device_limit_for_machine(
    db: sqlite3.Connection,
    user_row: sqlite3.Row,
    *,
    machine_id: str,
) -> tuple[bool, str]:
    active_count = current_active_tt_user_device_count(db, user_row["id"])
    max_devices = max(1, int(user_row["max_devices"] or ACCOUNT_DEFAULT_MAX_DEVICES))
    if active_count <= max_devices:
        return True, ""
    retained_rows = db.execute(
        """
        SELECT id, machine_id
        FROM tt_user_devices
        WHERE tt_user_id = ? AND (revoked_at IS NULL OR revoked_at = '')
        ORDER BY COALESCE(last_verified_at, logged_in_at, '') DESC, id DESC
        LIMIT ?
        """,
        (user_row["id"], max_devices),
    ).fetchall()
    retained_machine_ids = {str(row["machine_id"] or "").strip() for row in retained_rows}
    if str(machine_id or "").strip() in retained_machine_ids:
        return True, ""
    return (
        False,
        f"TT账号登录设备已超过上限（{active_count}/{max_devices}），请在管理后台解绑设备后重新登录",
    )


def current_active_activation_count(db: sqlite3.Connection, license_id: int) -> int:
    row = db.execute(
        """
        SELECT COUNT(*) AS cnt
        FROM license_activations
        WHERE license_id = ? AND (revoked_at IS NULL OR revoked_at = '')
        """,
        (license_id,),
    ).fetchone()
    return int(row["cnt"] if row else 0)


def resolve_user_from_account_token_payload(
    data: dict[str, Any],
) -> tuple[sqlite3.Row | None, dict[str, str], tuple[Any, int] | None]:
    payload, error = validate_account_auth_payload(data)
    if error:
        return None, payload, (jsonify({"ok": False, "message": error}), 400)
    if not payload["token"]:
        return None, payload, (jsonify({"ok": False, "message": "登录凭证不能为空"}), 400)

    try:
        token_payload = verify_account_token(payload["token"])
    except ValueError as exc:
        return None, payload, (jsonify({"ok": False, "message": str(exc)}), 400)
    if token_payload.get("machine_id") != payload["machine_id"]:
        return None, payload, (jsonify({"ok": False, "message": "登录凭证与当前机器不匹配"}), 400)

    db = get_db()
    user_row = db.execute(
        "SELECT * FROM users WHERE id = ?",
        (token_payload.get("user_id"),),
    ).fetchone()
    if not user_row:
        return None, payload, (jsonify({"ok": False, "message": "账号不存在"}), 404)
    if payload["account"] and payload["account"] not in {user_row["username"], user_row["email"] or ""}:
        return None, payload, (jsonify({"ok": False, "message": "登录凭证与当前账号不匹配"}), 400)
    ok, account_error = ensure_account_can_login(user_row)
    if not ok:
        return None, payload, (jsonify({"ok": False, "message": account_error}), 400)

    device_row = db.execute(
        """
        SELECT *
        FROM user_devices
        WHERE user_id = ? AND machine_id = ?
        """,
        (user_row["id"], payload["machine_id"]),
    ).fetchone()
    if not device_row:
        return None, payload, (jsonify({"ok": False, "message": "当前机器未登录或登录凭证已失效"}), 400)
    if str(device_row["revoked_at"] or "").strip():
        return None, payload, (jsonify({"ok": False, "message": "当前机器登录已退出，请重新登录"}), 400)
    if device_row["token_hash"] != hash_token(payload["token"]):
        return None, payload, (jsonify({"ok": False, "message": "登录凭证已失效，请重新登录"}), 400)
    return user_row, payload, None


def current_total_activation_count(db: sqlite3.Connection, license_id: int) -> int:
    row = db.execute(
        """
        SELECT COUNT(*) AS cnt
        FROM license_activations
        WHERE license_id = ?
        """,
        (license_id,),
    ).fetchone()
    return int(row["cnt"] if row else 0)


def latest_license_verification_at(db: sqlite3.Connection, license_id: int) -> str:
    row = db.execute(
        """
        SELECT MAX(last_verified_at) AS last_verified_at
        FROM license_activations
        WHERE license_id = ?
        """,
        (license_id,),
    ).fetchone()
    if not row:
        return ""
    return str(row["last_verified_at"] or "")


def serialize_license_row(db: sqlite3.Connection, row: sqlite3.Row) -> dict:
    item = dict(row)
    item["active_activations"] = current_active_activation_count(db, row["id"])
    item["total_activations"] = current_total_activation_count(db, row["id"])
    item["last_verified_at"] = latest_license_verification_at(db, row["id"])
    item["is_deleted"] = bool(item.get("deleted_at"))
    return item


def serialize_activation_row(row: sqlite3.Row) -> dict:
    return dict(row)


def get_license_row(
    db: sqlite3.Connection,
    license_id: int,
    *,
    include_deleted: bool = False,
) -> sqlite3.Row | None:
    sql = "SELECT * FROM licenses WHERE id = ?"
    params: list[object] = [license_id]
    if not include_deleted:
        sql += " AND deleted_at IS NULL"
    return db.execute(sql, params).fetchone()


def get_license_rows_by_ids(
    db: sqlite3.Connection,
    license_ids: list[int],
    *,
    include_deleted: bool = False,
) -> list[sqlite3.Row]:
    if not license_ids:
        return []
    placeholders = ",".join(["?"] * len(license_ids))
    sql = f"SELECT * FROM licenses WHERE id IN ({placeholders})"
    params: list[object] = list(license_ids)
    if not include_deleted:
        sql += " AND deleted_at IS NULL"
    sql += " ORDER BY id DESC"
    return db.execute(sql, params).fetchall()


def build_license_filter_clause(args) -> tuple[list[str], list[object]]:
    clauses: list[str] = []
    params: list[object] = []

    keyword = str(args.get("keyword") or "").strip()
    if keyword:
        like = f"%{keyword}%"
        clauses.append(
            "(license_key LIKE ? OR license_key_masked LIKE ? OR COALESCE(licensee, '') LIKE ? OR COALESCE(notes, '') LIKE ?)"
        )
        params.extend([like, like, like, like])

    edition = str(args.get("edition") or "").strip().lower()
    if edition in LICENSE_EDITION_VALUES:
        clauses.append("edition = ?")
        params.append(edition)

    status = str(args.get("status") or "").strip().lower()
    show_deleted = str(args.get("show_deleted") or "").strip() == "1"
    if status == "deleted":
        clauses.append("deleted_at IS NOT NULL")
    else:
        if not show_deleted:
            clauses.append("deleted_at IS NULL")
        if status in LICENSE_STATUS_VALUES:
            clauses.append("status = ?")
            params.append(status)

    return clauses, params


def parse_license_ids_from_payload(data: dict) -> tuple[list[int], str | None]:
    raw_ids = data.get("ids") or []
    if not isinstance(raw_ids, list) or not raw_ids:
        return [], "请选择至少一条授权码"
    ids: list[int] = []
    for item in raw_ids:
        try:
            value = int(item)
        except (TypeError, ValueError):
            continue
        if value > 0:
            ids.append(value)
    ids = list(dict.fromkeys(ids))
    if not ids:
        return [], "请选择至少一条有效的授权码"
    return ids, None


def revoke_license_activations(db: sqlite3.Connection, license_id: int) -> int:
    result = db.execute(
        """
        UPDATE license_activations
        SET revoked_at = ?
        WHERE license_id = ? AND (revoked_at IS NULL OR revoked_at = '')
        """,
        (now_iso(), license_id),
    )
    return int(result.rowcount or 0)


def soft_delete_license_row(db: sqlite3.Connection, row: sqlite3.Row, *, deleted_by: int | None) -> tuple[bool, str]:
    if row["deleted_at"]:
        return False, "该激活码已删除"
    revoke_license_activations(db, row["id"])
    db.execute(
        """
        UPDATE licenses
        SET status = 'disabled', deleted_at = ?, deleted_by = ?, updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (now_iso(), deleted_by, row["id"]),
    )
    return True, ""


def restore_license_row(db: sqlite3.Connection, row: sqlite3.Row) -> tuple[bool, str]:
    if not row["deleted_at"]:
        return False, "该激活码未删除，无需恢复"
    db.execute(
        """
        UPDATE licenses
        SET deleted_at = NULL, deleted_by = NULL, updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (row["id"],),
    )
    return True, ""


def update_license_status_row(
    db: sqlite3.Connection,
    row: sqlite3.Row,
    *,
    status: str,
) -> tuple[bool, str]:
    if row["deleted_at"]:
        return False, f"{row['license_key_masked']} 已删除，不能修改状态"
    if status not in LICENSE_STATUS_VALUES:
        return False, "无效的授权码状态"
    if row["status"] == status:
        return True, ""
    db.execute(
        "UPDATE licenses SET status = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (status, row["id"]),
    )
    return True, ""


def generate_remote_client_id() -> str:
    return f"rc_{secrets.token_hex(8)}"


def generate_remote_client_token() -> str:
    return secrets.token_urlsafe(24)


def hash_remote_client_token(token: str) -> str:
    return hashlib.sha256(str(token or "").encode("utf-8")).hexdigest()


def now_iso() -> str:
    return datetime.datetime.now().isoformat(timespec="seconds")


def now_timestamp() -> int:
    return int(datetime.datetime.now().timestamp())


def get_app_setting(setting_key: str) -> sqlite3.Row | None:
    key = str(setting_key or "").strip()
    if not key:
        return None
    return get_db().execute(
        "SELECT * FROM app_settings WHERE setting_key = ?",
        (key,),
    ).fetchone()


def get_app_setting_value(setting_key: str) -> str:
    row = get_app_setting(setting_key)
    return str(row["setting_value"] or "").strip() if row else ""


def mask_secret_value(secret: str) -> str:
    text = str(secret or "").strip()
    if len(text) >= 8:
        return f"{text[:4]}{'*' * 8}{text[-4:]}"
    if text:
        return "*" * len(text)
    return ""


def set_app_setting(setting_key: str, setting_value: str, *, is_secret: bool, updated_by: int | None = None) -> None:
    get_db().execute(
        """
        INSERT INTO app_settings (setting_key, setting_value, is_secret, updated_at, updated_by)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(setting_key) DO UPDATE SET
            setting_value = excluded.setting_value,
            is_secret = excluded.is_secret,
            updated_at = excluded.updated_at,
            updated_by = excluded.updated_by
        """,
        (
            str(setting_key or "").strip(),
            str(setting_value or "").strip(),
            1 if is_secret else 0,
            now_iso(),
            updated_by,
        ),
    )


def migrate_legacy_minidrama_settings(db: sqlite3.Connection) -> None:
    existing_count = db.execute("SELECT COUNT(*) FROM minidrama_apps").fetchone()[0]
    if existing_count:
        return
    app_id_row = db.execute(
        "SELECT * FROM app_settings WHERE setting_key = ?",
        (MINIDRAMA_APP_ID_SETTING_KEY,),
    ).fetchone()
    app_secret_row = db.execute(
        "SELECT * FROM app_settings WHERE setting_key = ?",
        (MINIDRAMA_APP_SECRET_SETTING_KEY,),
    ).fetchone()
    app_id = str(app_id_row["setting_value"] or "").strip() if app_id_row else ""
    app_secret = str(app_secret_row["setting_value"] or "").strip() if app_secret_row else ""
    if not app_id or not app_secret:
        return
    now = now_iso()
    updated_by = app_secret_row["updated_by"] if app_secret_row else app_id_row["updated_by"] if app_id_row else None
    db.execute(
        """
        INSERT OR IGNORE INTO minidrama_apps (
            app_id, app_secret, name, is_default, enabled, created_at, updated_at, updated_by
        ) VALUES (?, ?, ?, 1, 1, ?, ?, ?)
        """,
        (app_id, app_secret, "默认小程序", now, now, updated_by),
    )


init_db()


def serialize_minidrama_app(row: sqlite3.Row | dict[str, Any], *, include_secret: bool = False) -> dict[str, Any]:
    app_secret = str((row["app_secret"] if isinstance(row, sqlite3.Row) else row.get("app_secret")) or "").strip()
    payload = {
        "id": int(row["id"] if isinstance(row, sqlite3.Row) else row.get("id") or 0),
        "app_id": str((row["app_id"] if isinstance(row, sqlite3.Row) else row.get("app_id")) or "").strip(),
        "name": str((row["name"] if isinstance(row, sqlite3.Row) else row.get("name")) or "").strip(),
        "enabled": bool((row["enabled"] if isinstance(row, sqlite3.Row) else row.get("enabled", True))),
        "is_default": bool((row["is_default"] if isinstance(row, sqlite3.Row) else row.get("is_default", False))),
        "app_secret_configured": bool(app_secret),
        "app_secret_masked": mask_secret_value(app_secret),
        "source": "database",
        "updated_at": str((row["updated_at"] if isinstance(row, sqlite3.Row) else row.get("updated_at")) or ""),
        "updated_by": row["updated_by"] if isinstance(row, sqlite3.Row) else row.get("updated_by"),
    }
    if include_secret:
        payload["app_secret"] = app_secret
    return payload


def list_minidrama_app_rows(*, include_disabled: bool = True) -> list[sqlite3.Row]:
    query = "SELECT * FROM minidrama_apps"
    params: tuple[Any, ...] = ()
    if not include_disabled:
        query += " WHERE enabled = 1"
    query += " ORDER BY is_default DESC, updated_at DESC, id DESC"
    return list(get_db().execute(query, params).fetchall())


def get_minidrama_app_row(app_id: str) -> sqlite3.Row | None:
    normalized = str(app_id or "").strip()
    if not normalized:
        return None
    return get_db().execute("SELECT * FROM minidrama_apps WHERE app_id = ?", (normalized,)).fetchone()


def get_default_minidrama_app_row() -> sqlite3.Row | None:
    row = get_db().execute(
        "SELECT * FROM minidrama_apps WHERE enabled = 1 AND is_default = 1 ORDER BY id DESC LIMIT 1"
    ).fetchone()
    if row:
        return row
    return get_db().execute(
        "SELECT * FROM minidrama_apps WHERE enabled = 1 ORDER BY updated_at DESC, id DESC LIMIT 1"
    ).fetchone()


def save_minidrama_app(
    *,
    app_id: str,
    app_secret: str,
    name: str = "",
    enabled: bool = True,
    is_default: bool = False,
    updated_by: int | None = None,
) -> sqlite3.Row:
    db = get_db()
    normalized_app_id = str(app_id or "").strip()
    existing = get_minidrama_app_row(normalized_app_id)
    now = now_iso()
    if is_default:
        db.execute("UPDATE minidrama_apps SET is_default = 0 WHERE app_id <> ?", (normalized_app_id,))
    if existing:
        db.execute(
            """
            UPDATE minidrama_apps
            SET app_secret = ?, name = ?, enabled = ?, is_default = ?, updated_at = ?, updated_by = ?
            WHERE app_id = ?
            """,
            (
                str(app_secret or "").strip(),
                str(name or "").strip(),
                1 if enabled else 0,
                1 if is_default else int(existing["is_default"] or 0),
                now,
                updated_by,
                normalized_app_id,
            ),
        )
    else:
        if not is_default and not get_default_minidrama_app_row():
            is_default = True
        db.execute(
            """
            INSERT INTO minidrama_apps (
                app_id, app_secret, name, enabled, is_default, created_at, updated_at, updated_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                normalized_app_id,
                str(app_secret or "").strip(),
                str(name or "").strip(),
                1 if enabled else 0,
                1 if is_default else 0,
                now,
                now,
                updated_by,
            ),
        )
    db.execute("DELETE FROM minidrama_token_cache WHERE app_id = ?", (normalized_app_id,))
    if not get_default_minidrama_app_row():
        db.execute(
            """
            UPDATE minidrama_apps
            SET is_default = 1
            WHERE id = (
                SELECT id FROM minidrama_apps WHERE enabled = 1 ORDER BY updated_at DESC, id DESC LIMIT 1
            )
            """
        )
    return get_minidrama_app_row(normalized_app_id)


def serialize_kuaishou_app(row: sqlite3.Row | dict[str, Any], *, include_secret: bool = False) -> dict[str, Any]:
    app_secret = str((row["app_secret"] if isinstance(row, sqlite3.Row) else row.get("app_secret")) or "").strip()
    payload = {
        "id": int(row["id"] if isinstance(row, sqlite3.Row) else row.get("id") or 0),
        "app_id": str((row["app_id"] if isinstance(row, sqlite3.Row) else row.get("app_id")) or "").strip(),
        "advertiser_id": str((row["advertiser_id"] if isinstance(row, sqlite3.Row) else row.get("advertiser_id")) or "").strip(),
        "name": str((row["name"] if isinstance(row, sqlite3.Row) else row.get("name")) or "").strip(),
        "enabled": bool((row["enabled"] if isinstance(row, sqlite3.Row) else row.get("enabled", True))),
        "is_default": bool((row["is_default"] if isinstance(row, sqlite3.Row) else row.get("is_default", False))),
        "app_secret_configured": bool(app_secret),
        "app_secret_masked": mask_secret_value(app_secret),
        "access_token_configured": bool(str((row["access_token"] if isinstance(row, sqlite3.Row) else row.get("access_token")) or "").strip()),
        "refresh_token_configured": bool(str((row["refresh_token"] if isinstance(row, sqlite3.Row) else row.get("refresh_token")) or "").strip()),
        "access_token_expires_at": int((row["access_token_expires_at"] if isinstance(row, sqlite3.Row) else row.get("access_token_expires_at") or 0) or 0),
        "refresh_token_expires_at": int((row["refresh_token_expires_at"] if isinstance(row, sqlite3.Row) else row.get("refresh_token_expires_at") or 0) or 0),
        "source": "database",
        "updated_at": str((row["updated_at"] if isinstance(row, sqlite3.Row) else row.get("updated_at")) or ""),
        "updated_by": row["updated_by"] if isinstance(row, sqlite3.Row) else row.get("updated_by"),
    }
    if include_secret:
        payload["app_secret"] = app_secret
        payload["access_token"] = str((row["access_token"] if isinstance(row, sqlite3.Row) else row.get("access_token")) or "").strip()
        payload["refresh_token"] = str((row["refresh_token"] if isinstance(row, sqlite3.Row) else row.get("refresh_token")) or "").strip()
    return payload


def list_kuaishou_app_rows(*, include_disabled: bool = True) -> list[sqlite3.Row]:
    query = "SELECT * FROM kuaishou_apps"
    params: tuple[Any, ...] = ()
    if not include_disabled:
        query += " WHERE enabled = 1"
    query += " ORDER BY is_default DESC, updated_at DESC, id DESC"
    return list(get_db().execute(query, params).fetchall())


def get_kuaishou_app_row(app_id: str) -> sqlite3.Row | None:
    normalized = str(app_id or "").strip()
    if not normalized:
        return None
    return get_db().execute("SELECT * FROM kuaishou_apps WHERE app_id = ?", (normalized,)).fetchone()


def get_default_kuaishou_app_row() -> sqlite3.Row | None:
    row = get_db().execute(
        "SELECT * FROM kuaishou_apps WHERE enabled = 1 AND is_default = 1 ORDER BY id DESC LIMIT 1"
    ).fetchone()
    if row:
        return row
    return get_db().execute(
        "SELECT * FROM kuaishou_apps WHERE enabled = 1 ORDER BY updated_at DESC, id DESC LIMIT 1"
    ).fetchone()


def save_kuaishou_app(*, app_id: str, app_secret: str, advertiser_id: str = "", name: str = "", access_token: str = "", refresh_token: str = "", access_token_expires_at: int = 0, refresh_token_expires_at: int = 0, enabled: bool = True, is_default: bool = False, updated_by: int | None = None) -> sqlite3.Row:
    db = get_db()
    normalized_app_id = str(app_id or "").strip()
    existing = get_kuaishou_app_row(normalized_app_id)
    now = now_iso()
    if is_default:
        db.execute("UPDATE kuaishou_apps SET is_default = 0 WHERE app_id <> ?", (normalized_app_id,))
    if existing:
        db.execute(
            "UPDATE kuaishou_apps SET app_secret = ?, advertiser_id = ?, name = ?, access_token = ?, refresh_token = ?, access_token_expires_at = ?, refresh_token_expires_at = ?, enabled = ?, is_default = ?, updated_at = ?, updated_by = ? WHERE app_id = ?",
            (str(app_secret or "").strip(), str(advertiser_id or "").strip(), str(name or "").strip(), str(access_token or "").strip(), str(refresh_token or "").strip(), int(access_token_expires_at or 0), int(refresh_token_expires_at or 0), 1 if enabled else 0, 1 if is_default else int(existing["is_default"] or 0), now, updated_by, normalized_app_id),
        )
    else:
        if not is_default and not get_default_kuaishou_app_row():
            is_default = True
        db.execute(
            "INSERT INTO kuaishou_apps (app_id, app_secret, advertiser_id, name, access_token, refresh_token, access_token_expires_at, refresh_token_expires_at, enabled, is_default, created_at, updated_at, updated_by) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (normalized_app_id, str(app_secret or "").strip(), str(advertiser_id or "").strip(), str(name or "").strip(), str(access_token or "").strip(), str(refresh_token or "").strip(), int(access_token_expires_at or 0), int(refresh_token_expires_at or 0), 1 if enabled else 0, 1 if is_default else 0, now, now, updated_by),
        )
    if str(access_token or "").strip() and int(access_token_expires_at or 0) > 0:
        db.execute(
            "INSERT INTO kuaishou_token_cache (app_id, access_token, expires_at, refreshing_by, refreshing_until, last_error, created_at, updated_at) VALUES (?, ?, ?, NULL, 0, NULL, ?, ?) ON CONFLICT(app_id) DO UPDATE SET access_token = excluded.access_token, expires_at = excluded.expires_at, refreshing_by = NULL, refreshing_until = 0, last_error = NULL, updated_at = excluded.updated_at",
            (normalized_app_id, str(access_token or "").strip(), int(access_token_expires_at or 0), now, now),
        )
    else:
        db.execute("DELETE FROM kuaishou_token_cache WHERE app_id = ?", (normalized_app_id,))
    if not get_default_kuaishou_app_row():
        db.execute("UPDATE kuaishou_apps SET is_default = 1 WHERE id = (SELECT id FROM kuaishou_apps WHERE enabled = 1 ORDER BY updated_at DESC, id DESC LIMIT 1)")
    return get_kuaishou_app_row(normalized_app_id)


def get_kuaishou_server_settings(requested_app_id: str = "") -> dict[str, Any]:
    requested = str(requested_app_id or "").strip()
    if requested:
        row = get_kuaishou_app_row(requested)
        if row:
            return serialize_kuaishou_app(row, include_secret=True)
        return {"app_id": requested, "app_secret": "", "advertiser_id": "", "access_token": "", "refresh_token": "", "access_token_expires_at": 0, "refresh_token_expires_at": 0, "app_secret_configured": False, "source": "empty", "updated_at": "", "updated_by": None}
    row = get_default_kuaishou_app_row()
    if row:
        payload = serialize_kuaishou_app(row, include_secret=True)
        payload["apps"] = [serialize_kuaishou_app(item) for item in list_kuaishou_app_rows()]
        return payload
    return {"app_id": "", "app_secret": "", "advertiser_id": "", "access_token": "", "refresh_token": "", "access_token_expires_at": 0, "refresh_token_expires_at": 0, "app_secret_configured": False, "source": "empty", "updated_at": "", "updated_by": None, "apps": []}


def serialize_kuaishou_settings(settings: dict[str, Any]) -> dict[str, Any]:
    app_secret = str(settings.get("app_secret") or "").strip()
    access_token = str(settings.get("access_token") or "").strip()
    refresh_token = str(settings.get("refresh_token") or "").strip()
    return {"app_id": str(settings.get("app_id") or "").strip(), "advertiser_id": str(settings.get("advertiser_id") or "").strip(), "name": str(settings.get("name") or "").strip(), "enabled": bool(settings.get("enabled", True)), "is_default": bool(settings.get("is_default", False)), "app_secret_configured": bool(app_secret), "app_secret_masked": mask_secret_value(app_secret), "access_token_configured": bool(access_token), "refresh_token_configured": bool(refresh_token), "access_token_expires_at": int(settings.get("access_token_expires_at") or 0), "refresh_token_expires_at": int(settings.get("refresh_token_expires_at") or 0), "source": str(settings.get("source") or "empty"), "updated_at": str(settings.get("updated_at") or ""), "updated_by": settings.get("updated_by"), "apps": settings.get("apps") if isinstance(settings.get("apps"), list) else []}


def resolve_kuaishou_server_credentials(requested_app_id: str = "") -> tuple[sqlite3.Row | None, str | None]:
    settings = get_kuaishou_server_settings(requested_app_id)
    app_id = str(settings.get("app_id") or "").strip()
    app_secret = str(settings.get("app_secret") or "").strip()
    requested = str(requested_app_id or "").strip()
    if requested and app_id and requested != app_id:
        return None, "????? AppID ?????????"
    if not app_id:
        return None, "???????? AppID"
    if not app_secret:
        return None, f"???????? AppSecret?{app_id}"
    if settings.get("enabled") is False:
        return None, f"???????????{app_id}"
    return get_kuaishou_app_row(app_id), None


def refresh_kuaishou_access_token_from_api(app_id: str, app_secret: str, refresh_token: str) -> dict[str, Any]:
    body = json.dumps({"app_id": int(app_id) if str(app_id).isdigit() else str(app_id), "secret": str(app_secret or "").strip(), "refresh_token": str(refresh_token or "").strip()}, ensure_ascii=False).encode("utf-8")
    req = urlrequest.Request(f"{KUAISHOU_API_BASE}/rest/openapi/oauth2/authorize/refresh_token", data=body, method="POST", headers={"Content-Type": "application/json", "Accept": "application/json"})
    with urlrequest.urlopen(req, timeout=20) as response:
        payload = json.loads(response.read().decode("utf-8", errors="replace") or "{}")
    if not isinstance(payload, dict):
        raise ValueError("?? refresh token ????????")
    if int(payload.get("code") or 0) != 0:
        raise ValueError(f"?? refresh token ?????code={payload.get('code')}, message={payload.get('message') or ''}")
    data = payload.get("data")
    if not isinstance(data, dict) or not str(data.get("access_token") or "").strip():
        raise ValueError("?? refresh token ????? access_token")
    return data



def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _kuaishou_token_expiry_from_response(data: dict[str, Any], *, expires_in_key: str, expires_at_key: str, fallback: int = 0) -> int:
    now_ts = now_timestamp()
    expires_in = _safe_int(data.get(expires_in_key), 0)
    if expires_in > 0:
        return now_ts + max(0, expires_in - 120)
    expires_at = _safe_int(data.get(expires_at_key), 0)
    if expires_at > 0:
        return expires_at
    return int(fallback or 0)


def _kuaishou_app_needs_token_refresh(row: sqlite3.Row, *, now_ts: int | None = None) -> bool:
    current_ts = now_timestamp() if now_ts is None else int(now_ts)
    if not bool(row["enabled"]):
        return False
    if not str(row["app_secret"] or "").strip() or not str(row["refresh_token"] or "").strip():
        return False
    refresh_expires_at = _safe_int(row["refresh_token_expires_at"], 0)
    if refresh_expires_at > 0 and refresh_expires_at <= current_ts:
        return False
    access_token = str(row["access_token"] or "").strip()
    access_expires_at = _safe_int(row["access_token_expires_at"], 0)
    if not access_token or access_expires_at <= 0:
        return True
    if access_expires_at - current_ts <= KUAISHOU_ACCESS_TOKEN_RENEW_MARGIN_SECONDS:
        return True
    if refresh_expires_at > 0 and refresh_expires_at - current_ts <= KUAISHOU_REFRESH_TOKEN_RENEW_MARGIN_SECONDS:
        return True
    return False


def _kuaishou_token_payload(row: sqlite3.Row, *, cached: bool, refreshing: bool = False) -> dict[str, Any]:
    expires_at = _safe_int(row["access_token_expires_at"], 0)
    return {
        "app_id": str(row["app_id"] or "").strip(),
        "advertiser_id": str(row["advertiser_id"] or "").strip(),
        "access_token": str(row["access_token"] or "").strip(),
        "expires_at": expires_at,
        "access_token_expires_at": expires_at,
        "refresh_token_expires_at": _safe_int(row["refresh_token_expires_at"], 0),
        "expires_in": max(0, expires_at - now_timestamp()),
        "cached": cached,
        "refreshing": refreshing,
    }


def _release_kuaishou_refresh_lock(db: sqlite3.Connection, app_id: str, *, last_error: str | None = None) -> None:
    if last_error:
        db.execute(
            "UPDATE kuaishou_token_cache SET refreshing_by = NULL, refreshing_until = 0, last_error = ?, updated_at = ? WHERE app_id = ?",
            (last_error, now_iso(), app_id),
        )
    else:
        db.execute(
            "UPDATE kuaishou_token_cache SET refreshing_by = NULL, refreshing_until = 0, updated_at = ? WHERE app_id = ?",
            (now_iso(), app_id),
        )


def refresh_kuaishou_app_token_if_due(app_id: str, *, refreshing_by: str, force: bool = False) -> tuple[dict[str, Any] | None, str | None, int]:
    db = get_db()
    row, error_message = resolve_kuaishou_server_credentials(app_id)
    if error_message:
        return None, error_message, 400
    assert row is not None
    app_id = str(row["app_id"] or "").strip()
    now_ts = now_timestamp()
    now_text = now_iso()

    cache_row = db.execute("SELECT * FROM kuaishou_token_cache WHERE app_id = ?", (app_id,)).fetchone()
    if (
        not force
        and cache_row
        and str(cache_row["access_token"] or "").strip()
        and _safe_int(cache_row["expires_at"], 0) - now_ts > KUAISHOU_TOKEN_REFRESH_MARGIN_SECONDS
    ):
        return {
            "app_id": app_id,
            "advertiser_id": str(row["advertiser_id"] or "").strip(),
            "access_token": str(cache_row["access_token"] or "").strip(),
            "expires_at": _safe_int(cache_row["expires_at"], 0),
            "access_token_expires_at": _safe_int(cache_row["expires_at"], 0),
            "refresh_token_expires_at": _safe_int(row["refresh_token_expires_at"], 0),
            "expires_in": max(0, _safe_int(cache_row["expires_at"], 0) - now_ts),
            "cached": True,
        }, None, 200

    if not force and not _kuaishou_app_needs_token_refresh(row, now_ts=now_ts):
        return _kuaishou_token_payload(row, cached=True), None, 200

    db.execute(
        "INSERT OR IGNORE INTO kuaishou_token_cache (app_id, access_token, expires_at, refreshing_by, refreshing_until, created_at, updated_at) VALUES (?, NULL, 0, NULL, 0, ?, ?)",
        (app_id, now_text, now_text),
    )
    db.commit()
    lock_owner = str(refreshing_by or "system")
    lock_until = now_ts + KUAISHOU_TOKEN_REFRESH_LOCK_SECONDS
    cursor = db.execute(
        "UPDATE kuaishou_token_cache SET refreshing_by = ?, refreshing_until = ?, updated_at = ? WHERE app_id = ? AND (refreshing_until IS NULL OR refreshing_until <= ? OR refreshing_by = ?)",
        (lock_owner, lock_until, now_text, app_id, now_ts, lock_owner),
    )
    db.commit()
    if cursor.rowcount <= 0:
        latest = get_kuaishou_app_row(app_id)
        if latest and str(latest["access_token"] or "").strip() and _safe_int(latest["access_token_expires_at"], 0) > now_ts + 60:
            return _kuaishou_token_payload(latest, cached=True, refreshing=True), None, 200
        return None, "快手 token 正在刷新，请稍后重试", 409

    row = get_kuaishou_app_row(app_id)
    if row is None:
        _release_kuaishou_refresh_lock(db, app_id, last_error="快手配置不存在")
        db.commit()
        return None, "快手配置不存在", 404
    refresh_token = str(row["refresh_token"] or "").strip()
    if not refresh_token:
        _release_kuaishou_refresh_lock(db, app_id, last_error="服务端未配置快手 refresh_token")
        db.commit()
        return None, "服务端未配置快手 refresh_token，请先在桌面端登录并同步", 400
    refresh_expires_at = _safe_int(row["refresh_token_expires_at"], 0)
    if refresh_expires_at > 0 and refresh_expires_at <= now_ts:
        _release_kuaishou_refresh_lock(db, app_id, last_error="快手 refresh_token 已过期")
        db.commit()
        return None, "快手 refresh_token 已过期，请在桌面端重新登录后同步", 400

    try:
        token_data = refresh_kuaishou_access_token_from_api(app_id, str(row["app_secret"] or "").strip(), refresh_token)
    except Exception as exc:
        _release_kuaishou_refresh_lock(db, app_id, last_error=str(exc))
        db.commit()
        return None, str(exc), 502

    access_token = str(token_data.get("access_token") or "").strip()
    new_refresh_token = str(token_data.get("refresh_token") or refresh_token).strip()
    advertiser_id = str(token_data.get("advertiser_id") or token_data.get("user_id") or row["advertiser_id"] or "").strip()
    access_expires_at = _kuaishou_token_expiry_from_response(
        token_data,
        expires_in_key="access_token_expires_in",
        expires_at_key="access_token_expires_at",
        fallback=_safe_int(row["access_token_expires_at"], 0),
    )
    refresh_expires_at = _kuaishou_token_expiry_from_response(
        token_data,
        expires_in_key="refresh_token_expires_in",
        expires_at_key="refresh_token_expires_at",
        fallback=_safe_int(row["refresh_token_expires_at"], 0),
    )
    cursor = db.execute(
        "UPDATE kuaishou_apps SET advertiser_id = ?, access_token = ?, refresh_token = ?, access_token_expires_at = ?, refresh_token_expires_at = ?, updated_at = ? WHERE app_id = ? AND refresh_token = ?",
        (advertiser_id, access_token, new_refresh_token, access_expires_at, refresh_expires_at, now_iso(), app_id, refresh_token),
    )
    if cursor.rowcount <= 0:
        _release_kuaishou_refresh_lock(db, app_id)
        db.commit()
        latest = get_kuaishou_app_row(app_id)
        if latest and str(latest["access_token"] or "").strip():
            return _kuaishou_token_payload(latest, cached=True, refreshing=True), None, 200
        return None, "快手 token 已被其他进程刷新，请稍后重试", 409
    db.execute(
        "UPDATE kuaishou_token_cache SET access_token = ?, expires_at = ?, refreshing_by = NULL, refreshing_until = 0, last_error = NULL, updated_at = ? WHERE app_id = ?",
        (access_token, access_expires_at, now_iso(), app_id),
    )
    db.commit()
    latest = get_kuaishou_app_row(app_id)
    return _kuaishou_token_payload(latest, cached=False) if latest else None, None, 200


def refresh_due_kuaishou_apps_once() -> dict[str, int]:
    summary = {"checked": 0, "refreshed": 0, "failed": 0, "skipped": 0}
    db = get_db()
    rows = list(db.execute("SELECT * FROM kuaishou_apps WHERE enabled = 1 ORDER BY updated_at ASC, id ASC").fetchall())
    now_ts = now_timestamp()
    for row in rows:
        summary["checked"] += 1
        if not _kuaishou_app_needs_token_refresh(row, now_ts=now_ts):
            summary["skipped"] += 1
            continue
        payload, error_message, _status = refresh_kuaishou_app_token_if_due(
            str(row["app_id"] or ""),
            refreshing_by="system-scheduler",
            force=True,
        )
        if payload and not error_message:
            summary["refreshed"] += 1
        else:
            summary["failed"] += 1
            print(f"[kuaishou-token] refresh failed app_id={row['app_id']}: {error_message}")
    return summary


_kuaishou_scheduler_started = False
_kuaishou_scheduler_lock = threading.Lock()


def _kuaishou_token_scheduler_loop() -> None:
    if KUAISHOU_TOKEN_REFRESH_STARTUP_DELAY_SECONDS > 0:
        threading.Event().wait(KUAISHOU_TOKEN_REFRESH_STARTUP_DELAY_SECONDS)
    while True:
        try:
            with app.app_context():
                summary = refresh_due_kuaishou_apps_once()
                if summary.get("refreshed") or summary.get("failed"):
                    print(f"[kuaishou-token] scheduled refresh summary: {summary}")
        except Exception as exc:
            print(f"[kuaishou-token] scheduled refresh crashed: {type(exc).__name__}: {exc}")
        threading.Event().wait(max(60, KUAISHOU_TOKEN_REFRESH_SCHEDULER_INTERVAL_SECONDS))


def start_kuaishou_token_refresh_scheduler() -> None:
    global _kuaishou_scheduler_started
    with _kuaishou_scheduler_lock:
        if _kuaishou_scheduler_started:
            return
        _kuaishou_scheduler_started = True
        thread = threading.Thread(target=_kuaishou_token_scheduler_loop, name="kuaishou-token-refresh", daemon=True)
        thread.start()

def get_minidrama_server_settings(requested_app_id: str = "") -> dict[str, Any]:
    requested = str(requested_app_id or "").strip()
    if requested:
        row = get_minidrama_app_row(requested)
        if row:
            return serialize_minidrama_app(row, include_secret=True)
        return {
            "app_id": requested,
            "app_secret": "",
            "app_secret_configured": False,
            "source": "empty",
            "updated_at": "",
            "updated_by": None,
        }
    row = get_default_minidrama_app_row()
    if row:
        payload = serialize_minidrama_app(row, include_secret=True)
        payload["apps"] = [serialize_minidrama_app(item) for item in list_minidrama_app_rows()]
        return payload
    db_app_id = get_app_setting_value(MINIDRAMA_APP_ID_SETTING_KEY)
    db_app_secret = get_app_setting_value(MINIDRAMA_APP_SECRET_SETTING_KEY)
    env_app_id = (
        str(os.environ.get("WX_MINIDRAMA_APPID") or "").strip()
        or str(os.environ.get("MINIDRAMA_APP_ID") or "").strip()
    )
    env_app_secret = (
        str(os.environ.get("WX_MINIDRAMA_APPSECRET") or "").strip()
        or str(os.environ.get("MINIDRAMA_APP_SECRET") or "").strip()
    )
    app_id = db_app_id or env_app_id
    app_secret = db_app_secret or env_app_secret
    if db_app_id or db_app_secret:
        source = "database"
    elif env_app_id or env_app_secret:
        source = "environment"
    else:
        source = "empty"
    app_id_row = get_app_setting(MINIDRAMA_APP_ID_SETTING_KEY)
    app_secret_row = get_app_setting(MINIDRAMA_APP_SECRET_SETTING_KEY)
    updated_row = app_secret_row or app_id_row
    return {
        "app_id": app_id,
        "app_secret": app_secret,
        "app_secret_configured": bool(app_secret),
        "source": source,
        "updated_at": updated_row["updated_at"] if updated_row else "",
        "updated_by": updated_row["updated_by"] if updated_row else None,
        "apps": [],
    }


def serialize_minidrama_settings(settings: dict[str, Any]) -> dict[str, Any]:
    app_secret = str(settings.get("app_secret") or "").strip()
    return {
        "app_id": str(settings.get("app_id") or "").strip(),
        "name": str(settings.get("name") or "").strip(),
        "enabled": bool(settings.get("enabled", True)),
        "is_default": bool(settings.get("is_default", False)),
        "app_secret_configured": bool(app_secret),
        "app_secret_masked": mask_secret_value(app_secret),
        "source": str(settings.get("source") or "empty"),
        "updated_at": str(settings.get("updated_at") or ""),
        "updated_by": settings.get("updated_by"),
        "apps": settings.get("apps") if isinstance(settings.get("apps"), list) else [],
    }


def resolve_minidrama_server_credentials(requested_app_id: str = "") -> tuple[str, str, str | None]:
    settings = get_minidrama_server_settings(requested_app_id)
    app_id = str(settings.get("app_id") or "").strip()
    app_secret = str(settings.get("app_secret") or "").strip()
    requested = str(requested_app_id or "").strip()
    if requested and app_id and requested != app_id:
        return "", "", "请求的小程序 AppID 与服务端配置不一致"
    if not app_id:
        return "", "", "服务端未配置小程序 AppID"
    if not app_secret:
        return "", "", f"服务端未配置小程序 AppSecret：{app_id}"
    if settings.get("enabled") is False:
        return "", "", f"服务端小程序配置已停用：{app_id}"
    return app_id, app_secret, None


def fetch_minidrama_access_token_from_weixin(app_id: str, app_secret: str) -> tuple[str, int]:
    query = urlparse.urlencode(
        {
            "grant_type": "client_credential",
            "appid": str(app_id or "").strip(),
            "secret": str(app_secret or "").strip(),
        }
    )
    req = urlrequest.Request(f"{WEIXIN_API_BASE}/cgi-bin/token?{query}", method="GET")
    with urlrequest.urlopen(req, timeout=20) as response:
        raw = response.read().decode("utf-8", errors="replace")
    payload = json.loads(raw or "{}")
    if not isinstance(payload, dict):
        raise ValueError("微信 token 接口返回格式异常")
    errcode = payload.get("errcode")
    if errcode not in (None, 0):
        raise ValueError(f"微信 token 接口失败：errcode={errcode}, errmsg={payload.get('errmsg') or ''}")
    access_token = str(payload.get("access_token") or "").strip()
    expires_in = int(payload.get("expires_in") or 0)
    if not access_token:
        raise ValueError("微信 token 接口未返回 access_token")
    return access_token, expires_in


def serialize_remote_client(row: sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "client_id": row["client_id"],
        "client_name": row["client_name"],
        "owner_user_id": row["owner_user_id"],
        "machine_id": row["machine_id"] or "",
        "device_name": row["device_name"] or "",
        "app_version": row["app_version"] or "",
        "workspace_path": row["workspace_path"] or "",
        "status": row["status"] or "offline",
        "last_seen_at": row["last_seen_at"] or "",
        "created_at": row["created_at"] or "",
        "updated_at": row["updated_at"] or "",
    }


def serialize_remote_conversation(row: sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "remote_client_id": row["remote_client_id"],
        "owner_user_id": row["owner_user_id"],
        "title": row["title"] or "",
        "status": row["status"] or "active",
        "created_at": row["created_at"] or "",
        "updated_at": row["updated_at"] or "",
    }


def serialize_remote_message(row: sqlite3.Row, attachments: list[dict] | None = None) -> dict:
    payload = None
    result = None
    try:
        payload = json.loads(row["payload_json"]) if row["payload_json"] else None
    except Exception:
        payload = None
    try:
        result = json.loads(row["result_json"]) if row["result_json"] else None
    except Exception:
        result = None
    return {
        "id": row["id"],
        "conversation_id": row["conversation_id"],
        "sender_type": row["sender_type"],
        "sender_user_id": row["sender_user_id"],
        "remote_client_id": row["remote_client_id"],
        "message_type": row["message_type"],
        "content_text": row["content_text"] or "",
        "payload": payload,
        "status": row["status"] or "pending",
        "result": result,
        "created_at": row["created_at"] or "",
        "updated_at": row["updated_at"] or "",
        "attachments": attachments or [],
    }


def serialize_remote_attachment(row: sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "message_id": row["message_id"],
        "file_type": row["file_type"],
        "original_name": row["original_name"] or "",
        "stored_path": row["stored_path"],
        "content_type": row["content_type"] or "",
        "created_at": row["created_at"] or "",
        "download_url": url_for("download_remote_attachment", attachment_id=row["id"]),
    }


def get_remote_client_by_public_id(db: sqlite3.Connection, client_id: str) -> sqlite3.Row | None:
    return db.execute(
        "SELECT * FROM remote_clients WHERE client_id = ?",
        (str(client_id or "").strip(),),
    ).fetchone()


def authenticate_remote_client(db: sqlite3.Connection, client_id: str, client_token: str) -> sqlite3.Row | None:
    row = get_remote_client_by_public_id(db, client_id)
    if not row:
        return None
    if hash_remote_client_token(client_token) != row["client_token_hash"]:
        return None
    return row


def create_remote_client_record(
    db: sqlite3.Connection,
    *,
    owner_user_id: int,
    client_name: str,
) -> tuple[sqlite3.Row, str]:
    client_id = generate_remote_client_id()
    client_token = generate_remote_client_token()
    now = now_iso()
    db.execute(
        """
        INSERT INTO remote_clients (
            client_id, client_name, client_token_hash, owner_user_id, status, created_at, updated_at
        ) VALUES (?, ?, ?, ?, 'offline', ?, ?)
        """,
        (
            client_id,
            str(client_name or "").strip() or "默认设备",
            hash_remote_client_token(client_token),
            int(owner_user_id),
            now,
            now,
        ),
    )
    db.commit()
    row = get_remote_client_by_public_id(db, client_id)
    return row, client_token


def require_remote_client() -> tuple[sqlite3.Connection, sqlite3.Row] | tuple[sqlite3.Connection, None]:
    db = get_db()
    data = request.get_json(silent=True) or {}
    client_id = request.headers.get("X-Remote-Client-Id") or data.get("client_id") or request.args.get("client_id") or ""
    client_token = request.headers.get("X-Remote-Client-Token") or data.get("client_token") or request.args.get("client_token") or ""
    row = authenticate_remote_client(db, str(client_id).strip(), str(client_token).strip())
    return db, row


def normalize_upload_record_platform(value: object) -> str:
    normalized = str(value or "").strip().lower()
    aliases = {
        "weixin_video_channel": "video_channel",
        "wechat_video_channel": "video_channel",
        "weixin_channel": "video_channel",
        "video": "video_channel",
        "wechat_miniprogram": "miniprogram",
        "weixin_miniprogram": "miniprogram",
        "mini_program": "miniprogram",
        "mini": "miniprogram",
        "ks": "kuaishou",
        "kwai": "kuaishou",
    }
    return aliases.get(normalized, normalized)


def _upload_record_text(data: dict[str, Any], *keys: str, limit: int = 500) -> str:
    for key in keys:
        value = data.get(key)
        text = str(value or "").strip()
        if text:
            return text[:limit]
    return ""


def _upload_record_int(data: dict[str, Any], *keys: str) -> int | None:
    for key in keys:
        value = data.get(key)
        if value is None or value == "":
            continue
        parsed = to_int_or_none(value)
        if parsed is not None:
            return parsed
    return None


def _upload_record_key(payload: dict[str, Any], raw: dict[str, Any]) -> str:
    explicit = str(payload.get("sync_key") or "").strip()
    if explicit:
        return explicit[:128]
    source = {
        "platform": payload.get("platform"),
        "record_time": payload.get("record_time"),
        "project_path": payload.get("project_path"),
        "original_name": payload.get("original_name"),
        "new_name": payload.get("new_name"),
        "upload_status": payload.get("upload_status"),
        "step_label": payload.get("step_label"),
        "series_id": payload.get("series_id"),
        "mini_series_id": payload.get("mini_series_id"),
        "raw": raw,
    }
    text = json.dumps(source, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _find_upload_record_drama_id(db: sqlite3.Connection, payload: dict[str, Any]) -> int | None:
    drama_id = _upload_record_int(payload, "drama_id")
    if drama_id:
        row = db.execute("SELECT id FROM dramas WHERE id = ?", (drama_id,)).fetchone()
        if row:
            return int(row["id"])
    original_name = str(payload.get("original_name") or "").strip()
    new_name = str(payload.get("new_name") or "").strip()
    if not original_name:
        return None
    if new_name:
        row = db.execute(
            "SELECT id FROM dramas WHERE original_name = ? AND new_name = ? ORDER BY id DESC LIMIT 1",
            (original_name, new_name),
        ).fetchone()
        if row:
            return int(row["id"])
    row = db.execute(
        "SELECT id FROM dramas WHERE original_name = ? ORDER BY id DESC LIMIT 1",
        (original_name,),
    ).fetchone()
    return int(row["id"]) if row else None


def _upload_record_successful(payload: dict[str, Any]) -> bool:
    status = str(payload.get("upload_status") or "").strip().lower()
    if any(token in status for token in ("成功", "完成", "已上传", "success", "done", "submitted")):
        return True
    if any(token in status for token in ("失败", "错误", "failed", "error")):
        return False
    uploaded_count = payload.get("uploaded_video_count")
    video_count = payload.get("video_file_count")
    try:
        return int(uploaded_count or 0) > 0 and int(uploaded_count or 0) >= int(video_count or 0)
    except (TypeError, ValueError):
        return False


def sanitize_upload_record_payload(data: dict[str, Any]) -> tuple[dict[str, Any], str | None]:
    raw = data.get("raw") if isinstance(data.get("raw"), dict) else {}
    merged = {**raw, **data}
    platform = normalize_upload_record_platform(merged.get("platform"))
    if platform not in UPLOAD_RECORD_PLATFORMS:
        return {}, "platform 仅支持 video_channel / miniprogram / kuaishou"

    payload: dict[str, Any] = {
        "platform": platform,
        "platform_label": _upload_record_text(merged, "platform_label", "upload_platform", limit=50)
        or UPLOAD_RECORD_PLATFORM_LABELS.get(platform, platform),
        "record_time": _upload_record_text(merged, "record_time", "created_at", "time", limit=30),
        "date": _upload_record_text(merged, "date", limit=20),
        "upload_status": _upload_record_text(merged, "upload_status", "status", limit=100),
        "execution_mode": _upload_record_text(merged, "execution_mode", "mode", limit=100),
        "step_label": _upload_record_text(merged, "step_label", "step", limit=200),
        "project_name": _upload_record_text(merged, "project_name", "title", limit=200),
        "project_path": _upload_record_text(merged, "project_path", "path", limit=500),
        "original_name": _upload_record_text(merged, "original_name", "original_title", limit=200),
        "new_name": _upload_record_text(merged, "new_name", "new_title", limit=200),
        "episodes": _upload_record_int(merged, "episodes", "episode_count"),
        "video_file_count": _upload_record_int(merged, "video_file_count", "video_count"),
        "uploaded_video_count": _upload_record_int(merged, "uploaded_video_count", "uploaded_count"),
        "uploader_display": _upload_record_text(merged, "uploader_display", "uploader", "channel_nickname", limit=200),
        "account_profile_id": _upload_record_text(merged, "account_profile_id", limit=100),
        "account_profile_name": _upload_record_text(merged, "account_profile_name", "account_profile", limit=200),
        "device_name": _upload_record_text(merged, "device_name", limit=200),
        "failure_reason": _upload_record_text(merged, "failure_reason", "error_message", "error", limit=1000),
        "extra_info": _upload_record_text(merged, "extra_info", "details", limit=2000),
        "series_id": _upload_record_text(merged, "series_id", limit=100),
        "mini_series_id": _upload_record_text(merged, "mini_series_id", limit=100),
        "audit_status": _upload_record_text(merged, "audit_status", "audit_status_text", limit=100),
        "selling_status": _upload_record_text(merged, "selling_status", "selling_status_text", limit=100),
        "audit_reject_reason": _upload_record_text(merged, "audit_reject_reason", "reject_reason", "audit_reason", limit=1000),
        "audit_reject_detail": _upload_record_text(merged, "audit_reject_detail", "reject_reason_detail", "audit_reason_detail", limit=2000),
        "online_status": _upload_record_text(merged, "online_status", "listing_status", limit=100),
        "online_at": _upload_record_text(merged, "online_at", "listing_at", limit=50),
        "distribution_status": _upload_record_text(merged, "distribution_status", limit=100),
        "distribution_at": _upload_record_text(merged, "distribution_at", limit=50),
        "distribution_detail": _upload_record_text(merged, "distribution_detail", "distribution_error", limit=2000),
        "submitted_at": _upload_record_text(merged, "submitted_at", limit=50),
        "raw_json": json.dumps(raw or data, ensure_ascii=False, sort_keys=True),
    }
    payload["sync_key"] = _upload_record_text(merged, "sync_key", limit=128) or _upload_record_key(payload, raw or data)
    if not payload["record_time"]:
        payload["record_time"] = now_iso()
    if not payload["date"]:
        payload["date"] = str(payload["record_time"])[:10]
    if not payload["original_name"] and not payload["new_name"] and not payload["project_name"]:
        return {}, "记录缺少剧名或项目名称"
    return payload, None


def upsert_upload_record(
    db: sqlite3.Connection,
    *,
    owner_user_id: int,
    owner_username: str,
    remote_client_id: int | None,
    data: dict[str, Any],
) -> tuple[dict[str, Any] | None, str | None, bool]:
    payload, error = sanitize_upload_record_payload(data)
    if error:
        return None, error, False
    payload["owner_user_id"] = int(owner_user_id)
    payload["owner_username"] = str(owner_username or "").strip()
    payload["remote_client_id"] = remote_client_id
    payload["drama_id"] = _find_upload_record_drama_id(db, payload)

    now = now_iso()
    existing = db.execute(
        "SELECT id FROM upload_records WHERE owner_user_id = ? AND platform = ? AND sync_key = ?",
        (payload["owner_user_id"], payload["platform"], payload["sync_key"]),
    ).fetchone()
    payload["updated_at"] = now
    if existing:
        payload["id"] = int(existing["id"])
        set_columns = [
            key
            for key in payload.keys()
            if key not in {"id", "owner_user_id", "platform", "sync_key"}
        ]
        set_clause = ", ".join(f"{key} = :{key}" for key in set_columns)
        db.execute(
            f"UPDATE upload_records SET {set_clause} WHERE id = :id",
            payload,
        )
        created = False
        record_id = int(existing["id"])
    else:
        payload["created_at"] = now
        columns = list(payload.keys())
        placeholders = ", ".join(f":{key}" for key in columns)
        db.execute(
            f"INSERT INTO upload_records ({', '.join(columns)}) VALUES ({placeholders})",
            payload,
        )
        record_id = int(db.execute("SELECT last_insert_rowid()").fetchone()[0])
        created = True

    row = db.execute("SELECT * FROM upload_records WHERE id = ?", (record_id,)).fetchone()
    if row:
        refresh_drama_platform_status(db, row)
    return (serialize_upload_record(row) if row else None), None, created


def refresh_drama_platform_status(db: sqlite3.Connection, row: sqlite3.Row) -> None:
    drama_id = row["drama_id"]
    if not drama_id:
        return
    now = now_iso()
    db.execute(
        """
        INSERT INTO drama_platform_status (
            drama_id, platform, platform_label, latest_status, uploaded_video_count,
            video_file_count, last_record_id, external_series_id, audit_status,
            selling_status, audit_reject_reason, online_status, distribution_status,
            updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(drama_id, platform) DO UPDATE SET
            platform_label = excluded.platform_label,
            latest_status = excluded.latest_status,
            uploaded_video_count = excluded.uploaded_video_count,
            video_file_count = excluded.video_file_count,
            last_record_id = excluded.last_record_id,
            external_series_id = excluded.external_series_id,
            audit_status = excluded.audit_status,
            selling_status = excluded.selling_status,
            audit_reject_reason = excluded.audit_reject_reason,
            online_status = excluded.online_status,
            distribution_status = excluded.distribution_status,
            updated_at = excluded.updated_at
        """,
        (
            int(drama_id),
            row["platform"],
            row["platform_label"],
            row["upload_status"],
            row["uploaded_video_count"],
            row["video_file_count"],
            int(row["id"]),
            row["series_id"] or row["mini_series_id"],
            row["audit_status"],
            row["selling_status"],
            row["audit_reject_reason"],
            row["online_status"],
            row["distribution_status"],
            now,
        ),
    )
    if _upload_record_successful(row_to_dict(row)):
        uploader = row["uploader_display"] or row["owner_username"]
        db.execute(
            "UPDATE dramas SET uploaded = '是', uploader = ? WHERE id = ?",
            (uploader, int(drama_id)),
        )


def serialize_upload_record(row: sqlite3.Row | None) -> dict[str, Any]:
    if row is None:
        return {}
    payload = row_to_dict(row)
    payload["platform_label"] = payload.get("platform_label") or UPLOAD_RECORD_PLATFORM_LABELS.get(
        str(payload.get("platform") or ""),
        str(payload.get("platform") or ""),
    )
    return payload


def build_remote_command_summary(command: str, payload: dict[str, Any]) -> str:
    if command == REMOTE_COMMAND_KUAISHOU_START_QUEUE:
        return "快手执行队列"
    if command == REMOTE_COMMAND_KUAISHOU_STOP_QUEUE:
        return "快手停止队列"
    if command == REMOTE_COMMAND_KUAISHOU_QUERY_STATUS:
        return "快手查询状态"
    titles = payload.get("titles") if isinstance(payload.get("titles"), list) else []
    normalized_titles = [str(item).strip() for item in titles if str(item).strip()]
    if command == REMOTE_COMMAND_KUAISHOU_UPLOAD_SERIES:
        if not normalized_titles:
            return "快手上传短剧"
        preview = "、".join(normalized_titles[:3])
        if len(normalized_titles) > 3:
            preview += f" 等 {len(normalized_titles)} 部"
        return f"快手上传短剧：{preview}"
    if command != REMOTE_COMMAND_IMPORT_DRAMA_TITLES:
        return command or "远程命令"
    if not normalized_titles:
        return "导入短剧"
    preview = "、".join(normalized_titles[:3])
    if len(normalized_titles) > 3:
        preview += f" 等 {len(normalized_titles)} 部"
    return f"导入短剧：{preview}"


def sanitize_remote_command_payload(message_type: str, data: dict[str, Any]) -> tuple[dict[str, Any] | None, str | None, str]:
    content_text = str(data.get("content_text") or "").strip()
    payload = data.get("payload")
    if message_type != "command":
        if payload is None:
            return None, None, content_text
        if not isinstance(payload, dict):
            return None, "payload 必须是对象", content_text
        return payload, None, content_text

    if not isinstance(payload, dict):
        return None, "command 消息缺少 payload 对象", content_text

    command = str(payload.get("command") or "").strip().lower()
    if command not in REMOTE_COMMAND_ALLOWED_COMMANDS:
        return None, "仅支持 import_drama_titles / ks_upload_series / ks_start_queue / ks_stop_queue / ks_query_status 命令", content_text

    if command in {
        REMOTE_COMMAND_KUAISHOU_START_QUEUE,
        REMOTE_COMMAND_KUAISHOU_STOP_QUEUE,
        REMOTE_COMMAND_KUAISHOU_QUERY_STATUS,
    }:
        normalized_payload = {"command": command}
        if not content_text:
            content_text = build_remote_command_summary(command, normalized_payload)
        return normalized_payload, None, content_text

    raw_titles = payload.get("titles")
    if not isinstance(raw_titles, list):
        return None, "titles 必须是数组", content_text
    titles: list[str] = []
    seen_titles: set[str] = set()
    for raw_item in raw_titles:
        title = str(raw_item or "").strip()
        if not title or title in seen_titles:
            continue
        seen_titles.add(title)
        titles.append(title)
    if not titles:
        return None, "titles 不能为空", content_text

    if command == REMOTE_COMMAND_KUAISHOU_UPLOAD_SERIES:
        normalized_payload = {
            "command": REMOTE_COMMAND_KUAISHOU_UPLOAD_SERIES,
            "titles": titles,
            "skip_submitted": bool(payload.get("skip_submitted", True)),
            "auto_download": bool(payload.get("auto_download", True)),
        }
        if not content_text:
            content_text = build_remote_command_summary(REMOTE_COMMAND_KUAISHOU_UPLOAD_SERIES, normalized_payload)
        return normalized_payload, None, content_text

    workspace_path = str(payload.get("workspace_path") or "").strip()
    sync_download = bool(payload.get("sync_download", True))
    auto_run = bool(payload.get("auto_run", True))

    enabled_steps_value = payload.get("enabled_steps")
    enabled_steps: list[str] | None = None
    if enabled_steps_value is not None:
        if not isinstance(enabled_steps_value, list):
            return None, "enabled_steps 必须是数组", content_text
        enabled_steps = []
        seen_steps: set[str] = set()
        for raw_step in enabled_steps_value:
            step_key = str(raw_step or "").strip()
            if not step_key or step_key in seen_steps:
                continue
            if step_key not in REMOTE_IMPORT_DRAMA_ALLOWED_STEPS:
                return None, f"enabled_steps 包含不支持的步骤: {step_key}", content_text
            seen_steps.add(step_key)
            enabled_steps.append(step_key)

    on_project_error = str(payload.get("on_project_error") or "").strip().lower()
    if on_project_error and on_project_error not in REMOTE_IMPORT_DRAMA_ALLOWED_ERROR_STRATEGIES:
        return None, "on_project_error 仅支持 skip 或 stop", content_text

    parallel_projects = payload.get("parallel_projects")
    normalized_parallel_projects: int | None = None
    if parallel_projects not in (None, ""):
        try:
            normalized_parallel_projects = max(1, min(4, int(parallel_projects)))
        except (TypeError, ValueError):
            return None, "parallel_projects 必须是 1 到 4 之间的整数", content_text

    normalized_payload = {
        "command": REMOTE_COMMAND_IMPORT_DRAMA_TITLES,
        "titles": titles,
        "workspace_path": workspace_path,
        "sync_download": sync_download,
        "auto_run": auto_run,
    }
    if enabled_steps is not None:
        normalized_payload["enabled_steps"] = enabled_steps
    if on_project_error:
        normalized_payload["on_project_error"] = on_project_error
    if normalized_parallel_projects is not None:
        normalized_payload["parallel_projects"] = normalized_parallel_projects

    if not content_text:
        content_text = build_remote_command_summary(REMOTE_COMMAND_IMPORT_DRAMA_TITLES, normalized_payload)
    return normalized_payload, None, content_text


def ensure_remote_conversation_access(db: sqlite3.Connection, conversation_id: int, user_id: int, role: str) -> sqlite3.Row | None:
    row = db.execute(
        """
        SELECT rc.owner_user_id, c.*
        FROM remote_conversations c
        JOIN remote_clients rc ON rc.id = c.remote_client_id
        WHERE c.id = ?
        """,
        (conversation_id,),
    ).fetchone()
    if not row:
        return None
    if role != "admin" and int(row["owner_user_id"]) != int(user_id):
        return None
    return row


def get_or_create_remote_conversation(db: sqlite3.Connection, remote_client_row: sqlite3.Row, *, title: str = "") -> sqlite3.Row:
    row = db.execute(
        """
        SELECT *
        FROM remote_conversations
        WHERE remote_client_id = ? AND status = 'active'
        ORDER BY updated_at DESC, id DESC
        LIMIT 1
        """,
        (remote_client_row["id"],),
    ).fetchone()
    if row:
        return row
    db.execute(
        """
        INSERT INTO remote_conversations (remote_client_id, owner_user_id, title, status, created_at, updated_at)
        VALUES (?, ?, ?, 'active', ?, ?)
        """,
        (
            remote_client_row["id"],
            remote_client_row["owner_user_id"],
            title or f"{remote_client_row['client_name']} 会话",
            now_iso(),
            now_iso(),
        ),
    )
    db.commit()
    return db.execute(
        "SELECT * FROM remote_conversations WHERE id = last_insert_rowid()"
    ).fetchone()


def validate_client_license_payload(data: dict) -> tuple[dict, str | None]:
    payload = {
        "license_key": str(data.get("license_key") or "").strip(),
        "machine_id": str(data.get("machine_id") or "").strip(),
        "app_name": str(data.get("app_name") or "").strip(),
        "app_version": str(data.get("app_version") or "").strip(),
        "token": str(data.get("token") or "").strip(),
    }
    if not payload["license_key"]:
        return payload, "激活码不能为空"
    if not payload["machine_id"]:
        return payload, "机器码不能为空"
    return payload, None


def validate_account_auth_payload(data: dict, *, require_registration: bool = False) -> tuple[dict, str | None]:
    payload = {
        "account": str(data.get("account") or data.get("username") or data.get("email") or "").strip(),
        "username": str(data.get("username") or "").strip(),
        "email": normalize_email(str(data.get("email") or "")),
        "password": str(data.get("password") or ""),
        "machine_id": str(data.get("machine_id") or "").strip(),
        "device_name": str(data.get("device_name") or "").strip(),
        "app_name": str(data.get("app_name") or "").strip(),
        "app_version": str(data.get("app_version") or "").strip(),
        "token": str(data.get("token") or "").strip(),
        # 设备数达到上限时默认拒绝新设备登录，只有客户端显式传 force_login=true
        # 才允许顶掉旧设备，避免 max_devices=1 时被静默替换。
        "force_login": parse_json_bool(data.get("force_login"), default=False),
    }
    if require_registration:
        if not USERNAME_RE.match(payload["username"]):
            return payload, "用户名需为 2-30 位字母、数字或下划线，或使用有效邮箱格式"
        if not EMAIL_RE.match(payload["email"]):
            return payload, "请输入有效邮箱"
    elif not payload["account"]:
        return payload, "请输入用户名或邮箱"
    if require_registration or payload["password"]:
        if len(payload["password"]) < 6:
            return payload, "密码至少需要 6 位"
    if not payload["machine_id"]:
        return payload, "机器码不能为空"
    return payload, None


def parse_json_bool(value: object, *, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    text = str(value).strip().lower()
    if not text:
        return default
    if text in {"1", "true", "yes", "y", "on"}:
        return True
    if text in {"0", "false", "no", "n", "off"}:
        return False
    return default


def build_account_auth_response(
    user_row: sqlite3.Row,
    *,
    machine_id: str,
    token: str,
    logged_in_at: str | None = None,
    last_verified_at: str | None = None,
    replaced_device_count: int = 0,
) -> dict:
    now_iso = datetime.datetime.now().isoformat(timespec="seconds")
    offline_grace_until = (
        datetime.datetime.now() + datetime.timedelta(hours=ACCOUNT_OFFLINE_GRACE_HOURS)
    ).isoformat(timespec="seconds")
    username = str(user_row["username"] or "")
    return {
        "username": username,
        "account_username": username,
        "email": user_row["email"] or "",
        "license_key_masked": username,
        "machine_id": machine_id,
        "token": token,
        "activated_at": logged_in_at or last_verified_at or now_iso,
        "last_verified_at": last_verified_at or now_iso,
        "offline_grace_until": offline_grace_until,
        "expires_at": user_row["expires_at"] or "",
        "edition": user_row["edition"],
        "licensee": username,
        "max_devices": int(user_row["max_devices"] or ACCOUNT_DEFAULT_MAX_DEVICES),
        "replaced_device_count": int(replaced_device_count or 0),
    }


def ensure_account_can_login(user_row: sqlite3.Row) -> tuple[bool, str]:
    if str(user_row["status"] or "active") != "active":
        return False, "账号已停用"
    if is_user_account_expired(user_row):
        return False, "账号已过期"
    return True, ""


def activate_account_for_machine(
    db: sqlite3.Connection,
    *,
    user_row: sqlite3.Row,
    machine_id: str,
    device_name: str,
    app_name: str,
    app_version: str,
    force_login: bool = False,
) -> dict:
    ok, error = ensure_account_can_login(user_row)
    if not ok:
        raise ValueError(error)

    device_row = db.execute(
        """
        SELECT *
        FROM user_devices
        WHERE user_id = ? AND machine_id = ?
        """,
        (user_row["id"], machine_id),
    ).fetchone()

    token = issue_account_token(user_row=user_row, machine_id=machine_id)
    token_hash = hash_token(token)
    now_iso = datetime.datetime.now().isoformat(timespec="seconds")
    replaced_device_count = 0

    active_current_row = db.execute(
        """
        SELECT id
        FROM user_devices
        WHERE user_id = ? AND machine_id = ? AND (revoked_at IS NULL OR revoked_at = '')
        """,
        (user_row["id"], machine_id),
    ).fetchone()
    if active_current_row is None:
        active_count = current_active_user_device_count(db, user_row["id"])
        max_devices = max(1, int(user_row["max_devices"] or ACCOUNT_DEFAULT_MAX_DEVICES))
        overflow_count = active_count - max_devices + 1
        if overflow_count > 0:
            if not force_login:
                raise ValueError("账号已在其他电脑登录")
            old_rows = db.execute(
                """
                SELECT id
                FROM user_devices
                WHERE user_id = ?
                  AND machine_id != ?
                  AND (revoked_at IS NULL OR revoked_at = '')
                ORDER BY COALESCE(last_verified_at, logged_in_at, '') ASC, id ASC
                LIMIT ?
                """,
                (user_row["id"], machine_id, overflow_count),
            ).fetchall()
            for old_row in old_rows:
                db.execute(
                    """
                    UPDATE user_devices
                    SET revoked_at = ?
                    WHERE id = ? AND (revoked_at IS NULL OR revoked_at = '')
                    """,
                    (now_iso, old_row["id"]),
                )
            replaced_device_count = len(old_rows)

    if device_row:
        db.execute(
            """
            UPDATE user_devices
            SET token_hash = ?, device_name = ?, app_name = ?, app_version = ?,
                last_verified_at = ?, revoked_at = NULL
            WHERE id = ?
            """,
            (
                token_hash,
                device_name or None,
                app_name or None,
                app_version or None,
                now_iso,
                device_row["id"],
            ),
        )
    else:
        db.execute(
            """
            INSERT INTO user_devices (
                user_id, machine_id, device_name, app_name, app_version,
                token_hash, logged_in_at, last_verified_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                user_row["id"],
                machine_id,
                device_name or None,
                app_name or None,
                app_version or None,
                token_hash,
                now_iso,
                now_iso,
            ),
        )

    db.execute(
        "UPDATE users SET updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (user_row["id"],),
    )
    db.commit()
    return build_account_auth_response(
        user_row,
        machine_id=machine_id,
        token=token,
        logged_in_at=(
            str(device_row["logged_in_at"])
            if device_row and device_row["logged_in_at"]
            else now_iso
        ),
        last_verified_at=now_iso,
        replaced_device_count=replaced_device_count,
    )


def build_tt_account_auth_response(
    user_row: sqlite3.Row,
    *,
    machine_id: str,
    token: str,
    logged_in_at: str | None = None,
    last_verified_at: str | None = None,
    replaced_device_count: int = 0,
) -> dict:
    now_value = datetime.datetime.now()
    now_iso_value = now_value.isoformat(timespec="seconds")
    offline_grace_until = (
        now_value + datetime.timedelta(hours=ACCOUNT_OFFLINE_GRACE_HOURS)
    ).isoformat(timespec="seconds")
    username = str(user_row["username"] or "")
    return {
        "username": username,
        "account_username": username,
        "email": user_row["email"] or "",
        "license_key_masked": username,
        "machine_id": machine_id,
        "token": token,
        "activated_at": logged_in_at or last_verified_at or now_iso_value,
        "last_verified_at": last_verified_at or now_iso_value,
        "offline_grace_until": offline_grace_until,
        "expires_at": user_row["expires_at"] or "",
        "edition": user_row["edition"],
        "licensee": username,
        "max_devices": int(user_row["max_devices"] or ACCOUNT_DEFAULT_MAX_DEVICES),
        "replaced_device_count": int(replaced_device_count or 0),
    }


def ensure_tt_account_can_login(user_row: sqlite3.Row) -> tuple[bool, str]:
    if str(user_row["status"] or "active") != "active":
        return False, "TT账号已停用"
    if is_user_account_expired(user_row):
        return False, "TT账号已过期"
    return True, ""


def activate_tt_account_for_machine(
    db: sqlite3.Connection,
    *,
    user_row: sqlite3.Row,
    machine_id: str,
    device_name: str,
    app_name: str,
    app_version: str,
    force_login: bool = False,
) -> dict:
    ok, error = ensure_tt_account_can_login(user_row)
    if not ok:
        raise ValueError(error)

    device_row = db.execute(
        """
        SELECT *
        FROM tt_user_devices
        WHERE tt_user_id = ? AND machine_id = ?
        """,
        (user_row["id"], machine_id),
    ).fetchone()

    token = issue_tt_account_token(user_row=user_row, machine_id=machine_id)
    token_hash = hash_token(token)
    now_value = now_iso()
    replaced_device_count = 0

    active_current_row = db.execute(
        """
        SELECT id
        FROM tt_user_devices
        WHERE tt_user_id = ? AND machine_id = ? AND (revoked_at IS NULL OR revoked_at = '')
        """,
        (user_row["id"], machine_id),
    ).fetchone()
    if active_current_row is not None:
        ok, device_limit_error = ensure_tt_account_device_limit_for_machine(
            db,
            user_row,
            machine_id=machine_id,
        )
        if not ok:
            raise ValueError(device_limit_error)
    if active_current_row is None:
        active_count = current_active_tt_user_device_count(db, user_row["id"])
        max_devices = max(1, int(user_row["max_devices"] or ACCOUNT_DEFAULT_MAX_DEVICES))
        overflow_count = active_count - max_devices + 1
        if overflow_count > 0:
            if not force_login:
                raise ValueError("TT账号已在其他电脑登录")
            old_rows = db.execute(
                """
                SELECT id
                FROM tt_user_devices
                WHERE tt_user_id = ?
                  AND machine_id != ?
                  AND (revoked_at IS NULL OR revoked_at = '')
                ORDER BY COALESCE(last_verified_at, logged_in_at, '') ASC, id ASC
                LIMIT ?
                """,
                (user_row["id"], machine_id, overflow_count),
            ).fetchall()
            for old_row in old_rows:
                db.execute(
                    """
                    UPDATE tt_user_devices
                    SET revoked_at = ?
                    WHERE id = ? AND (revoked_at IS NULL OR revoked_at = '')
                    """,
                    (now_value, old_row["id"]),
                )
            replaced_device_count = len(old_rows)

    if device_row:
        db.execute(
            """
            UPDATE tt_user_devices
            SET token_hash = ?, device_name = ?, app_name = ?, app_version = ?,
                last_verified_at = ?, revoked_at = NULL
            WHERE id = ?
            """,
            (
                token_hash,
                device_name or None,
                app_name or None,
                app_version or None,
                now_value,
                device_row["id"],
            ),
        )
    else:
        db.execute(
            """
            INSERT INTO tt_user_devices (
                tt_user_id, machine_id, device_name, app_name, app_version,
                token_hash, logged_in_at, last_verified_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                user_row["id"],
                machine_id,
                device_name or None,
                app_name or None,
                app_version or None,
                token_hash,
                now_value,
                now_value,
            ),
        )

    db.execute(
        "UPDATE tt_users SET updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (user_row["id"],),
    )
    db.commit()
    return build_tt_account_auth_response(
        user_row,
        machine_id=machine_id,
        token=token,
        logged_in_at=(
            str(device_row["logged_in_at"])
            if device_row and device_row["logged_in_at"]
            else now_value
        ),
        last_verified_at=now_value,
        replaced_device_count=replaced_device_count,
    )


def build_client_license_response(
    license_row: sqlite3.Row,
    *,
    machine_id: str,
    token: str,
    activated_at: str | None = None,
    last_verified_at: str | None = None,
) -> dict:
    now_iso = datetime.datetime.now().isoformat(timespec="seconds")
    return {
        "license_key_masked": license_row["license_key_masked"],
        "machine_id": machine_id,
        "token": token,
        "activated_at": activated_at or last_verified_at or now_iso,
        "last_verified_at": last_verified_at or now_iso,
        "expires_at": license_row["expires_at"] or "",
        "edition": license_row["edition"],
        "licensee": license_row["licensee"] or "",
    }


def activate_license_for_machine(
    db: sqlite3.Connection,
    *,
    license_row: sqlite3.Row,
    machine_id: str,
    app_name: str,
    app_version: str,
) -> dict:
    active_row = db.execute(
        """
        SELECT *
        FROM license_activations
        WHERE license_id = ? AND machine_id = ? AND (revoked_at IS NULL OR revoked_at = '')
        """,
        (license_row["id"], machine_id),
    ).fetchone()

    if not active_row:
        active_count = current_active_activation_count(db, license_row["id"])
        if active_count >= int(license_row["max_activations"] or 1):
            raise ValueError("该激活码已达到最大设备绑定数量")

    token = issue_license_token(license_row=license_row, machine_id=machine_id)
    token_hash = hash_token(token)
    now_iso = datetime.datetime.now().isoformat(timespec="seconds")

    if active_row:
        db.execute(
            """
            UPDATE license_activations
            SET token_hash = ?, app_name = ?, app_version = ?, last_verified_at = ?, revoked_at = NULL
            WHERE id = ?
            """,
            (token_hash, app_name or None, app_version or None, now_iso, active_row["id"]),
        )
    else:
        db.execute(
            """
            INSERT INTO license_activations (
                license_id, machine_id, app_name, app_version, token_hash, activated_at, last_verified_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (license_row["id"], machine_id, app_name or None, app_version or None, token_hash, now_iso, now_iso),
        )

    db.execute(
        "UPDATE licenses SET updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (license_row["id"],),
    )
    db.commit()
    return build_client_license_response(
        license_row,
        machine_id=machine_id,
        token=token,
        activated_at=(
            str(active_row["activated_at"])
            if active_row and active_row["activated_at"]
            else now_iso
        ),
        last_verified_at=now_iso,
    )


def sanitize_license_payload(data: dict) -> tuple[dict, str | None]:
    license_key = str(data.get("license_key") or "").strip().upper()
    edition = str(data.get("edition") or "pro").strip().lower()
    status = str(data.get("status") or "active").strip().lower()
    licensee = str(data.get("licensee") or "").strip()
    notes = str(data.get("notes") or "").strip()
    expires_at = str(data.get("expires_at") or "").strip()
    try:
        max_activations = int(data.get("max_activations") or 1)
    except (TypeError, ValueError):
        return {}, "最大激活数必须是正整数"
    if max_activations < 1:
        return {}, "最大激活数必须是正整数"
    if edition not in LICENSE_EDITION_VALUES:
        edition = "pro"
    if status not in LICENSE_STATUS_VALUES:
        status = "active"
    if expires_at:
        try:
            datetime.datetime.fromisoformat(expires_at)
        except ValueError:
            return {}, "到期时间格式不正确，请使用 YYYY-MM-DD 或 ISO 日期时间"
    if not license_key:
        license_key = generate_license_key()
    return {
        "license_key": license_key,
        "license_key_masked": mask_license_key(license_key),
        "status": status,
        "edition": edition,
        "licensee": licensee or None,
        "max_activations": max_activations,
        "expires_at": expires_at or None,
        "notes": notes or None,
    }, None


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        db = get_db()
        user = db.execute(
            """
            SELECT id, username, password_hash, role
            FROM users
            WHERE username = ? OR lower(COALESCE(email, '')) = lower(?)
            """,
            (username, username),
        ).fetchone()
        if user and check_password_hash(user["password_hash"], password):
            session.clear()
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["role"] = user["role"]
            return redirect(url_for("index"))
        error = "用户名或密码错误"
    return render_template("login.html", error=error)


@app.route("/logout")
@login_required
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def index():
    return redirect(url_for("video_channel_dramas_page"))


@app.route("/dramas/video-channel")
@login_required
def video_channel_dramas_page():
    return render_template("video_channel_dramas.html")


@app.route("/dramas/miniprogram")
@login_required
def miniprogram_dramas_page():
    return render_template("miniprogram_dramas.html")


@app.route("/dramas/kuaishou")
@login_required
def kuaishou_dramas_page():
    return render_template("kuaishou_dramas.html")


@app.route("/monitor")
@login_required
@admin_required
def monitor_dashboard():
    return render_template("monitor.html")


@app.route("/upload-records")
@login_required
@admin_required
def upload_records_page():
    return render_template("upload_records.html")


@app.route("/licenses")
@login_required
@admin_required
def license_management():
    return render_template("licenses.html")


@app.route("/users")
@login_required
@admin_required
def user_management():
    return render_template("users.html")


@app.route("/tt-users")
@login_required
@admin_required
def tt_user_management():
    return render_template("users.html", user_page_variant="tt_users")


@app.route("/settings/minidrama")
@login_required
@admin_required
def minidrama_settings_page():
    return render_template("minidrama_settings.html")


@app.route("/settings/kuaishou")
@login_required
@admin_required
def kuaishou_settings_page():
    return render_template("kuaishou_settings.html")


@app.route("/remote")
@login_required
@admin_required
def remote_management():
    return render_template("remote.html")


@app.route("/api/me", methods=["GET"])
@login_required
def api_me():
    return jsonify(
        {
            "user_id": session.get("user_id"),
            "username": session.get("username", ""),
            "role": session.get("role", "user"),
        }
    )


@app.route("/api/settings/kuaishou", methods=["GET"])
@login_required
@admin_required
def get_kuaishou_settings_api():
    return jsonify(serialize_kuaishou_settings(get_kuaishou_server_settings()))


@app.route("/api/settings/kuaishou", methods=["PUT"])
@login_required
@admin_required
def update_kuaishou_settings_api():
    data = request.get_json(silent=True) or {}
    app_id = str(data.get("app_id") or "").strip()
    app_secret = str(data.get("app_secret") or "").strip()
    if not app_id:
        return jsonify({"error": "快手 AppID 不能为空"}), 400
    if len(app_id) > 64:
        return jsonify({"error": "快手 AppID 过长"}), 400
    existing = get_kuaishou_server_settings(app_id)
    if not app_secret:
        app_secret = str(existing.get("app_secret") or "").strip()
    if not app_secret:
        return jsonify({"error": "快手 AppSecret 不能为空"}), 400
    advertiser_id = str(data.get("advertiser_id") or existing.get("advertiser_id") or "").strip()
    access_token = str(data.get("access_token") or existing.get("access_token") or "").strip()
    refresh_token = str(data.get("refresh_token") or existing.get("refresh_token") or "").strip()
    access_token_expires_at = int(data.get("access_token_expires_at") or existing.get("access_token_expires_at") or 0)
    refresh_token_expires_at = int(data.get("refresh_token_expires_at") or existing.get("refresh_token_expires_at") or 0)
    saved_row = save_kuaishou_app(
        app_id=app_id,
        app_secret=app_secret,
        advertiser_id=advertiser_id,
        name=str(data.get("name") or existing.get("name") or "").strip(),
        access_token=access_token,
        refresh_token=refresh_token,
        access_token_expires_at=access_token_expires_at,
        refresh_token_expires_at=refresh_token_expires_at,
        enabled=bool(data.get("enabled", existing.get("enabled", True))),
        is_default=bool(data.get("is_default")),
        updated_by=session.get("user_id"),
    )
    get_db().commit()
    payload = serialize_kuaishou_settings(get_kuaishou_server_settings())
    payload["saved"] = serialize_kuaishou_app(saved_row)
    return jsonify(payload)


@app.route("/api/settings/kuaishou/<path:app_id>", methods=["DELETE"])
@login_required
@admin_required
def delete_kuaishou_settings_api(app_id: str):
    normalized = str(app_id or "").strip()
    row = get_kuaishou_app_row(normalized)
    if not row:
        return jsonify({"error": "快手配置不存在"}), 404
    db = get_db()
    db.execute("DELETE FROM kuaishou_apps WHERE app_id = ?", (normalized,))
    db.execute("DELETE FROM kuaishou_token_cache WHERE app_id = ?", (normalized,))
    if not get_default_kuaishou_app_row():
        db.execute("UPDATE kuaishou_apps SET is_default = 1 WHERE id = (SELECT id FROM kuaishou_apps WHERE enabled = 1 ORDER BY updated_at DESC, id DESC LIMIT 1)")
    db.commit()
    return jsonify(serialize_kuaishou_settings(get_kuaishou_server_settings()))


@app.route("/api/settings/minidrama", methods=["GET"])
@login_required
@admin_required
def get_minidrama_settings_api():
    return jsonify(serialize_minidrama_settings(get_minidrama_server_settings()))


@app.route("/api/settings/minidrama", methods=["PUT"])
@login_required
@admin_required
def update_minidrama_settings_api():
    data = request.get_json(silent=True) or {}
    app_id = str(data.get("app_id") or "").strip()
    app_secret_input = str(data.get("app_secret") or "").strip()
    name = str(data.get("name") or "").strip()
    enabled = bool(data.get("enabled", True))
    is_default = bool(data.get("is_default"))
    clear_app_secret = bool(data.get("clear_app_secret"))

    if not app_id:
        return jsonify({"error": "小程序 AppID 不能为空"}), 400
    if len(app_id) > 64:
        return jsonify({"error": "小程序 AppID 过长"}), 400

    existing_settings = get_minidrama_server_settings(app_id)
    existing_secret = str(existing_settings.get("app_secret") or "").strip()
    app_secret = "" if clear_app_secret else (app_secret_input or existing_secret)
    if not app_secret:
        return jsonify({"error": "小程序 AppSecret 不能为空"}), 400
    if len(app_secret) > 256:
        return jsonify({"error": "小程序 AppSecret 过长"}), 400

    db = get_db()
    user_id = session.get("user_id")
    saved_row = save_minidrama_app(
        app_id=app_id,
        app_secret=app_secret,
        name=name,
        enabled=enabled,
        is_default=is_default,
        updated_by=user_id,
    )
    db.commit()

    payload = serialize_minidrama_settings(get_minidrama_server_settings())
    payload["saved"] = serialize_minidrama_app(saved_row)
    return jsonify(payload)


@app.route("/api/settings/minidrama/<path:app_id>", methods=["DELETE"])
@login_required
@admin_required
def delete_minidrama_settings_api(app_id: str):
    normalized = str(app_id or "").strip()
    row = get_minidrama_app_row(normalized)
    if not row:
        return jsonify({"error": "小程序配置不存在"}), 404
    db = get_db()
    db.execute("DELETE FROM minidrama_apps WHERE app_id = ?", (normalized,))
    db.execute("DELETE FROM minidrama_token_cache WHERE app_id = ?", (normalized,))
    if not get_default_minidrama_app_row():
        db.execute(
            """
            UPDATE minidrama_apps
            SET is_default = 1
            WHERE id = (
                SELECT id FROM minidrama_apps WHERE enabled = 1 ORDER BY updated_at DESC, id DESC LIMIT 1
            )
            """
        )
    db.commit()
    return jsonify(serialize_minidrama_settings(get_minidrama_server_settings()))


@app.route("/client-api/account/register", methods=["POST"])
@app.route("/account/register", methods=["POST"])
def client_register_account():
    data = request.get_json(silent=True) or {}
    payload, error = validate_account_auth_payload(data, require_registration=True)
    if error:
        return jsonify({"ok": False, "message": error}), 400

    db = get_db()
    conflict_message = _find_existing_registration_conflict(
        db,
        username=payload["username"],
        email=payload["email"],
    )
    if conflict_message:
        return jsonify({"ok": False, "message": conflict_message}), 400

    try:
        db.execute(
            """
            INSERT INTO users (
                username, email, password_hash, role, status, max_devices, edition, updated_at
            ) VALUES (?, ?, ?, 'user', 'active', ?, 'pro', CURRENT_TIMESTAMP)
            """,
            (
                payload["username"],
                payload["email"],
                generate_password_hash(payload["password"]),
                ACCOUNT_DEFAULT_MAX_DEVICES,
            ),
        )
        db.commit()
    except sqlite3.IntegrityError:
        conflict_message = _find_existing_registration_conflict(
            db,
            username=payload["username"],
            email=payload["email"],
        )
        return jsonify({"ok": False, "message": conflict_message or "用户名或邮箱已存在"}), 400

    user_row = get_user_by_account(db, payload["username"])
    try:
        result = activate_account_for_machine(
            db,
            user_row=user_row,
            machine_id=payload["machine_id"],
            device_name=payload["device_name"],
            app_name=payload["app_name"],
            app_version=payload["app_version"],
            force_login=payload["force_login"],
        )
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    return jsonify({"ok": True, "data": result})


@app.route("/client-api/account/login", methods=["POST"])
@app.route("/account/login", methods=["POST"])
def client_login_account():
    data = request.get_json(silent=True) or {}
    payload, error = validate_account_auth_payload(data)
    if error:
        return jsonify({"ok": False, "message": error}), 400
    if not payload["password"]:
        return jsonify({"ok": False, "message": "请输入密码"}), 400

    db = get_db()
    user_row = get_user_by_account(db, payload["account"])
    if not user_row or not check_password_hash(user_row["password_hash"], payload["password"]):
        return jsonify({"ok": False, "message": "用户名/邮箱或密码不正确"}), 401

    try:
        result = activate_account_for_machine(
            db,
            user_row=user_row,
            machine_id=payload["machine_id"],
            device_name=payload["device_name"],
            app_name=payload["app_name"],
            app_version=payload["app_version"],
            force_login=payload["force_login"],
        )
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    return jsonify({"ok": True, "data": result})


@app.route("/client-api/account/logout", methods=["POST"])
@app.route("/account/logout", methods=["POST"])
def client_logout_account():
    data = request.get_json(silent=True) or {}
    machine_id = str(data.get("machine_id") or "").strip()
    token = str(data.get("token") or "").strip()
    account = str(data.get("account") or data.get("username") or data.get("email") or "").strip()
    if not machine_id:
        return jsonify({"ok": False, "message": "机器码不能为空"}), 400
    if not token:
        return jsonify({"ok": False, "message": "登录凭证不能为空"}), 400

    try:
        token_payload = verify_account_token(token)
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    if token_payload.get("machine_id") != machine_id:
        return jsonify({"ok": False, "message": "登录凭证与当前机器不匹配"}), 400

    db = get_db()
    user_row = db.execute(
        "SELECT * FROM users WHERE id = ?",
        (token_payload.get("user_id"),),
    ).fetchone()
    if not user_row:
        return jsonify({"ok": False, "message": "账号不存在"}), 404
    if account and account not in {user_row["username"], user_row["email"] or ""}:
        return jsonify({"ok": False, "message": "登录凭证与当前账号不匹配"}), 400

    result = db.execute(
        """
        UPDATE user_devices
        SET revoked_at = ?
        WHERE user_id = ?
          AND machine_id = ?
          AND token_hash = ?
          AND (revoked_at IS NULL OR revoked_at = '')
        """,
        (now_iso(), user_row["id"], machine_id, hash_token(token)),
    )
    db.execute(
        "UPDATE users SET updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (user_row["id"],),
    )
    db.commit()
    if result.rowcount == 0:
        return jsonify({"ok": False, "message": "当前机器未登录或登录凭证已失效"}), 400
    return jsonify(
        {
            "ok": True,
            "message": "退出登录成功",
            "data": {"username": user_row["username"], "machine_id": machine_id},
        }
    )


@app.route("/client-api/account/verify", methods=["POST"])
@app.route("/account/verify", methods=["POST"])
def client_verify_account():
    data = request.get_json(silent=True) or {}
    payload, error = validate_account_auth_payload(data)
    if error:
        return jsonify({"ok": False, "message": error}), 400
    if not payload["token"]:
        return jsonify({"ok": False, "message": "登录凭证不能为空"}), 400

    try:
        token_payload = verify_account_token(payload["token"])
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400

    if token_payload.get("machine_id") != payload["machine_id"]:
        return jsonify({"ok": False, "message": "登录凭证与当前机器不匹配"}), 400

    db = get_db()
    user_row = db.execute(
        "SELECT * FROM users WHERE id = ?",
        (token_payload.get("user_id"),),
    ).fetchone()
    if not user_row:
        return jsonify({"ok": False, "message": "账号不存在"}), 404
    if payload["account"] and payload["account"] not in {user_row["username"], user_row["email"] or ""}:
        return jsonify({"ok": False, "message": "登录凭证与当前账号不匹配"}), 400

    device_row = db.execute(
        """
        SELECT *
        FROM user_devices
        WHERE user_id = ? AND machine_id = ? AND (revoked_at IS NULL OR revoked_at = '')
        """,
        (user_row["id"], payload["machine_id"]),
    ).fetchone()
    if not device_row:
        return jsonify({"ok": False, "message": "当前机器未登录"}), 400
    if device_row["token_hash"] != hash_token(payload["token"]):
        return jsonify({"ok": False, "message": "登录凭证已失效，请重新登录"}), 400

    try:
        result = activate_account_for_machine(
            db,
            user_row=user_row,
            machine_id=payload["machine_id"],
            device_name=payload["device_name"],
            app_name=payload["app_name"],
            app_version=payload["app_version"],
        )
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    return jsonify({"ok": True, "data": result})


@app.route("/client-api/tt/account/login", methods=["POST"])
@app.route("/tt/account/login", methods=["POST"])
def client_login_tt_account():
    data = request.get_json(silent=True) or {}
    payload, error = validate_account_auth_payload(data)
    if error:
        return jsonify({"ok": False, "message": error}), 400
    if not payload["password"]:
        return jsonify({"ok": False, "message": "请输入密码"}), 400

    db = get_db()
    user_row = get_tt_user_by_account(db, payload["account"])
    if not user_row or not check_password_hash(user_row["password_hash"], payload["password"]):
        return jsonify({"ok": False, "message": "TT用户名、邮箱或密码不正确"}), 401

    try:
        result = activate_tt_account_for_machine(
            db,
            user_row=user_row,
            machine_id=payload["machine_id"],
            device_name=payload["device_name"],
            app_name=payload["app_name"],
            app_version=payload["app_version"],
            force_login=payload["force_login"],
        )
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    return jsonify({"ok": True, "data": result})


@app.route("/client-api/tt/account/logout", methods=["POST"])
@app.route("/tt/account/logout", methods=["POST"])
def client_logout_tt_account():
    data = request.get_json(silent=True) or {}
    machine_id = str(data.get("machine_id") or "").strip()
    token = str(data.get("token") or "").strip()
    account = str(data.get("account") or data.get("username") or data.get("email") or "").strip()
    if not machine_id:
        return jsonify({"ok": False, "message": "机器码不能为空"}), 400
    if not token:
        return jsonify({"ok": False, "message": "登录凭证不能为空"}), 400

    try:
        token_payload = verify_tt_account_token(token)
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    if token_payload.get("machine_id") != machine_id:
        return jsonify({"ok": False, "message": "登录凭证与当前机器不匹配"}), 400

    db = get_db()
    user_row = db.execute(
        "SELECT * FROM tt_users WHERE id = ?",
        (token_payload.get("tt_user_id"),),
    ).fetchone()
    if not user_row:
        return jsonify({"ok": False, "message": "TT账号不存在"}), 404
    if account and account not in {user_row["username"], user_row["email"] or ""}:
        return jsonify({"ok": False, "message": "登录凭证与当前TT账号不匹配"}), 400

    result = db.execute(
        """
        UPDATE tt_user_devices
        SET revoked_at = ?
        WHERE tt_user_id = ?
          AND machine_id = ?
          AND token_hash = ?
          AND (revoked_at IS NULL OR revoked_at = '')
        """,
        (now_iso(), user_row["id"], machine_id, hash_token(token)),
    )
    db.execute(
        "UPDATE tt_users SET updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (user_row["id"],),
    )
    db.commit()
    if result.rowcount == 0:
        return jsonify({"ok": False, "message": "当前机器未登录或登录凭证已失效"}), 400
    return jsonify(
        {
            "ok": True,
            "message": "退出登录成功",
            "data": {"username": user_row["username"], "machine_id": machine_id},
        }
    )


@app.route("/client-api/tt/account/verify", methods=["POST"])
@app.route("/tt/account/verify", methods=["POST"])
def client_verify_tt_account():
    data = request.get_json(silent=True) or {}
    payload, error = validate_account_auth_payload(data)
    if error:
        return jsonify({"ok": False, "message": error}), 400
    if not payload["token"]:
        return jsonify({"ok": False, "message": "登录凭证不能为空"}), 400

    try:
        token_payload = verify_tt_account_token(payload["token"])
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400

    if token_payload.get("machine_id") != payload["machine_id"]:
        return jsonify({"ok": False, "message": "登录凭证与当前机器不匹配"}), 400

    db = get_db()
    user_row = db.execute(
        "SELECT * FROM tt_users WHERE id = ?",
        (token_payload.get("tt_user_id"),),
    ).fetchone()
    if not user_row:
        return jsonify({"ok": False, "message": "TT账号不存在"}), 404
    if payload["account"] and payload["account"] not in {user_row["username"], user_row["email"] or ""}:
        return jsonify({"ok": False, "message": "登录凭证与当前TT账号不匹配"}), 400

    device_row = db.execute(
        """
        SELECT *
        FROM tt_user_devices
        WHERE tt_user_id = ? AND machine_id = ? AND (revoked_at IS NULL OR revoked_at = '')
        """,
        (user_row["id"], payload["machine_id"]),
    ).fetchone()
    if not device_row:
        return jsonify({"ok": False, "message": "当前机器未登录"}), 400
    if device_row["token_hash"] != hash_token(payload["token"]):
        return jsonify({"ok": False, "message": "登录凭证已失效，请重新登录"}), 400

    try:
        result = activate_tt_account_for_machine(
            db,
            user_row=user_row,
            machine_id=payload["machine_id"],
            device_name=payload["device_name"],
            app_name=payload["app_name"],
            app_version=payload["app_version"],
        )
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    return jsonify({"ok": True, "data": result})


@app.route("/client-api/account/remote-client", methods=["POST"])
def client_create_remote_client_from_account():
    data = request.get_json(silent=True) or {}
    payload, error = validate_account_auth_payload(data)
    if error:
        return jsonify({"ok": False, "message": error}), 400
    if not payload["token"]:
        return jsonify({"ok": False, "message": "登录凭证不能为空"}), 400

    try:
        token_payload = verify_account_token(payload["token"])
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    if token_payload.get("machine_id") != payload["machine_id"]:
        return jsonify({"ok": False, "message": "登录凭证与当前机器不匹配"}), 400

    db = get_db()
    user_row = db.execute(
        "SELECT * FROM users WHERE id = ?",
        (token_payload.get("user_id"),),
    ).fetchone()
    if not user_row:
        return jsonify({"ok": False, "message": "账号不存在"}), 404
    if payload["account"] and payload["account"] not in {user_row["username"], user_row["email"] or ""}:
        return jsonify({"ok": False, "message": "登录凭证与当前账号不匹配"}), 400
    ok, account_error = ensure_account_can_login(user_row)
    if not ok:
        return jsonify({"ok": False, "message": account_error}), 400

    device_row = db.execute(
        """
        SELECT *
        FROM user_devices
        WHERE user_id = ? AND machine_id = ?
        """,
        (user_row["id"], payload["machine_id"]),
    ).fetchone()
    if not device_row:
        return jsonify({"ok": False, "message": "当前机器未登录或登录凭证已失效"}), 400
    if str(device_row["revoked_at"] or "").strip():
        return jsonify({"ok": False, "message": "当前机器登录已退出，请重新登录"}), 400
    if device_row["token_hash"] != hash_token(payload["token"]):
        return jsonify({"ok": False, "message": "登录凭证已失效，请重新登录"}), 400

    client_name = str(data.get("client_name") or "").strip() or "默认设备"
    row, client_token = create_remote_client_record(
        db,
        owner_user_id=int(user_row["id"]),
        client_name=client_name,
    )
    return jsonify(
        {
            "ok": True,
            "data": {
                "item": serialize_remote_client(row),
                "client_token": client_token,
                "user_role": user_row["role"] or "user",
            },
        }
    ), 201


@app.route("/client-api/account/management-session", methods=["POST"])
def client_create_management_session_from_account():
    data = request.get_json(silent=True) or {}
    user_row, _payload, error_response = resolve_user_from_account_token_payload(data)
    if error_response is not None:
        return error_response

    session.clear()
    session["user_id"] = user_row["id"]
    session["username"] = user_row["username"]
    session["role"] = user_row["role"] or "user"
    return jsonify(
        {
            "ok": True,
            "data": {
                "user_id": user_row["id"],
                "username": user_row["username"],
                "role": user_row["role"] or "user",
            },
        }
    )


@app.route("/client-api/licenses/activate", methods=["POST"])
@app.route("/client-api/license/activate", methods=["POST"])
@app.route("/license/activate", methods=["POST"])
def client_activate_license():
    data = request.get_json(silent=True) or {}
    payload, error = validate_client_license_payload(data)
    if error:
        return jsonify({"ok": False, "message": error}), 400

    db = get_db()
    license_row = db.execute(
        "SELECT * FROM licenses WHERE license_key = ? AND deleted_at IS NULL",
        (payload["license_key"],),
    ).fetchone()
    if not license_row:
        return jsonify({"ok": False, "message": "激活码无效"}), 404
    if license_row["status"] != "active":
        return jsonify({"ok": False, "message": "该激活码已被停用"}), 400
    if is_license_expired(license_row):
        db.execute(
            "UPDATE licenses SET status = 'expired', updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (license_row["id"],),
        )
        db.commit()
        return jsonify({"ok": False, "message": "该激活码已过期"}), 400

    try:
        result = activate_license_for_machine(
            db,
            license_row=license_row,
            machine_id=payload["machine_id"],
            app_name=payload["app_name"],
            app_version=payload["app_version"],
        )
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    return jsonify({"ok": True, "data": result})


@app.route("/client-api/licenses/verify", methods=["POST"])
@app.route("/client-api/license/verify", methods=["POST"])
@app.route("/license/verify", methods=["POST"])
def client_verify_license():
    data = request.get_json(silent=True) or {}
    payload, error = validate_client_license_payload(data)
    if error:
        return jsonify({"ok": False, "message": error}), 400
    if not payload["token"]:
        return jsonify({"ok": False, "message": "token 不能为空"}), 400

    db = get_db()
    license_row = db.execute(
        "SELECT * FROM licenses WHERE license_key = ? AND deleted_at IS NULL",
        (payload["license_key"],),
    ).fetchone()
    if not license_row:
        return jsonify({"ok": False, "message": "激活码无效"}), 404
    if license_row["status"] != "active":
        return jsonify({"ok": False, "message": "该激活码已被停用"}), 400
    if is_license_expired(license_row):
        db.execute(
            "UPDATE licenses SET status = 'expired', updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (license_row["id"],),
        )
        db.commit()
        return jsonify({"ok": False, "message": "该激活码已过期"}), 400

    try:
        token_payload = verify_license_token(payload["token"])
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400

    if token_payload.get("license_key") != license_row["license_key"]:
        return jsonify({"ok": False, "message": "授权 token 与激活码不匹配"}), 400
    if token_payload.get("machine_id") != payload["machine_id"]:
        return jsonify({"ok": False, "message": "授权 token 与当前机器不匹配"}), 400

    activation_row = db.execute(
        """
        SELECT *
        FROM license_activations
        WHERE license_id = ? AND machine_id = ? AND (revoked_at IS NULL OR revoked_at = '')
        """,
        (license_row["id"], payload["machine_id"]),
    ).fetchone()
    if not activation_row:
        return jsonify({"ok": False, "message": "当前机器未绑定该激活码"}), 400
    if activation_row["token_hash"] != hash_token(payload["token"]):
        return jsonify({"ok": False, "message": "授权 token 已失效，请重新激活"}), 400

    result = activate_license_for_machine(
        db,
        license_row=license_row,
        machine_id=payload["machine_id"],
        app_name=payload["app_name"],
        app_version=payload["app_version"],
    )
    return jsonify({"ok": True, "data": result})


@app.route("/api/dramas", methods=["GET"])
@login_required
def list_dramas():
    page = max(1, int(request.args.get("page", 1) or 1))
    page_size = int(request.args.get("page_size", 20) or 20)
    page_size = min(100, max(1, page_size))
    sort_by = request.args.get("sort_by", DEFAULT_SORT_FIELD)
    sort_dir = request.args.get("sort_dir", DEFAULT_SORT_DIR).lower()
    if sort_by not in SORTABLE_FIELDS:
        sort_by = DEFAULT_SORT_FIELD
    if sort_dir not in {"asc", "desc"}:
        sort_dir = DEFAULT_SORT_DIR

    clauses, params = build_filter_clause(request.args)
    where_sql = " AND ".join(["1=1"] + clauses)

    db = get_db()
    total = db.execute(
        f"SELECT COUNT(*) as cnt FROM dramas WHERE {where_sql}", params
    ).fetchone()[0]

    offset = (page - 1) * page_size
    rows = db.execute(
        f"SELECT * FROM dramas WHERE {where_sql} ORDER BY {sort_by} {sort_dir.upper()} LIMIT ? OFFSET ?",
        params + [page_size, offset],
    ).fetchall()

    pages = math.ceil(total / page_size) if total else 0

    items = [row_to_dict(row) for row in rows]
    return jsonify(
        {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": pages,
        }
    )


@app.route("/api/monitor/daily", methods=["GET"])
@login_required
@admin_required
def monitor_daily():
    clauses, params, mode, date_from, date_to = build_monitor_filters(request.args)
    date_field = "date(created_at, 'localtime')" if mode == "created" else "date"
    where_sql = " AND ".join(
        [f"{date_field} IS NOT NULL", f"{date_field} >= ?", f"{date_field} <= ?", *clauses]
    )
    query_params = [date_from.isoformat(), date_to.isoformat(), *params]

    db = get_db()
    rows = db.execute(
        f"""
        SELECT
            {date_field} AS stat_day,
            COUNT(*) AS new_count,
            SUM(CASE WHEN review_passed = '是' THEN 1 ELSE 0 END) AS review_passed_count,
            SUM(CASE WHEN uploaded = '是' THEN 1 ELSE 0 END) AS uploaded_count
        FROM dramas
        WHERE {where_sql}
        GROUP BY stat_day
        ORDER BY stat_day ASC
        """,
        query_params,
    ).fetchall()

    row_map: dict[str, dict[str, Any]] = {}
    for row in rows:
        day = row["stat_day"]
        new_count = int(row["new_count"] or 0)
        review_passed_count = int(row["review_passed_count"] or 0)
        uploaded_count = int(row["uploaded_count"] or 0)
        row_map[day] = {
            "day": day,
            "new_count": new_count,
            "review_passed_count": review_passed_count,
            "uploaded_count": uploaded_count,
            "review_rate": (review_passed_count / new_count) if new_count else 0,
            "upload_rate": (uploaded_count / new_count) if new_count else 0,
        }

    items: list[dict[str, Any]] = []
    cursor = date_from
    while cursor <= date_to:
        day = cursor.isoformat()
        items.append(
            row_map.get(
                day,
                {
                    "day": day,
                    "new_count": 0,
                    "review_passed_count": 0,
                    "uploaded_count": 0,
                    "review_rate": 0,
                    "upload_rate": 0,
                },
            )
        )
        cursor += datetime.timedelta(days=1)

    today_key = datetime.date.today().isoformat()
    today_item = row_map.get(
        today_key,
        {
            "new_count": 0,
            "review_passed_count": 0,
            "uploaded_count": 0,
        },
    )
    range_new_count = sum(item["new_count"] for item in items)
    range_review_passed_count = sum(item["review_passed_count"] for item in items)
    range_uploaded_count = sum(item["uploaded_count"] for item in items)

    return jsonify(
        {
            "summary": {
                "today_new_count": today_item["new_count"],
                "today_review_passed_count": today_item["review_passed_count"],
                "today_uploaded_count": today_item["uploaded_count"],
                "range_new_count": range_new_count,
                "range_review_passed_count": range_review_passed_count,
                "range_uploaded_count": range_uploaded_count,
                "range_review_rate": (range_review_passed_count / range_new_count) if range_new_count else 0,
                "range_uploaded_rate": (range_uploaded_count / range_new_count) if range_new_count else 0,
                "label_suffix": "录入日期" if mode == "created" else "上线日期",
            },
            "rows": items,
        }
    )


@app.route("/api/dramas", methods=["POST"])
@login_required
@admin_required
def create_drama():
    data = request.get_json(silent=True) or {}
    payload, error = sanitize_drama_payload(data)
    if error:
        return jsonify({"error": error}), 400
    db = get_db()
    try:
        placeholders = ", ".join([f":{k}" for k in payload.keys()])
        columns = ", ".join(payload.keys())
        db.execute(f"INSERT INTO dramas ({columns}) VALUES ({placeholders})", payload)
        db.commit()
    except sqlite3.IntegrityError:
        return jsonify({"error": "已存在相同的原剧名和新剧名组合"}), 400
    return jsonify({"message": "创建成功"}), 201


@app.route("/api/dramas/<int:drama_id>", methods=["PUT"])
@login_required
@admin_required
def update_drama(drama_id: int):
    data = request.get_json(silent=True) or {}
    payload, error = sanitize_drama_payload(data)
    if error:
        return jsonify({"error": error}), 400
    db = get_db()
    set_clause = ", ".join([f"{key} = :{key}" for key in payload.keys()])
    payload["id"] = drama_id
    try:
        result = db.execute(
            f"UPDATE dramas SET {set_clause} WHERE id = :id",
            payload,
        )
        if result.rowcount == 0:
            return jsonify({"error": "未找到该短剧"}), 404
        db.commit()
    except sqlite3.IntegrityError:
        return jsonify({"error": "已存在相同的原剧名和新剧名组合"}), 400
    return jsonify({"message": "更新成功"})


@app.route("/api/dramas/<int:drama_id>", methods=["DELETE"])
@login_required
@admin_required
def delete_drama(drama_id: int):
    db = get_db()
    result = db.execute("DELETE FROM dramas WHERE id = ?", (drama_id,))
    db.commit()
    if result.rowcount == 0:
        return jsonify({"error": "未找到该短剧"}), 404
    return jsonify({"message": "删除成功"})


@app.route("/api/dramas/batch-delete", methods=["POST"])
@login_required
@admin_required
def batch_delete():
    data = request.get_json(silent=True) or {}
    ids = data.get("ids") or []
    if not isinstance(ids, list) or not ids:
        return jsonify({"error": "请选择要删除的短剧"}), 400
    placeholders = ",".join(["?"] * len(ids))
    db = get_db()
    db.execute(f"DELETE FROM dramas WHERE id IN ({placeholders})", ids)
    db.commit()
    return jsonify({"message": "批量删除完成"})


@app.route("/api/dramas/<int:drama_id>/upload", methods=["PATCH"])
@login_required
def toggle_upload(drama_id: int):
    db = get_db()
    row = db.execute(
        "SELECT uploaded FROM dramas WHERE id = ?",
        (drama_id,),
    ).fetchone()
    if not row:
        return jsonify({"error": "未找到该短剧"}), 404
    new_value = "否" if row["uploaded"] == "是" else "是"
    uploader_value = session.get("username") if new_value == "是" else None
    db.execute(
        "UPDATE dramas SET uploaded = ?, uploader = ? WHERE id = ?",
        (new_value, uploader_value, drama_id),
    )
    db.commit()
    return jsonify({"id": drama_id, "uploaded": new_value, "uploader": uploader_value})


@app.route("/api/dramas/<int:drama_id>/upload-state", methods=["PATCH"])
@login_required
def set_upload_state(drama_id: int):
    data = request.get_json(silent=True) or {}
    uploaded_value = normalize_flag(data.get("uploaded"))
    if uploaded_value not in ALLOWED_FLAGS:
        return jsonify({"error": "uploaded 只能是“是”或“否”"}), 400
    db = get_db()
    row = db.execute(
        "SELECT id FROM dramas WHERE id = ?",
        (drama_id,),
    ).fetchone()
    if not row:
        return jsonify({"error": "未找到该短剧"}), 404
    uploader_value = session.get("username") if uploaded_value == "是" else None
    db.execute(
        "UPDATE dramas SET uploaded = ?, uploader = ? WHERE id = ?",
        (uploaded_value, uploader_value, drama_id),
    )
    db.commit()
    return jsonify({"id": drama_id, "uploaded": uploaded_value, "uploader": uploader_value})


@app.route("/client-api/upload-records/batch", methods=["POST"])
def client_sync_upload_records():
    db, client_row = require_remote_client()
    if client_row is None:
        return jsonify({"ok": False, "message": "client_id 或 client_token 无效"}), 401
    data = request.get_json(silent=True) or {}
    raw_records = data.get("records")
    if not isinstance(raw_records, list):
        return jsonify({"ok": False, "message": "records 必须是数组"}), 400
    if len(raw_records) > 200:
        return jsonify({"ok": False, "message": "单次最多同步 200 条上传记录"}), 400

    owner_user_id = int(client_row["owner_user_id"])
    user_row = db.execute(
        "SELECT id, username FROM users WHERE id = ?",
        (owner_user_id,),
    ).fetchone()
    if not user_row:
        return jsonify({"ok": False, "message": "远程客户端归属用户不存在"}), 400

    created_count = 0
    updated_count = 0
    failed: list[dict[str, Any]] = []
    items: list[dict[str, Any]] = []
    for index, raw_item in enumerate(raw_records, start=1):
        if not isinstance(raw_item, dict):
            failed.append({"index": index, "message": "记录必须是对象"})
            continue
        item, error, created = upsert_upload_record(
            db,
            owner_user_id=owner_user_id,
            owner_username=str(user_row["username"] or ""),
            remote_client_id=int(client_row["id"]),
            data=raw_item,
        )
        if error:
            failed.append({"index": index, "message": error})
            continue
        if created:
            created_count += 1
        else:
            updated_count += 1
        if item:
            items.append(item)
    db.commit()
    return jsonify(
        {
            "ok": not failed,
            "data": {
                "created": created_count,
                "updated": updated_count,
                "failed": len(failed),
                "failed_items": failed,
                "items": items[:20],
            },
        }
    )


@app.route("/api/upload-records", methods=["GET"])
@login_required
@admin_required
def list_upload_records():
    user_id = int(session["user_id"])
    role = str(session.get("role") or "user").strip().lower()
    page = max(1, int(request.args.get("page", 1) or 1))
    page_size = min(100, max(1, int(request.args.get("page_size", 20) or 20)))

    clauses = ["1=1"]
    params: list[object] = []
    if role != "admin":
        clauses.append("ur.owner_user_id = ?")
        params.append(user_id)
    else:
        requested_user_id = str(request.args.get("user_id") or "").strip()
        if requested_user_id.isdigit():
            clauses.append("ur.owner_user_id = ?")
            params.append(int(requested_user_id))

    platform = normalize_upload_record_platform(request.args.get("platform"))
    if platform in UPLOAD_RECORD_PLATFORMS:
        clauses.append("ur.platform = ?")
        params.append(platform)
    status = str(request.args.get("status") or "").strip()
    if status:
        clauses.append("ur.upload_status LIKE ?")
        params.append(f"%{status}%")
    search = str(request.args.get("search") or "").strip()
    if search:
        like = f"%{search}%"
        clauses.append(
            "(ur.original_name LIKE ? OR ur.new_name LIKE ? OR ur.project_name LIKE ? OR ur.uploader_display LIKE ?)"
        )
        params.extend([like, like, like, like])
    date_from = str(request.args.get("date_from") or "").strip()
    if date_from:
        clauses.append("COALESCE(ur.date, substr(ur.record_time, 1, 10)) >= ?")
        params.append(date_from)
    date_to = str(request.args.get("date_to") or "").strip()
    if date_to:
        clauses.append("COALESCE(ur.date, substr(ur.record_time, 1, 10)) <= ?")
        params.append(date_to)

    where_sql = " AND ".join(clauses)
    db = get_db()
    total = db.execute(
        f"SELECT COUNT(*) FROM upload_records ur WHERE {where_sql}",
        params,
    ).fetchone()[0]
    offset = (page - 1) * page_size
    rows = db.execute(
        f"""
        SELECT ur.*, d.company
        FROM upload_records ur
        LEFT JOIN dramas d ON d.id = ur.drama_id
        WHERE {where_sql}
        ORDER BY COALESCE(ur.record_time, ur.created_at) DESC, ur.id DESC
        LIMIT ? OFFSET ?
        """,
        params + [page_size, offset],
    ).fetchall()
    return jsonify(
        {
            "items": [serialize_upload_record(row) for row in rows],
            "page": page,
            "page_size": page_size,
            "total": total,
            "pages": math.ceil(total / page_size) if total else 0,
            "can_view_all_users": role == "admin",
            "platforms": UPLOAD_RECORD_PLATFORM_LABELS,
        }
    )


@app.route("/api/platform-dramas", methods=["GET"])
@login_required
def list_platform_dramas():
    user_id = int(session["user_id"])
    role = str(session.get("role") or "user").strip().lower()
    platform = normalize_upload_record_platform(request.args.get("platform") or "video_channel")
    if platform not in UPLOAD_RECORD_PLATFORMS:
        return jsonify({"error": "platform 仅支持 video_channel / miniprogram / kuaishou"}), 400

    page = max(1, int(request.args.get("page", 1) or 1))
    page_size = min(100, max(1, int(request.args.get("page_size", 20) or 20)))
    sort_by = str(request.args.get("sort_by") or "record_time").strip()
    sort_dir = str(request.args.get("sort_dir") or "desc").strip().lower()
    sort_sql = PLATFORM_DRAMA_SORTABLE_FIELDS.get(sort_by, PLATFORM_DRAMA_SORTABLE_FIELDS["record_time"])
    if sort_dir not in {"asc", "desc"}:
        sort_dir = "desc"

    clauses = ["ur.platform = ?"]
    params: list[object] = [platform]
    if role != "admin":
        clauses.append("ur.owner_user_id = ?")
        params.append(user_id)
    else:
        requested_user_id = str(request.args.get("user_id") or "").strip()
        if requested_user_id.isdigit():
            clauses.append("ur.owner_user_id = ?")
            params.append(int(requested_user_id))

    search = str(request.args.get("search") or "").strip()
    if search:
        like = f"%{search}%"
        clauses.append(
            """
            (
                ur.original_name LIKE ? OR ur.new_name LIKE ? OR ur.project_name LIKE ?
                OR d.original_name LIKE ? OR d.new_name LIKE ?
            )
            """
        )
        params.extend([like, like, like, like, like])
    company = str(request.args.get("company") or "").strip()
    if company:
        clauses.append("d.company = ?")
        params.append(company)
    review_passed = normalize_flag(request.args.get("review_passed"))
    if review_passed in ALLOWED_FLAGS:
        clauses.append("d.review_passed = ?")
        params.append(review_passed)
    status = str(request.args.get("status") or request.args.get("uploaded") or "").strip()
    if status:
        clauses.append("ur.upload_status LIKE ?")
        params.append(f"%{status}%")
    uploader = str(request.args.get("uploader") or "").strip()
    if uploader:
        clauses.append("(ur.uploader_display LIKE ? OR ur.account_profile_name LIKE ? OR ur.owner_username LIKE ?)")
        params.extend([f"%{uploader}%", f"%{uploader}%", f"%{uploader}%"])
    date_from = str(request.args.get("date_from") or "").strip()
    if date_from:
        clauses.append("COALESCE(ur.date, substr(ur.record_time, 1, 10), d.date) >= ?")
        params.append(date_from)
    date_to = str(request.args.get("date_to") or "").strip()
    if date_to:
        clauses.append("COALESCE(ur.date, substr(ur.record_time, 1, 10), d.date) <= ?")
        params.append(date_to)

    where_sql = " AND ".join(clauses)
    db = get_db()
    total = db.execute(
        f"SELECT COUNT(*) FROM upload_records ur LEFT JOIN dramas d ON d.id = ur.drama_id WHERE {where_sql}",
        params,
    ).fetchone()[0]
    offset = (page - 1) * page_size
    rows = db.execute(
        f"""
        SELECT
            ur.id AS record_id,
            ur.owner_user_id,
            ur.owner_username,
            ur.remote_client_id,
            ur.drama_id,
            ur.platform,
            ur.platform_label,
            ur.sync_key,
            ur.record_time,
            COALESCE(ur.date, substr(ur.record_time, 1, 10), d.date) AS date,
            ur.upload_status,
            ur.execution_mode,
            ur.step_label,
            ur.project_name,
            ur.project_path,
            COALESCE(ur.original_name, d.original_name) AS original_name,
            COALESCE(ur.new_name, d.new_name) AS new_name,
            COALESCE(ur.episodes, d.episodes) AS episodes,
            d.duration,
            d.review_passed,
            d.uploaded,
            d.materials,
            d.promo_text,
            d.description,
            d.company,
            d.uploader AS drama_uploader,
            ur.video_file_count,
            ur.uploaded_video_count,
            ur.uploader_display,
            ur.account_profile_id,
            ur.account_profile_name,
            ur.device_name,
            ur.failure_reason,
            ur.extra_info,
            ur.series_id,
            ur.mini_series_id,
            ur.audit_status,
            ur.selling_status,
            ur.audit_reject_reason,
            ur.audit_reject_detail,
            ur.online_status,
            ur.online_at,
            ur.distribution_status,
            ur.distribution_at,
            ur.distribution_detail,
            ur.submitted_at,
            ur.created_at,
            ur.updated_at
        FROM upload_records ur
        LEFT JOIN dramas d ON d.id = ur.drama_id
        WHERE {where_sql}
        ORDER BY {sort_sql} {sort_dir.upper()}, ur.id DESC
        LIMIT ? OFFSET ?
        """,
        params + [page_size, offset],
    ).fetchall()
    items = []
    for row in rows:
        item = row_to_dict(row)
        item["id"] = item.get("drama_id") or 0
        item["row_key"] = f"{item.get('platform')}-{item.get('record_id')}"
        item["platform_label"] = item.get("platform_label") or UPLOAD_RECORD_PLATFORM_LABELS.get(platform, platform)
        items.append(item)
    return jsonify(
        {
            "items": items,
            "page": page,
            "page_size": page_size,
            "total": total,
            "pages": math.ceil(total / page_size) if total else 0,
            "platform": platform,
            "platform_label": UPLOAD_RECORD_PLATFORM_LABELS.get(platform, platform),
            "can_view_all_users": role == "admin",
        }
    )


@app.route("/api/dramas/<int:drama_id>/upload-records", methods=["GET"])
@login_required
def list_drama_upload_records(drama_id: int):
    user_id = int(session["user_id"])
    role = str(session.get("role") or "user").strip().lower()
    clauses = ["ur.drama_id = ?"]
    params: list[object] = [drama_id]
    if role != "admin":
        clauses.append("ur.owner_user_id = ?")
        params.append(user_id)
    db = get_db()
    rows = db.execute(
        f"""
        SELECT ur.*
        FROM upload_records ur
        WHERE {' AND '.join(clauses)}
        ORDER BY COALESCE(ur.record_time, ur.created_at) DESC, ur.id DESC
        LIMIT 200
        """,
        params,
    ).fetchall()
    return jsonify({"items": [serialize_upload_record(row) for row in rows]})


@app.route("/api/companies", methods=["GET"])
@login_required
def list_companies():
    db = get_db()
    role = str(session.get("role") or "user").strip().lower()
    if role == "admin":
        rows = db.execute(
            "SELECT DISTINCT company FROM dramas WHERE company IS NOT NULL AND company <> '' ORDER BY company ASC"
        ).fetchall()
    else:
        rows = db.execute(
            """
            SELECT DISTINCT d.company
            FROM upload_records ur
            JOIN dramas d ON d.id = ur.drama_id
            WHERE ur.owner_user_id = ?
              AND d.company IS NOT NULL
              AND d.company <> ''
            ORDER BY d.company ASC
            """,
            (int(session["user_id"]),),
        ).fetchall()
    return jsonify([row["company"] for row in rows])


@app.route("/api/export", methods=["GET"])
@login_required
def export_excel():
    clauses, params = build_filter_clause(request.args)
    where_sql = " AND ".join(["1=1"] + clauses)
    db = get_db()
    rows = db.execute(
        f"SELECT * FROM dramas WHERE {where_sql} ORDER BY date DESC",
        params,
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "短剧数据"
    ws.append(EXPORT_HEADERS)

    for row in rows:
        ws.append(
            [
                row["date"],
                row["original_name"],
                row["new_name"],
                row["episodes"],
                row["duration"],
                row["review_passed"],
                row["uploaded"],
                row["materials"],
                row["promo_text"],
                row["description"],
                row["company"],
                row["uploader"],
            ]
        )

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except ValueError:
                continue
        ws.column_dimensions[column].width = min(max_length + 2, 40)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"短剧数据_{datetime.date.today().isoformat()}.xlsx"
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/import", methods=["POST"])
@login_required
@admin_required
def import_excel():
    uploaded_file = request.files.get("file")
    if not uploaded_file:
        return jsonify({"error": "请上传Excel文件"}), 400
    try:
        workbook = load_workbook(uploaded_file, data_only=True)
    except Exception:
        return jsonify({"error": "无法读取Excel文件"}), 400

    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
    header_texts = {(header or "").strip() for header in header_row if header}
    header_map = {}
    for idx, header in enumerate(header_row, start=1):
        header_text = (header or "").strip()
        if header_text in HEADER_MAP:
            header_map[idx] = HEADER_MAP[header_text]
    if "original_name" not in header_map.values() or "new_name" not in header_map.values():
        return jsonify({"error": "Excel缺少必要的原剧名或新剧名列"}), 400
    has_uploader_column = "uploader" in header_map.values()
    has_uploaded_column = "uploaded" in header_map.values()
    is_submission_record_import = {"记录时间", "记录类型", "提交状态"}.issubset(header_texts)

    db = get_db()
    existing_rows = db.execute("SELECT original_name, new_name FROM dramas").fetchall()
    existing_pairs = {(row["original_name"], row["new_name"]) for row in existing_rows}

    new_count = 0
    duplicate_count = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        empty = True
        for idx, value in enumerate(row, start=1):
            if idx in header_map:
                row_data[header_map[idx]] = value
                if value not in (None, ""):
                    empty = False
        if empty:
            continue
        normalized = normalize_row(row_data)
        original_name = normalized.get("original_name")
        new_name = normalized.get("new_name")
        if not original_name or not new_name:
            continue
        pair = (original_name, new_name)
        if pair in existing_pairs:
            duplicate_count += 1
            continue
        uploaded_value = normalized.get("uploaded") or None
        if is_submission_record_import and not has_uploaded_column:
            uploaded_value = "是"
        elif uploaded_value not in ALLOWED_FLAGS:
            uploaded_value = "否"
        uploader_value = normalized.get("uploader")
        if not has_uploader_column and uploaded_value == "是":
            uploader_value = session.get("username") or None
        insert_payload = {
            "date": normalized.get("date"),
            "original_name": original_name,
            "new_name": new_name,
            "episodes": normalized.get("episodes"),
            "duration": normalized.get("duration"),
            "review_passed": normalized.get("review_passed") or "否",
            "uploaded": uploaded_value,
            "uploader": uploader_value,
            "materials": normalized.get("materials"),
            "promo_text": normalized.get("promo_text"),
            "description": normalized.get("description"),
            "company": normalized.get("company"),
            "remark1": normalized.get("remark1"),
            "remark2": normalized.get("remark2"),
            "remark3": normalized.get("remark3"),
        }
        placeholders = ", ".join([f":{k}" for k in insert_payload.keys()])
        columns = ", ".join(insert_payload.keys())
        db.execute(
            f"INSERT INTO dramas ({columns}) VALUES ({placeholders})",
            insert_payload,
        )
        db.commit()
        existing_pairs.add(pair)
        new_count += 1

    return jsonify(
        {
            "new_count": new_count,
            "duplicate_count": duplicate_count,
            "conflicts": [],
        }
    )


def sanitize_user_payload(data: dict, *, require_password: bool = False) -> tuple[dict, str | None]:
    username = str(data.get("username") or "").strip()
    email = normalize_email(str(data.get("email") or ""))
    password = str(data.get("password") or "")
    role = str(data.get("role") or "user").strip().lower()
    status = str(data.get("status") or "active").strip().lower()
    edition = str(data.get("edition") or "pro").strip().lower()
    expires_at = str(data.get("expires_at") or "").strip()
    try:
        max_devices = int(data.get("max_devices") or ACCOUNT_DEFAULT_MAX_DEVICES)
    except (TypeError, ValueError):
        return {}, "最大设备数必须是正整数"
    if not USERNAME_RE.match(username):
        return {}, "用户名需为 2-30 位字母数字下划线，或使用邮箱格式"
    if require_password and len(password) < 6:
        return {}, "密码至少6位"
    if email and not EMAIL_RE.match(email):
        return {}, "邮箱格式不正确"
    if role not in {"admin", "user"}:
        role = "user"
    if status not in {"active", "disabled"}:
        return {}, "账号状态不正确"
    if edition not in LICENSE_EDITION_VALUES:
        edition = "pro"
    if max_devices < 1:
        return {}, "最大设备数必须是正整数"
    if expires_at:
        try:
            datetime.datetime.fromisoformat(expires_at)
        except ValueError:
            return {}, "到期时间格式不正确"
    return {
        "username": username,
        "email": email,
        "password": password,
        "role": role,
        "status": status,
        "edition": edition,
        "max_devices": max_devices,
        "expires_at": expires_at or None,
    }, None


def sanitize_tt_user_payload(data: dict, *, require_password: bool = False) -> tuple[dict, str | None]:
    payload, error = sanitize_user_payload(data, require_password=require_password)
    if error:
        return {}, error
    payload.pop("role", None)
    return payload, None


@app.route("/api/users", methods=["GET"])
@login_required
@admin_required
def list_users():
    db = get_db()
    rows = db.execute(
        """
        SELECT
            u.id, u.username, u.email, u.role, u.status, u.max_devices,
            u.edition, u.expires_at, u.created_at,
            (SELECT COUNT(*) FROM user_devices d WHERE d.user_id = u.id AND (d.revoked_at IS NULL OR d.revoked_at = '')) AS active_devices,
            (SELECT COUNT(*) FROM user_devices d WHERE d.user_id = u.id) AS total_devices,
            (SELECT COALESCE(MAX(d.last_verified_at), '') FROM user_devices d WHERE d.user_id = u.id) AS last_verified_at
        FROM users u
        ORDER BY u.created_at DESC
        """
    ).fetchall()
    return jsonify([dict(row) for row in rows])


@app.route("/api/users", methods=["POST"])
@login_required
@admin_required
def create_user():
    data = request.get_json(silent=True) or {}
    payload, error = sanitize_user_payload(data, require_password=True)
    if error:
        return jsonify({"error": error}), 400
    db = get_db()
    try:
        db.execute(
            """
            INSERT INTO users (
                username, email, password_hash, role, status,
                max_devices, edition, expires_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            """,
            (
                payload["username"],
                payload["email"] or None,
                generate_password_hash(payload["password"]),
                payload["role"],
                payload["status"],
                payload["max_devices"],
                payload["edition"],
                payload["expires_at"],
            ),
        )
        db.commit()
    except sqlite3.IntegrityError:
        return jsonify({"error": "用户名或邮箱已存在"}), 400
    return jsonify({"message": "创建用户成功"}), 201


@app.route("/api/users/<int:user_id>", methods=["PUT"])
@login_required
@admin_required
def update_user(user_id: int):
    data = request.get_json(silent=True) or {}
    payload, error = sanitize_user_payload(data)
    if error:
        return jsonify({"error": error}), 400
    if session.get("user_id") == user_id:
        if payload["status"] != "active":
            return jsonify({"error": "不能停用当前登录用户"}), 400
        if payload["role"] != "admin":
            return jsonify({"error": "不能将当前登录管理员改为普通用户"}), 400
    db = get_db()
    try:
        result = db.execute(
            """
            UPDATE users
            SET username = ?, email = ?, role = ?, status = ?, max_devices = ?,
                edition = ?, expires_at = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                payload["username"],
                payload["email"] or None,
                payload["role"],
                payload["status"],
                payload["max_devices"],
                payload["edition"],
                payload["expires_at"],
                user_id,
            ),
        )
        db.commit()
    except sqlite3.IntegrityError:
        return jsonify({"error": "用户名或邮箱已存在"}), 400
    if result.rowcount == 0:
        return jsonify({"error": "未找到该用户"}), 404
    return jsonify({"message": "用户已更新"})


@app.route("/api/users/<int:user_id>/devices", methods=["GET"])
@login_required
@admin_required
def list_user_devices(user_id: int):
    db = get_db()
    user = db.execute(
        "SELECT id, username, email, status, max_devices, edition, expires_at FROM users WHERE id = ?",
        (user_id,),
    ).fetchone()
    if not user:
        return jsonify({"error": "未找到该用户"}), 404
    rows = db.execute(
        """
        SELECT id, machine_id, device_name, app_name, app_version, logged_in_at, last_verified_at, revoked_at
        FROM user_devices
        WHERE user_id = ?
        ORDER BY COALESCE(last_verified_at, logged_in_at, '') DESC, id DESC
        """,
        (user_id,),
    ).fetchall()
    return jsonify({"user": dict(user), "devices": [dict(row) for row in rows]})


@app.route("/api/users/<int:user_id>/devices/<int:device_id>/revoke", methods=["POST"])
@login_required
@admin_required
def revoke_user_device(user_id: int, device_id: int):
    db = get_db()
    result = db.execute(
        """
        UPDATE user_devices
        SET revoked_at = ?
        WHERE id = ? AND user_id = ? AND (revoked_at IS NULL OR revoked_at = '')
        """,
        (now_iso(), device_id, user_id),
    )
    db.commit()
    if result.rowcount == 0:
        return jsonify({"error": "未找到可解绑的设备"}), 404
    return jsonify({"message": "设备已解绑"})


@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@login_required
@admin_required
def delete_user(user_id: int):
    if session.get("user_id") == user_id:
        return jsonify({"error": "不能删除当前登录用户"}), 400
    db = get_db()
    result = db.execute("DELETE FROM users WHERE id = ?", (user_id,))
    db.commit()
    if result.rowcount == 0:
        return jsonify({"error": "未找到该用户"}), 404
    return jsonify({"message": "删除成功"})


@app.route("/api/users/<int:user_id>/password", methods=["PUT"])
@login_required
@admin_required
def change_password(user_id: int):
    data = request.get_json(silent=True) or {}
    new_password = (data.get("new_password") or "").strip()
    if len(new_password) < 4:
        return jsonify({"error": "新密码至少4位"}), 400
    db = get_db()
    result = db.execute(
        "UPDATE users SET password_hash = ? WHERE id = ?",
        (generate_password_hash(new_password), user_id),
    )
    db.commit()
    if result.rowcount == 0:
        return jsonify({"error": "未找到该用户"}), 404
    return jsonify({"message": "密码已更新"})


@app.route("/api/tt-users", methods=["GET"])
@login_required
@admin_required
def list_tt_users():
    db = get_db()
    rows = db.execute(
        """
        SELECT
            u.id, u.username, u.email, 'tt_user' AS role, u.status, u.max_devices,
            u.edition, u.expires_at, u.created_at,
            (SELECT COUNT(*) FROM tt_user_devices d WHERE d.tt_user_id = u.id AND (d.revoked_at IS NULL OR d.revoked_at = '')) AS active_devices,
            (SELECT COUNT(*) FROM tt_user_devices d WHERE d.tt_user_id = u.id) AS total_devices,
            (SELECT COALESCE(MAX(d.last_verified_at), '') FROM tt_user_devices d WHERE d.tt_user_id = u.id) AS last_verified_at
        FROM tt_users u
        ORDER BY u.created_at DESC
        """
    ).fetchall()
    return jsonify([dict(row) for row in rows])


@app.route("/api/tt-users", methods=["POST"])
@login_required
@admin_required
def create_tt_user():
    data = request.get_json(silent=True) or {}
    payload, error = sanitize_tt_user_payload(data, require_password=True)
    if error:
        return jsonify({"error": error}), 400
    db = get_db()
    try:
        db.execute(
            """
            INSERT INTO tt_users (
                username, email, password_hash, status,
                max_devices, edition, expires_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            """,
            (
                payload["username"],
                payload["email"] or None,
                generate_password_hash(payload["password"]),
                payload["status"],
                payload["max_devices"],
                payload["edition"],
                payload["expires_at"],
            ),
        )
        db.commit()
    except sqlite3.IntegrityError:
        return jsonify({"error": "TT用户名或邮箱已存在"}), 400
    return jsonify({"message": "创建TT用户成功"}), 201


@app.route("/api/tt-users/<int:user_id>", methods=["PUT"])
@login_required
@admin_required
def update_tt_user(user_id: int):
    data = request.get_json(silent=True) or {}
    payload, error = sanitize_tt_user_payload(data)
    if error:
        return jsonify({"error": error}), 400
    db = get_db()
    try:
        result = db.execute(
            """
            UPDATE tt_users
            SET username = ?, email = ?, status = ?, max_devices = ?,
                edition = ?, expires_at = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                payload["username"],
                payload["email"] or None,
                payload["status"],
                payload["max_devices"],
                payload["edition"],
                payload["expires_at"],
                user_id,
            ),
        )
        db.commit()
    except sqlite3.IntegrityError:
        return jsonify({"error": "TT用户名或邮箱已存在"}), 400
    if result.rowcount == 0:
        return jsonify({"error": "未找到该TT用户"}), 404
    return jsonify({"message": "TT用户已更新"})


@app.route("/api/tt-users/<int:user_id>/devices", methods=["GET"])
@login_required
@admin_required
def list_tt_user_devices(user_id: int):
    db = get_db()
    user = db.execute(
        "SELECT id, username, email, status, max_devices, edition, expires_at FROM tt_users WHERE id = ?",
        (user_id,),
    ).fetchone()
    if not user:
        return jsonify({"error": "未找到该TT用户"}), 404
    rows = db.execute(
        """
        SELECT id, machine_id, device_name, app_name, app_version, logged_in_at, last_verified_at, revoked_at
        FROM tt_user_devices
        WHERE tt_user_id = ?
        ORDER BY COALESCE(last_verified_at, logged_in_at, '') DESC, id DESC
        """,
        (user_id,),
    ).fetchall()
    return jsonify({"user": dict(user), "devices": [dict(row) for row in rows]})


@app.route("/api/tt-users/<int:user_id>/devices/<int:device_id>/revoke", methods=["POST"])
@login_required
@admin_required
def revoke_tt_user_device(user_id: int, device_id: int):
    db = get_db()
    result = db.execute(
        """
        UPDATE tt_user_devices
        SET revoked_at = ?
        WHERE id = ? AND tt_user_id = ? AND (revoked_at IS NULL OR revoked_at = '')
        """,
        (now_iso(), device_id, user_id),
    )
    db.commit()
    if result.rowcount == 0:
        return jsonify({"error": "未找到可解绑的TT设备"}), 404
    return jsonify({"message": "TT设备已解绑"})


@app.route("/api/tt-users/<int:user_id>", methods=["DELETE"])
@login_required
@admin_required
def delete_tt_user(user_id: int):
    db = get_db()
    result = db.execute("DELETE FROM tt_users WHERE id = ?", (user_id,))
    db.commit()
    if result.rowcount == 0:
        return jsonify({"error": "未找到该TT用户"}), 404
    return jsonify({"message": "删除成功"})


@app.route("/api/tt-users/<int:user_id>/password", methods=["PUT"])
@login_required
@admin_required
def change_tt_user_password(user_id: int):
    data = request.get_json(silent=True) or {}
    new_password = (data.get("new_password") or "").strip()
    if len(new_password) < 6:
        return jsonify({"error": "新密码至少6位"}), 400
    db = get_db()
    result = db.execute(
        "UPDATE tt_users SET password_hash = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (generate_password_hash(new_password), user_id),
    )
    db.commit()
    if result.rowcount == 0:
        return jsonify({"error": "未找到该TT用户"}), 404
    return jsonify({"message": "TT用户密码已更新"})


@app.route("/api/users/<int:user_id>/devices-legacy", methods=["GET"])
@login_required
@admin_required
def list_user_devices_legacy(user_id: int):
    db = get_db()
    user = db.execute(
        "SELECT id, username, max_devices FROM users WHERE id = ?",
        (user_id,),
    ).fetchone()
    if not user:
        return jsonify({"error": "未找到该用户"}), 404
    rows = db.execute(
        """
        SELECT id, user_id, machine_id, device_name, app_name, app_version,
               logged_in_at, last_verified_at, revoked_at
        FROM user_devices
        WHERE user_id = ?
        ORDER BY
            CASE WHEN revoked_at IS NULL OR revoked_at = '' THEN 0 ELSE 1 END,
            COALESCE(last_verified_at, logged_in_at, '') DESC
        """,
        (user_id,),
    ).fetchall()
    return jsonify({
        "user": dict(user),
        "items": [dict(row) for row in rows],
    })


@app.route("/api/users/<int:user_id>/devices/unbind", methods=["POST"])
@login_required
@admin_required
def unbind_user_device(user_id: int):
    data = request.get_json(silent=True) or {}
    machine_id = str(data.get("machine_id") or "").strip()
    if not machine_id:
        return jsonify({"error": "machine_id 不能为空"}), 400
    db = get_db()
    user = db.execute("SELECT id FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        return jsonify({"error": "未找到该用户"}), 404
    result = db.execute(
        """
        UPDATE user_devices
        SET revoked_at = ?
        WHERE user_id = ? AND machine_id = ? AND (revoked_at IS NULL OR revoked_at = '')
        """,
        (
            datetime.datetime.now().isoformat(timespec="seconds"),
            user_id,
            machine_id,
        ),
    )
    db.commit()
    if result.rowcount == 0:
        return jsonify({"error": "未找到可解绑的设备记录"}), 404
    return jsonify({"message": "设备解绑成功"})


@app.route("/api/licenses", methods=["GET"])
@login_required
@admin_required
def list_licenses():
    page = max(1, int(request.args.get("page", 1) or 1))
    page_size = int(request.args.get("page_size", 10) or 10)
    page_size = min(100, max(1, page_size))
    sort_by = str(
        request.args.get("sort_by", LICENSE_LIST_DEFAULT_SORT_FIELD) or LICENSE_LIST_DEFAULT_SORT_FIELD
    ).strip()
    sort_dir = str(
        request.args.get("sort_dir", LICENSE_LIST_DEFAULT_SORT_DIR) or LICENSE_LIST_DEFAULT_SORT_DIR
    ).strip().lower()
    if sort_by not in LICENSE_LIST_SORTABLE_FIELDS:
        sort_by = LICENSE_LIST_DEFAULT_SORT_FIELD
    if sort_dir not in {"asc", "desc"}:
        sort_dir = LICENSE_LIST_DEFAULT_SORT_DIR

    clauses, params = build_license_filter_clause(request.args)
    where_sql = " AND ".join(["1=1"] + clauses)

    db = get_db()
    total = db.execute(
        f"SELECT COUNT(*) AS cnt FROM licenses WHERE {where_sql}",
        params,
    ).fetchone()[0]
    offset = (page - 1) * page_size
    rows = db.execute(
        f"""
        SELECT *
        FROM licenses
        WHERE {where_sql}
        ORDER BY {LICENSE_LIST_SORTABLE_FIELDS[sort_by]} {sort_dir.upper()}, id DESC
        LIMIT ? OFFSET ?
        """,
        params + [page_size, offset],
    ).fetchall()
    pages = math.ceil(total / page_size) if total else 1
    return jsonify(
        {
            "items": [serialize_license_row(db, row) for row in rows],
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": pages,
        }
    )


@app.route("/api/licenses", methods=["POST"])
@login_required
@admin_required
def create_license():
    data = request.get_json(silent=True) or {}
    payload, error = sanitize_license_payload(data)
    if error:
        return jsonify({"error": error}), 400
    db = get_db()
    try:
        db.execute(
            """
            INSERT INTO licenses (
                license_key, license_key_masked, status, edition, licensee, max_activations, expires_at, notes
            ) VALUES (
                :license_key, :license_key_masked, :status, :edition, :licensee, :max_activations, :expires_at, :notes
            )
            """,
            payload,
        )
        db.commit()
    except sqlite3.IntegrityError:
        return jsonify({"error": "激活码已存在，请更换后重试"}), 400

    row = db.execute(
        "SELECT * FROM licenses WHERE license_key = ?",
        (payload["license_key"],),
    ).fetchone()
    return jsonify(
        {
            "message": "激活码创建成功",
            "item": serialize_license_row(db, row),
        }
    ), 201


@app.route("/api/licenses/export", methods=["GET"])
@login_required
@admin_required
def export_licenses():
    clauses, params = build_license_filter_clause(request.args)
    where_sql = " AND ".join(["1=1"] + clauses)
    db = get_db()
    rows = db.execute(
        f"""
        SELECT *
        FROM licenses
        WHERE {where_sql}
        ORDER BY created_at DESC, id DESC
        """,
        params,
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "授权码"
    ws.append(LICENSE_EXPORT_HEADERS)

    for row in rows:
        item = serialize_license_row(db, row)
        status_text = "已删除" if item.get("deleted_at") else item.get("status") or ""
        ws.append(
            [
                item.get("license_key") or "",
                item.get("license_key_masked") or "",
                item.get("licensee") or "",
                item.get("edition") or "",
                status_text,
                item.get("max_activations") or 0,
                item.get("active_activations") or 0,
                item.get("total_activations") or 0,
                item.get("expires_at") or "",
                item.get("last_verified_at") or "",
                item.get("notes") or "",
                item.get("created_at") or "",
                item.get("updated_at") or "",
                item.get("deleted_at") or "",
            ]
        )

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except ValueError:
                continue
        ws.column_dimensions[column].width = min(max_length + 2, 40)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"授权码数据_{datetime.date.today().isoformat()}.xlsx"
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/licenses/<int:license_id>/activations", methods=["GET"])
@login_required
@admin_required
def list_license_activations(license_id: int):
    db = get_db()
    license_row = get_license_row(db, license_id, include_deleted=True)
    if not license_row:
        return jsonify({"error": "未找到该激活码"}), 404
    rows = db.execute(
        """
        SELECT *
        FROM license_activations
        WHERE license_id = ?
        ORDER BY activated_at DESC, id DESC
        """,
        (license_id,),
    ).fetchall()
    return jsonify(
        {
            "license": serialize_license_row(db, license_row),
            "items": [serialize_activation_row(row) for row in rows],
        }
    )


@app.route("/api/licenses/<int:license_id>/secret", methods=["GET"])
@login_required
@admin_required
def get_license_secret(license_id: int):
    db = get_db()
    row = get_license_row(db, license_id, include_deleted=True)
    if not row:
        return jsonify({"error": "未找到该激活码"}), 404
    return jsonify(
        {
            "id": row["id"],
            "license_key": row["license_key"],
            "license_key_masked": row["license_key_masked"],
            "licensee": row["licensee"] or "",
            "edition": row["edition"] or "",
            "status": row["status"] or "",
            "deleted_at": row["deleted_at"] or "",
        }
    )


@app.route("/api/licenses/<int:license_id>/disable", methods=["POST"])
@login_required
@admin_required
def disable_license(license_id: int):
    db = get_db()
    row = get_license_row(db, license_id, include_deleted=True)
    if not row:
        return jsonify({"error": "未找到该激活码"}), 404
    ok, message = update_license_status_row(db, row, status="disabled")
    if not ok:
        return jsonify({"error": message}), 400
    db.commit()
    return jsonify({"message": "激活码已停用"})


@app.route("/api/licenses/<int:license_id>/enable", methods=["POST"])
@login_required
@admin_required
def enable_license(license_id: int):
    db = get_db()
    row = get_license_row(db, license_id, include_deleted=True)
    if not row:
        return jsonify({"error": "未找到该激活码"}), 404
    ok, message = update_license_status_row(db, row, status="active")
    if not ok:
        return jsonify({"error": message}), 400
    db.commit()
    return jsonify({"message": "激活码已启用"})


@app.route("/api/licenses/batch-disable", methods=["POST"])
@login_required
@admin_required
def batch_disable_licenses():
    data = request.get_json(silent=True) or {}
    license_ids, error = parse_license_ids_from_payload(data)
    if error:
        return jsonify({"error": error}), 400

    db = get_db()
    rows = get_license_rows_by_ids(db, license_ids, include_deleted=True)
    found_ids = {int(row["id"]) for row in rows}
    missing_ids = [license_id for license_id in license_ids if license_id not in found_ids]
    if missing_ids:
        return jsonify({"error": f"存在未找到的授权码：{', '.join(map(str, missing_ids))}"}), 404

    for row in rows:
        ok, message = update_license_status_row(db, row, status="disabled")
        if not ok:
            db.rollback()
            return jsonify({"error": message}), 400

    db.commit()
    return jsonify({"message": f"已停用 {len(rows)} 条授权码"})


@app.route("/api/licenses/batch-enable", methods=["POST"])
@login_required
@admin_required
def batch_enable_licenses():
    data = request.get_json(silent=True) or {}
    license_ids, error = parse_license_ids_from_payload(data)
    if error:
        return jsonify({"error": error}), 400

    db = get_db()
    rows = get_license_rows_by_ids(db, license_ids, include_deleted=True)
    found_ids = {int(row["id"]) for row in rows}
    missing_ids = [license_id for license_id in license_ids if license_id not in found_ids]
    if missing_ids:
        return jsonify({"error": f"存在未找到的授权码：{', '.join(map(str, missing_ids))}"}), 404

    for row in rows:
        ok, message = update_license_status_row(db, row, status="active")
        if not ok:
            db.rollback()
            return jsonify({"error": message}), 400

    db.commit()
    return jsonify({"message": f"已启用 {len(rows)} 条授权码"})


@app.route("/api/licenses/<int:license_id>/unbind", methods=["POST"])
@login_required
@admin_required
def unbind_license_machine(license_id: int):
    data = request.get_json(silent=True) or {}
    machine_id = str(data.get("machine_id") or "").strip()
    if not machine_id:
        return jsonify({"error": "machine_id 不能为空"}), 400
    db = get_db()
    if not get_license_row(db, license_id):
        return jsonify({"error": "未找到该激活码"}), 404
    result = db.execute(
        """
        UPDATE license_activations
        SET revoked_at = ?
        WHERE license_id = ? AND machine_id = ? AND (revoked_at IS NULL OR revoked_at = '')
        """,
        (
            datetime.datetime.now().isoformat(timespec="seconds"),
            license_id,
            machine_id,
        ),
    )
    db.commit()
    if result.rowcount == 0:
        return jsonify({"error": "未找到可解绑的设备记录"}), 404
    return jsonify({"message": "设备解绑成功"})


@app.route("/api/licenses/<int:license_id>", methods=["DELETE"])
@login_required
@admin_required
def delete_license(license_id: int):
    db = get_db()
    row = get_license_row(db, license_id, include_deleted=True)
    if not row:
        return jsonify({"error": "未找到该激活码"}), 404
    ok, message = soft_delete_license_row(db, row, deleted_by=session.get("user_id"))
    if not ok:
        return jsonify({"error": message}), 400
    db.commit()
    return jsonify({"message": "激活码已删除"})


@app.route("/api/licenses/batch-delete", methods=["POST"])
@login_required
@admin_required
def batch_delete_licenses():
    data = request.get_json(silent=True) or {}
    license_ids, error = parse_license_ids_from_payload(data)
    if error:
        return jsonify({"error": error}), 400

    db = get_db()
    rows = get_license_rows_by_ids(db, license_ids, include_deleted=True)
    found_ids = {int(row["id"]) for row in rows}
    missing_ids = [license_id for license_id in license_ids if license_id not in found_ids]
    if missing_ids:
        return jsonify({"error": f"存在未找到的授权码：{', '.join(map(str, missing_ids))}"}), 404

    for row in rows:
        ok, message = soft_delete_license_row(db, row, deleted_by=session.get("user_id"))
        if not ok:
            db.rollback()
            return jsonify({"error": message}), 400

    db.commit()
    return jsonify({"message": f"已删除 {len(rows)} 条授权码"})


@app.route("/api/licenses/<int:license_id>/restore", methods=["POST"])
@login_required
@admin_required
def restore_license(license_id: int):
    db = get_db()
    row = get_license_row(db, license_id, include_deleted=True)
    if not row:
        return jsonify({"error": "未找到该激活码"}), 404
    ok, message = restore_license_row(db, row)
    if not ok:
        return jsonify({"error": message}), 400
    db.commit()
    return jsonify({"message": "激活码已恢复"})


@app.route("/api/licenses/batch-restore", methods=["POST"])
@login_required
@admin_required
def batch_restore_licenses():
    data = request.get_json(silent=True) or {}
    license_ids, error = parse_license_ids_from_payload(data)
    if error:
        return jsonify({"error": error}), 400

    db = get_db()
    rows = get_license_rows_by_ids(db, license_ids, include_deleted=True)
    found_ids = {int(row["id"]) for row in rows}
    missing_ids = [license_id for license_id in license_ids if license_id not in found_ids]
    if missing_ids:
        return jsonify({"error": f"存在未找到的授权码：{', '.join(map(str, missing_ids))}"}), 404

    for row in rows:
        ok, message = restore_license_row(db, row)
        if not ok:
            db.rollback()
            return jsonify({"error": message}), 400

    db.commit()
    return jsonify({"message": f"已恢复 {len(rows)} 条授权码"})


@app.route("/api/remote/clients", methods=["GET"])
@login_required
@admin_required
def list_remote_clients():
    db = get_db()
    user_id = int(session["user_id"])
    role = str(session.get("role") or "user")
    if role == "admin":
        rows = db.execute(
            "SELECT * FROM remote_clients ORDER BY created_at DESC, id DESC"
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT * FROM remote_clients WHERE owner_user_id = ? ORDER BY created_at DESC, id DESC",
            (user_id,),
        ).fetchall()
    return jsonify([serialize_remote_client(row) for row in rows])


@app.route("/api/remote/clients", methods=["POST"])
@login_required
@admin_required
def create_remote_client():
    data = request.get_json(silent=True) or {}
    client_name = str(data.get("client_name") or "").strip() or "默认设备"
    db = get_db()
    row, client_token = create_remote_client_record(
        db,
        owner_user_id=int(session["user_id"]),
        client_name=client_name,
    )
    return jsonify(
        {
            "item": serialize_remote_client(row),
            "client_token": client_token,
        }
    ), 201


@app.route("/api/remote/conversations", methods=["GET"])
@login_required
@admin_required
def list_remote_conversations():
    db = get_db()
    user_id = int(session["user_id"])
    role = str(session.get("role") or "user")
    client_id = str(request.args.get("client_id") or "").strip()
    params: list[Any] = []
    sql = (
        "SELECT c.*, rc.client_id AS public_client_id, rc.client_name AS client_name "
        "FROM remote_conversations c "
        "JOIN remote_clients rc ON rc.id = c.remote_client_id "
    )
    conditions: list[str] = []
    if role != "admin":
        conditions.append("c.owner_user_id = ?")
        params.append(user_id)
    if client_id:
        conditions.append("rc.client_id = ?")
        params.append(client_id)
    if conditions:
        sql += " WHERE " + " AND ".join(conditions)
    sql += " ORDER BY c.updated_at DESC, c.id DESC"
    rows = db.execute(sql, params).fetchall()
    items = []
    for row in rows:
        item = serialize_remote_conversation(row)
        item["client_id"] = row["public_client_id"]
        item["client_name"] = row["client_name"]
        items.append(item)
    return jsonify(items)


@app.route("/api/remote/conversations", methods=["POST"])
@login_required
@admin_required
def create_remote_conversation():
    data = request.get_json(silent=True) or {}
    client_id = str(data.get("client_id") or "").strip()
    if not client_id:
        return jsonify({"error": "client_id 不能为空"}), 400
    db = get_db()
    client_row = get_remote_client_by_public_id(db, client_id)
    if not client_row:
        return jsonify({"error": "未找到对应客户端"}), 404
    user_id = int(session["user_id"])
    role = str(session.get("role") or "user")
    if role != "admin" and int(client_row["owner_user_id"]) != user_id:
        return jsonify({"error": "权限不足"}), 403
    title = str(data.get("title") or "").strip() or f"{client_row['client_name']} 会话"
    now = now_iso()
    db.execute(
        """
        INSERT INTO remote_conversations (remote_client_id, owner_user_id, title, status, created_at, updated_at)
        VALUES (?, ?, ?, 'active', ?, ?)
        """,
        (client_row["id"], client_row["owner_user_id"], title, now, now),
    )
    db.commit()
    row = db.execute(
        "SELECT * FROM remote_conversations WHERE id = last_insert_rowid()"
    ).fetchone()
    return jsonify(serialize_remote_conversation(row)), 201


@app.route("/api/remote/conversations/<int:conversation_id>/messages", methods=["GET"])
@login_required
@admin_required
def list_remote_messages(conversation_id: int):
    db = get_db()
    user_id = int(session["user_id"])
    role = str(session.get("role") or "user")
    conversation_row = ensure_remote_conversation_access(db, conversation_id, user_id, role)
    if not conversation_row:
        return jsonify({"error": "未找到会话或权限不足"}), 404
    message_rows = db.execute(
        """
        SELECT *
        FROM remote_messages
        WHERE conversation_id = ?
        ORDER BY id ASC
        """,
        (conversation_id,),
    ).fetchall()
    items = []
    for row in message_rows:
        attachments = db.execute(
            "SELECT * FROM remote_attachments WHERE message_id = ? ORDER BY id ASC",
            (row["id"],),
        ).fetchall()
        items.append(serialize_remote_message(row, [serialize_remote_attachment(item) for item in attachments]))
    return jsonify(items)


@app.route("/api/remote/conversations/<int:conversation_id>/messages", methods=["POST"])
@login_required
@admin_required
def create_remote_message(conversation_id: int):
    db = get_db()
    user_id = int(session["user_id"])
    role = str(session.get("role") or "user")
    conversation_row = ensure_remote_conversation_access(db, conversation_id, user_id, role)
    if not conversation_row:
        return jsonify({"error": "未找到会话或权限不足"}), 404
    data = request.get_json(silent=True) or {}
    message_type = str(data.get("message_type") or "text").strip().lower()
    if message_type not in REMOTE_MESSAGE_TYPE_VALUES:
        return jsonify({"error": "message_type 不支持"}), 400
    normalized_payload, payload_error, content_text = sanitize_remote_command_payload(message_type, data)
    if payload_error:
        return jsonify({"error": payload_error}), 400
    payload_json = json.dumps(normalized_payload, ensure_ascii=False) if normalized_payload is not None else None
    status = "pending" if message_type == "command" else "success"
    now = now_iso()
    db.execute(
        """
        INSERT INTO remote_messages (
            conversation_id, sender_type, sender_user_id, remote_client_id, message_type, content_text, payload_json, status, created_at, updated_at
        ) VALUES (?, 'user', ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            conversation_id,
            user_id,
            conversation_row["remote_client_id"],
            message_type,
            content_text or None,
            payload_json,
            status,
            now,
            now,
        ),
    )
    db.execute(
        "UPDATE remote_conversations SET updated_at = ? WHERE id = ?",
        (now, conversation_id),
    )
    db.commit()
    row = db.execute(
        "SELECT * FROM remote_messages WHERE id = last_insert_rowid()"
    ).fetchone()
    return jsonify(serialize_remote_message(row)), 201


@app.route("/client-api/remote/register", methods=["POST"])
def client_register_remote():
    db, client_row = require_remote_client()
    if client_row is None:
        return jsonify({"ok": False, "message": "client_id 或 client_token 无效"}), 401
    data = request.get_json(silent=True) or {}
    now = now_iso()
    db.execute(
        """
        UPDATE remote_clients
        SET machine_id = ?, device_name = ?, app_version = ?, workspace_path = ?, status = 'online', last_seen_at = ?, updated_at = ?
        WHERE id = ?
        """,
        (
            str(data.get("machine_id") or "").strip() or None,
            str(data.get("device_name") or "").strip() or None,
            str(data.get("app_version") or "").strip() or None,
            str(data.get("workspace_path") or "").strip() or None,
            now,
            now,
            client_row["id"],
        ),
    )
    db.commit()
    refreshed = get_remote_client_by_public_id(db, client_row["client_id"])
    return jsonify({"ok": True, "data": serialize_remote_client(refreshed)})


@app.route("/client-api/kuaishou/settings/status", methods=["GET"])
def client_get_kuaishou_settings_status():
    db, client_row = require_remote_client()
    if client_row is None:
        return jsonify({"ok": False, "message": "client_id ? client_token ??"}), 401
    requested_app_id = str(request.args.get("app_id") or "").strip()
    return jsonify({"ok": True, "data": serialize_kuaishou_settings(get_kuaishou_server_settings(requested_app_id))})


@app.route("/client-api/kuaishou/settings/ensure", methods=["PUT", "POST"])
def client_ensure_kuaishou_settings():
    db, client_row = require_remote_client()
    if client_row is None:
        return jsonify({"ok": False, "message": "client_id ? client_token ??"}), 401
    data = request.get_json(silent=True) or {}
    app_id = str(data.get("app_id") or "").strip()
    app_secret = str(data.get("app_secret") or "").strip()
    if not app_id:
        return jsonify({"ok": False, "message": "?? AppID ????"}), 400
    if len(app_id) > 64:
        return jsonify({"ok": False, "message": "?? AppID ??"}), 400

    existing_row = get_kuaishou_app_row(app_id)
    existing_settings = get_kuaishou_server_settings(app_id)
    existing_secret = str(existing_settings.get("app_secret") or "").strip()
    effective_secret = app_secret or existing_secret
    if not effective_secret:
        return jsonify({"ok": False, "message": "?? AppSecret ????"}), 400

    existing_access_token = str(existing_settings.get("access_token") or "").strip()
    existing_refresh_token = str(existing_settings.get("refresh_token") or "").strip()
    existing_access_expires_at = int(existing_settings.get("access_token_expires_at") or 0)
    existing_refresh_expires_at = int(existing_settings.get("refresh_token_expires_at") or 0)

    incoming_access_token = str(data.get("access_token") or "").strip()
    incoming_refresh_token = str(data.get("refresh_token") or "").strip()
    incoming_access_expires_at = int(data.get("access_token_expires_at") or 0)
    incoming_refresh_expires_at = int(data.get("refresh_token_expires_at") or 0)

    merged_access_token = existing_access_token
    merged_access_expires_at = existing_access_expires_at
    if incoming_access_token and (incoming_access_expires_at > existing_access_expires_at or not existing_access_token):
        merged_access_token = incoming_access_token
        merged_access_expires_at = incoming_access_expires_at

    merged_refresh_token = existing_refresh_token
    merged_refresh_expires_at = existing_refresh_expires_at
    if incoming_refresh_token and (incoming_refresh_expires_at > existing_refresh_expires_at or not existing_refresh_token):
        merged_refresh_token = incoming_refresh_token
        merged_refresh_expires_at = incoming_refresh_expires_at

    merged_advertiser_id = str(data.get("advertiser_id") or existing_settings.get("advertiser_id") or "").strip()
    merged_name = str(data.get("name") or existing_settings.get("name") or "").strip()
    owner_user_id = int(client_row["owner_user_id"] or 0) or None
    saved_row = save_kuaishou_app(
        app_id=app_id,
        app_secret=effective_secret,
        advertiser_id=merged_advertiser_id,
        name=merged_name,
        access_token=merged_access_token,
        refresh_token=merged_refresh_token,
        access_token_expires_at=merged_access_expires_at,
        refresh_token_expires_at=merged_refresh_expires_at,
        enabled=bool(existing_settings.get("enabled", True)) if existing_row else True,
        is_default=bool(existing_settings.get("is_default", False)) if existing_row else (not bool(get_default_kuaishou_app_row())),
        updated_by=owner_user_id,
    )
    db.commit()

    action = "updated"
    updated = True
    if existing_row is not None:
        token_changed = (
            merged_access_token != existing_access_token
            or merged_refresh_token != existing_refresh_token
            or merged_access_expires_at != existing_access_expires_at
            or merged_refresh_expires_at != existing_refresh_expires_at
            or merged_advertiser_id != str(existing_settings.get("advertiser_id") or "").strip()
            or effective_secret != existing_secret
        )
        if not token_changed:
            action = "unchanged"
            updated = False

    return jsonify({"ok": True, "data": {**serialize_kuaishou_app(saved_row), "action": action, "updated": updated}})


@app.route("/client-api/kuaishou/token", methods=["POST"])
def client_get_kuaishou_token():
    db, client_row = require_remote_client()
    if client_row is None:
        return jsonify({"ok": False, "message": "client_id 或 client_token 无效"}), 401
    data = request.get_json(silent=True) or {}
    requested_app_id = str(data.get("app_id") or request.args.get("app_id") or "").strip()
    payload, error_message, status_code = refresh_kuaishou_app_token_if_due(
        requested_app_id,
        refreshing_by=str(client_row["client_id"] or "client"),
        force=False,
    )
    if error_message:
        return jsonify({"ok": False, "message": error_message}), status_code
    return jsonify({"ok": True, "data": payload})


@app.route("/client-api/minidrama/settings/status", methods=["GET"])
def client_get_minidrama_settings_status():
    db, client_row = require_remote_client()
    if client_row is None:
        return jsonify({"ok": False, "message": "client_id 或 client_token 无效"}), 401
    requested_app_id = str(request.args.get("app_id") or "").strip()
    return jsonify({"ok": True, "data": serialize_minidrama_settings(get_minidrama_server_settings(requested_app_id))})


@app.route("/client-api/minidrama/settings/ensure", methods=["PUT", "POST"])
def client_ensure_minidrama_settings():
    db, client_row = require_remote_client()
    if client_row is None:
        return jsonify({"ok": False, "message": "client_id 或 client_token 无效"}), 401

    data = request.get_json(silent=True) or {}
    app_id = str(data.get("app_id") or "").strip()
    app_secret = str(data.get("app_secret") or "").strip()
    if not app_id:
        return jsonify({"ok": False, "message": "小程序 AppID 不能为空"}), 400
    if len(app_id) > 64:
        return jsonify({"ok": False, "message": "小程序 AppID 过长"}), 400
    if not app_secret:
        return jsonify({"ok": False, "message": "小程序 AppSecret 不能为空"}), 400
    if len(app_secret) > 256:
        return jsonify({"ok": False, "message": "小程序 AppSecret 过长"}), 400

    existing_settings = get_minidrama_server_settings(app_id)
    existing_app_id = str(existing_settings.get("app_id") or "").strip()
    existing_app_secret = str(existing_settings.get("app_secret") or "").strip()
    if existing_app_id and existing_app_secret:
        return jsonify(
            {
                "ok": True,
                "data": {
                    **serialize_minidrama_settings(existing_settings),
                    "action": "unchanged",
                    "updated": False,
                },
            }
        )

    owner_user_id = int(client_row["owner_user_id"] or 0) or None
    saved_row = save_minidrama_app(
        app_id=app_id,
        app_secret=app_secret,
        name=str(data.get("name") or "").strip(),
        enabled=True,
        is_default=not bool(get_default_minidrama_app_row()),
        updated_by=owner_user_id,
    )
    db.commit()

    return jsonify(
        {
            "ok": True,
            "data": {
                **serialize_minidrama_app(saved_row),
                "action": "updated",
                "updated": True,
            },
        }
    )


@app.route("/client-api/minidrama/token", methods=["POST"])
def client_get_minidrama_token():
    db, client_row = require_remote_client()
    if client_row is None:
        return jsonify({"ok": False, "message": "client_id 或 client_token 无效"}), 401
    data = request.get_json(silent=True) or {}
    requested_app_id = str(data.get("app_id") or request.args.get("app_id") or "").strip()
    app_id, app_secret, error_message = resolve_minidrama_server_credentials(requested_app_id)
    if error_message:
        return jsonify({"ok": False, "message": error_message}), 400

    now_ts = now_timestamp()
    now_text = now_iso()
    row = db.execute("SELECT * FROM minidrama_token_cache WHERE app_id = ?", (app_id,)).fetchone()
    if row and str(row["access_token"] or "").strip() and int(row["expires_at"] or 0) - now_ts > MINIDRAMA_TOKEN_REFRESH_MARGIN_SECONDS:
        return jsonify(
            {
                "ok": True,
                "data": {
                    "app_id": app_id,
                    "access_token": row["access_token"],
                    "expires_at": int(row["expires_at"] or 0),
                    "expires_in": max(0, int(row["expires_at"] or 0) - now_ts),
                    "cached": True,
                },
            }
        )

    db.execute(
        """
        INSERT OR IGNORE INTO minidrama_token_cache (
            app_id, access_token, expires_at, refreshing_by, refreshing_until, created_at, updated_at
        ) VALUES (?, NULL, 0, NULL, 0, ?, ?)
        """,
        (app_id, now_text, now_text),
    )
    db.commit()
    lock_until = now_ts + MINIDRAMA_TOKEN_REFRESH_LOCK_SECONDS
    cursor = db.execute(
        """
        UPDATE minidrama_token_cache
        SET refreshing_by = ?, refreshing_until = ?, updated_at = ?
        WHERE app_id = ? AND (refreshing_until IS NULL OR refreshing_until <= ? OR refreshing_by = ?)
        """,
        (client_row["client_id"], lock_until, now_text, app_id, now_ts, client_row["client_id"]),
    )
    db.commit()
    if cursor.rowcount <= 0:
        row = db.execute("SELECT * FROM minidrama_token_cache WHERE app_id = ?", (app_id,)).fetchone()
        if row and str(row["access_token"] or "").strip() and int(row["expires_at"] or 0) > now_ts + 60:
            return jsonify(
                {
                    "ok": True,
                    "data": {
                        "app_id": app_id,
                        "access_token": row["access_token"],
                        "expires_at": int(row["expires_at"] or 0),
                        "expires_in": max(0, int(row["expires_at"] or 0) - now_ts),
                        "cached": True,
                        "refreshing": True,
                    },
                }
            )
        return jsonify({"ok": False, "message": "小程序 token 正在刷新，请稍后重试"}), 409

    try:
        access_token, expires_in = fetch_minidrama_access_token_from_weixin(app_id, app_secret)
    except Exception as exc:
        db.execute(
            """
            UPDATE minidrama_token_cache
            SET refreshing_by = NULL, refreshing_until = 0, last_error = ?, updated_at = ?
            WHERE app_id = ?
            """,
            (str(exc), now_iso(), app_id),
        )
        db.commit()
        return jsonify({"ok": False, "message": str(exc)}), 502

    expires_at = now_timestamp() + max(0, int(expires_in) - 120)
    db.execute(
        """
        UPDATE minidrama_token_cache
        SET access_token = ?, expires_at = ?, refreshing_by = NULL, refreshing_until = 0, last_error = NULL, updated_at = ?
        WHERE app_id = ?
        """,
        (access_token, expires_at, now_iso(), app_id),
    )
    db.commit()
    return jsonify(
        {
            "ok": True,
            "data": {
                "app_id": app_id,
                "access_token": access_token,
                "expires_at": expires_at,
                "expires_in": max(0, expires_at - now_timestamp()),
                "cached": False,
            },
        }
    )


@app.route("/client-api/remote/poll", methods=["GET"])
def client_poll_remote():
    db, client_row = require_remote_client()
    if client_row is None:
        return jsonify({"ok": False, "message": "client_id 或 client_token 无效"}), 401
    now = now_iso()
    db.execute(
        "UPDATE remote_clients SET status = 'online', last_seen_at = ?, updated_at = ? WHERE id = ?",
        (now, now, client_row["id"]),
    )
    db.commit()
    row = db.execute(
        """
        SELECT m.*
        FROM remote_messages m
        JOIN remote_conversations c ON c.id = m.conversation_id
        WHERE c.remote_client_id = ? AND m.message_type = 'command' AND m.status = 'pending'
        ORDER BY m.id ASC
        LIMIT 1
        """,
        (client_row["id"],),
    ).fetchone()
    if not row:
        return jsonify({"ok": True, "data": None})
    db.execute(
        "UPDATE remote_messages SET status = 'sent', updated_at = ? WHERE id = ?",
        (now, row["id"]),
    )
    db.execute(
        "UPDATE remote_conversations SET updated_at = ? WHERE id = ?",
        (now, row["conversation_id"]),
    )
    db.commit()
    attachments = db.execute(
        "SELECT * FROM remote_attachments WHERE message_id = ? ORDER BY id ASC",
        (row["id"],),
    ).fetchall()
    refreshed = db.execute("SELECT * FROM remote_messages WHERE id = ?", (row["id"],)).fetchone()
    return jsonify({"ok": True, "data": serialize_remote_message(refreshed, [serialize_remote_attachment(item) for item in attachments])})


@app.route("/client-api/remote/messages/<int:message_id>/complete", methods=["POST"])
def client_complete_remote_message(message_id: int):
    db, client_row = require_remote_client()
    if client_row is None:
        return jsonify({"ok": False, "message": "client_id 或 client_token 无效"}), 401
    data = request.get_json(silent=True) or {}
    status = str(data.get("status") or "success").strip().lower()
    if status not in REMOTE_MESSAGE_STATUS_VALUES:
        status = "success"
    result_json = json.dumps(data.get("result") or {}, ensure_ascii=False)
    row = db.execute(
        """
        SELECT m.*, c.remote_client_id
        FROM remote_messages m
        JOIN remote_conversations c ON c.id = m.conversation_id
        WHERE m.id = ?
        """,
        (message_id,),
    ).fetchone()
    if not row or int(row["remote_client_id"]) != int(client_row["id"]):
        return jsonify({"ok": False, "message": "未找到消息"}), 404
    now = now_iso()
    db.execute(
        "UPDATE remote_messages SET status = ?, result_json = ?, updated_at = ? WHERE id = ?",
        (status, result_json, now, message_id),
    )
    db.execute(
        "UPDATE remote_conversations SET updated_at = ? WHERE id = ?",
        (now, row["conversation_id"]),
    )
    db.commit()
    return jsonify({"ok": True})


@app.route("/client-api/remote/upload-image", methods=["POST"])
def client_upload_remote_image():
    db, client_row = require_remote_client()
    if client_row is None:
        return jsonify({"ok": False, "message": "client_id 或 client_token 无效"}), 401
    upload = request.files.get("file")
    if upload is None or not upload.filename:
        return jsonify({"ok": False, "message": "缺少图片文件"}), 400
    message_text = str(request.form.get("message") or "").strip()
    conversation = get_or_create_remote_conversation(db, client_row, title=f"{client_row['client_name']} 远程会话")
    now = now_iso()
    db.execute(
        """
        INSERT INTO remote_messages (
            conversation_id, sender_type, sender_user_id, remote_client_id, message_type, content_text, payload_json, status, created_at, updated_at
        ) VALUES (?, 'client', NULL, ?, 'image', ?, NULL, 'success', ?, ?)
        """,
        (conversation["id"], client_row["id"], message_text or None, now, now),
    )
    message_row = db.execute("SELECT * FROM remote_messages WHERE id = last_insert_rowid()").fetchone()
    ext = os.path.splitext(upload.filename or "")[1] or ".png"
    stored_name = f"{message_row['id']}_{secrets.token_hex(4)}{ext}"
    stored_path = os.path.join(REMOTE_UPLOAD_DIR, stored_name)
    upload.save(stored_path)
    db.execute(
        """
        INSERT INTO remote_attachments (message_id, file_type, original_name, stored_path, content_type)
        VALUES (?, 'image', ?, ?, ?)
        """,
        (message_row["id"], upload.filename, stored_path, upload.mimetype or "image/png"),
    )
    db.execute(
        "UPDATE remote_conversations SET updated_at = ? WHERE id = ?",
        (now, conversation["id"]),
    )
    db.commit()
    attachment_row = db.execute("SELECT * FROM remote_attachments WHERE id = last_insert_rowid()").fetchone()
    return jsonify(
        {
            "ok": True,
            "data": {
                "message": serialize_remote_message(message_row, [serialize_remote_attachment(attachment_row)]),
            },
        }
    )


@app.before_request
def _ensure_background_schedulers_started():
    start_kuaishou_token_refresh_scheduler()


@app.route("/api/remote/attachments/<int:attachment_id>", methods=["GET"])
@login_required
@admin_required
def download_remote_attachment(attachment_id: int):
    db = get_db()
    user_id = int(session["user_id"])
    role = str(session.get("role") or "user")
    row = db.execute(
        """
        SELECT a.*, c.owner_user_id
        FROM remote_attachments a
        JOIN remote_messages m ON m.id = a.message_id
        JOIN remote_conversations c ON c.id = m.conversation_id
        WHERE a.id = ?
        """,
        (attachment_id,),
    ).fetchone()
    if not row:
        return jsonify({"error": "未找到附件"}), 404
    if role != "admin" and int(row["owner_user_id"]) != user_id:
        return jsonify({"error": "权限不足"}), 403
    return send_file(
        row["stored_path"],
        mimetype=row["content_type"] or "application/octet-stream",
        as_attachment=False,
        download_name=row["original_name"] or os.path.basename(row["stored_path"]),
    )


@app.route("/api/profile/password", methods=["POST"])
@login_required
def update_profile_password():
    data = request.get_json(silent=True) or {}
    current_password = (data.get("current_password") or "").strip()
    new_password = (data.get("new_password") or "").strip()
    if not current_password:
        return jsonify({"error": "当前密码不能为空"}), 400
    if len(new_password) < 6:
        return jsonify({"error": "新密码至少6位"}), 400
    user_id = session.get("user_id")
    db = get_db()
    user = db.execute(
        "SELECT id, password_hash FROM users WHERE id = ?",
        (user_id,),
    ).fetchone()
    if not user or not check_password_hash(user["password_hash"], current_password):
        return jsonify({"error": "当前密码错误"}), 400
    db.execute(
        "UPDATE users SET password_hash = ? WHERE id = ?",
        (generate_password_hash(new_password), user["id"]),
    )
    db.commit()
    return jsonify({"message": "密码修改成功"})


def build_filter_clause(args):
    clauses: list[str] = []
    params: list[object] = []
    drama_id = (args.get("id") or "").strip()
    if drama_id.isdigit():
        clauses.append("id = ?")
        params.append(int(drama_id))
    search = (args.get("search") or "").strip()
    if search:
        like = f"%{search}%"
        clauses.append("(original_name LIKE ? OR new_name LIKE ?)")
        params.extend([like, like])
    company = (args.get("company") or "").strip()
    if company:
        clauses.append("company = ?")
        params.append(company)
    review_passed = (args.get("review_passed") or "").strip()
    if review_passed in ALLOWED_FLAGS:
        clauses.append("review_passed = ?")
        params.append(review_passed)
    uploaded = (args.get("uploaded") or "").strip()
    if uploaded in ALLOWED_FLAGS:
        clauses.append("uploaded = ?")
        params.append(uploaded)
    uploader = (args.get("uploader") or "").strip()
    if uploader:
        clauses.append("uploader LIKE ?")
        params.append(f"%{uploader}%")
    date_from = (args.get("date_from") or "").strip()
    if date_from:
        clauses.append("date >= ?")
        params.append(date_from)
    date_to = (args.get("date_to") or "").strip()
    if date_to:
        clauses.append("date <= ?")
        params.append(date_to)
    return clauses, params


def parse_iso_date(value: str | None) -> datetime.date | None:
    if not value:
        return None
    try:
        return datetime.date.fromisoformat(value)
    except ValueError:
        return None


def build_monitor_filters(args) -> tuple[list[str], list[object], str, datetime.date, datetime.date]:
    clauses: list[str] = []
    params: list[object] = []

    company = (args.get("company") or "").strip()
    if company:
        clauses.append("company = ?")
        params.append(company)

    mode = (args.get("mode") or "created").strip().lower()
    if mode not in {"created", "online"}:
        mode = "created"

    today = datetime.date.today()
    default_from = today - datetime.timedelta(days=29)
    date_from = parse_iso_date((args.get("date_from") or "").strip()) or default_from
    date_to = parse_iso_date((args.get("date_to") or "").strip()) or today
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    return clauses, params, mode, date_from, date_to


def parse_iso_date(value: str | None) -> datetime.date | None:
    if not value:
        return None
    try:
        return datetime.date.fromisoformat(value)
    except ValueError:
        return None


def build_monitor_filters(args) -> tuple[list[str], list[object], str, datetime.date, datetime.date]:
    clauses: list[str] = []
    params: list[object] = []

    company = (args.get("company") or "").strip()
    if company:
        clauses.append("company = ?")
        params.append(company)

    mode = (args.get("mode") or "created").strip().lower()
    if mode not in {"created", "online"}:
        mode = "created"

    today = datetime.date.today()
    default_from = today - datetime.timedelta(days=29)
    date_from = parse_iso_date((args.get("date_from") or "").strip()) or default_from
    date_to = parse_iso_date((args.get("date_to") or "").strip()) or today
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    return clauses, params, mode, date_from, date_to


def row_to_dict(row: sqlite3.Row) -> dict:
    return {key: row[key] for key in row.keys()}


def sanitize_drama_payload(data: dict) -> tuple[dict, str | None]:
    payload: dict[str, object | None] = {}
    original_name = (data.get("original_name") or "").strip()
    new_name = (data.get("new_name") or "").strip()
    if not original_name or not new_name:
        return payload, "原剧名和新剧名不能为空"

    payload["original_name"] = original_name
    payload["new_name"] = new_name

    date_value = data.get("date")
    if isinstance(date_value, (datetime.date, datetime.datetime)):
        payload["date"] = date_value.strftime("%Y-%m-%d")
    else:
        payload["date"] = (date_value or "").strip() or None

    payload["episodes"] = to_int_or_none(data.get("episodes"))
    payload["duration"] = to_int_or_none(data.get("duration"))
    payload["review_passed"] = normalize_flag(data.get("review_passed"))
    payload["uploaded"] = normalize_flag(data.get("uploaded"))
    payload["uploader"] = (data.get("uploader") or "").strip() or None
    payload["materials"] = (data.get("materials") or "").strip() or None
    payload["promo_text"] = (data.get("promo_text") or "").strip() or None
    payload["description"] = (data.get("description") or "").strip() or None
    payload["company"] = (data.get("company") or "").strip() or None
    payload["remark1"] = (data.get("remark1") or "").strip() or None
    payload["remark2"] = (data.get("remark2") or "").strip() or None
    payload["remark3"] = (data.get("remark3") or "").strip() or None
    for key in ("uploader", "remark1", "remark2", "remark3"):
        if payload[key] and len(payload[key]) > 200:
            payload[key] = payload[key][:200]
    return payload, None


def normalize_row(row_data: dict) -> dict:
    normalized = {}
    date_value = row_data.get("date")
    if isinstance(date_value, datetime.datetime):
        normalized["date"] = date_value.strftime("%Y-%m-%d")
    elif isinstance(date_value, datetime.date):
        normalized["date"] = date_value.strftime("%Y-%m-%d")
    else:
        normalized["date"] = (str(date_value).strip() if date_value else None)
    normalized["original_name"] = (row_data.get("original_name") or "").strip()
    normalized["new_name"] = (row_data.get("new_name") or "").strip()
    normalized["episodes"] = to_int_or_none(row_data.get("episodes"))
    normalized["duration"] = to_int_or_none(row_data.get("duration"))
    normalized["review_passed"] = normalize_flag(row_data.get("review_passed"))
    normalized["uploaded"] = normalize_flag(row_data.get("uploaded"))
    normalized["uploader"] = normalize_text(row_data.get("uploader"))
    normalized["materials"] = normalize_text(row_data.get("materials"))
    normalized["promo_text"] = normalize_text(row_data.get("promo_text"))
    normalized["description"] = normalize_text(row_data.get("description"))
    normalized["company"] = normalize_text(row_data.get("company"))
    normalized["remark1"] = normalize_text(row_data.get("remark1"))
    normalized["remark2"] = normalize_text(row_data.get("remark2"))
    normalized["remark3"] = normalize_text(row_data.get("remark3"))
    for key in ("remark1", "remark2", "remark3"):
        if normalized[key] and len(normalized[key]) > 200:
            normalized[key] = normalized[key][:200]
    return normalized


def normalize_text(value):
    if value is None:
        return None
    return str(value).strip() or None


def normalize_flag(value):
    if isinstance(value, str) and value.strip() == "是":
        return "是"
    return "否"


def to_int_or_none(value):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000, debug=True)
